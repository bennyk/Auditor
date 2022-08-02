
from openpyxl import load_workbook, Workbook
from typing import List
import re

max_row = max_col = 99


def colnum_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string


class Table:
    col_limit = 0
    date_range = []

    def __init__(self, sheet_ranges):
        # configure spreadsheet based on number of cols.
        for j in range(2, max_col):
            c0 = "{}{}".format(colnum_string(j), 1)
            Table.date_range.append(sheet_ranges[c0].value)
            if re.match(r'CAGR$', sheet_ranges[c0].value):
                Table.col_limit = j+1
                break

        self.tab = []
        for i in range(1, max_row):
            c0 = "{}{}".format(colnum_string(1), i)
            if sheet_ranges[c0].value is None:
                break
            r = []
            for j in range(1, Table.col_limit):
                c1 = "{}{}".format(colnum_string(j), i)
                r.append(sheet_ranges[c1].value)
            self.tab.append(r)

    def match_title(self, reg):
        result = None
        for _ in self.tab:
            if re.match(reg, _[0]):
                result = _
                break
        assert result is not None
        return result

    def share_out_filing(self) -> float:
        # TODO based on the current or future one year later when data is available
        x = self.match_title('Total Shares Out\.')
        result = list(filter(None, reversed(x[1:])))[0]
        return result

    @staticmethod
    def strip(l):
        return l[1:][:-3]

    def list_over_share(self, l):
        return list(map(lambda e: e / self.share_out_filing(), Table.strip(l)) )

    def list_over_list(self, x, y, percent=False):
        if percent:
            return list(map(lambda n1, n2: 100 * (n1 / n2), x, y))
        return list(map(lambda n1, n2: n1 / n2, x, y))

    @staticmethod
    def cagr(l: [float]) -> float:
        return (l[len(l)-1] / l[0])**(1/len(l))-1

    @staticmethod
    def average(l: [float]):
        l = Table.strip(l)
        return sum(l)/len(l)

    def compute_ebit(self):
        t = self.match_title('EBIT$')
        # including dilution of share in percent
        ebit_over_share_per = Table.cagr(self.list_over_share(t))
        print("Sampled {} years of EBIT/share for {:.2f}% CAGR growth rate, including 3 year of analysts proj"
              .format(len(Table.strip(t)), 100*ebit_over_share_per))
        if abs(Table.cagr(Table.strip(t)) - ebit_over_share_per) > .01:
            ebit_per = Table.cagr(Table.strip(t))
            print("EBIT {:.2f} in CAGR percent vs {:.2f} on per share basis".format(ebit_per*100, ebit_over_share_per*100))

    def revenue(self):
        revs = Table.strip(self.match_title('Revenue$'))
        rev_per_share = Table.cagr(list(map(lambda r: r / self.share_out_filing(), revs)))
        print("Revenue per share from {} to {} for {} years at CAGR {:.2f}% for {}".format(
            revs[0], revs[-1], len(revs),
            rev_per_share*100, revs))
        if abs(Table.cagr(revs) - rev_per_share) > .01:
            print("Revenue {:.2f} in percent vs {:.2f} on per share basis".format(
                Table.cagr(revs)*100, rev_per_share*100))

    def eps(self):
        eps = Table.strip(self.match_title('EPS \(GAAP\)$'))
        print("EPS from {} to {} for {} years for {}".format(eps[0], eps[-1], len(eps), eps))

    def fcf(self):
        fcf = Table.strip(self.match_title('Free Cash Flow'))
        fcf_per_share = Table.cagr(list(map(lambda f: f / self.share_out_filing(), fcf)))
        print("FCF per share from {} to {} at CAGR {:.2f}% for {}".format(
            fcf[0], fcf[-1],
            fcf_per_share * 100, fcf))
        if abs(Table.cagr(fcf) - fcf_per_share) > .01:
            print("FCF {:.2f} in percent vs {:.2f} on per share basis".format(
                Table.cagr(fcf) * 100, fcf_per_share * 100))

    def book_value(self):
        bv = Table.strip(self.match_title('Book Value / Share$'))
        bv_cagr = Table.cagr(bv)
        print("Book value per share from {} to {} at CAGR {:.2f}% for {}".format(
            bv[0], bv[-1],
            bv_cagr*100, bv))

    def return_equity(self):
        roe = Table.strip(self.match_title('Return on Equity'))
        roe_cagr = Table.cagr(roe)
        print("Return on equity from {} to {} at CAGR {:.2f}% for {}".format(
            roe[0], roe[-1],
            roe_cagr * 100, roe))

    def net_debt_over_ebit(self):
        net_debt = Table.strip(self.match_title('Net Debt'))
        ebit = Table.strip(self.match_title('EBIT$'))
        # ebitda = self.match_title('EBITDA$')
        print("Net debt over EBIT average {:.2f} years for {}".format(
            Table.average(self.list_over_list(net_debt, ebit)),
            list(map(lambda x: round(x, 2), self.list_over_list(net_debt, ebit)))
        ))

    def ev_over_ebit(self):
        # TODO need to obtain market cap = price * shares outstanding
        pass

    def ebit_margin(self):
        ebits = Table.strip(self.match_title('EBIT$'))
        revs = Table.strip(self.match_title('Revenue$'))
        ebit_margins = self.list_over_list(ebits, revs, percent=True)
        print("EBIT margin average {:.2f}% for (numbers in percent) {}".format(
            Table.average(ebit_margins),
            list(map(lambda x: round(x, 2), ebit_margins))
        ))


wb = load_workbook('malaysia reit returns.xlsx')
for index, name in enumerate(wb.sheetnames):
    if index < 4:
        continue

    print("XXX", index, name)
    # type: List
    # sheet_ranges = wb[name]
    _t = Table(wb[name])
    # _t.compute_ebit()
    _t.revenue()
    _t.eps()
    _t.fcf()
    _t.book_value()
    _t.return_equity()
    _t.net_debt_over_ebit()
    _t.ebit_margin()
    # if index == 1:
    #     break

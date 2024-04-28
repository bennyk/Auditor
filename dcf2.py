import re
import pandas as pd

from spread import Spread
from utils import *
from bcolors import colour_print, bcolors

from openpyxl import Workbook, load_workbook, worksheet
from openpyxl.styles import Font, Alignment
from collections import OrderedDict
from typing import List
import yfinance as yf
from tabulate import tabulate

total_main_col = 12
total_half_col = int(total_main_col/2)

class ExcelOut:
    start_col = 2
    row_margin = 1

    def __init__(self, tick: str, entries, styles=None, headers=None):
        cls = self.__class__
        self.od = entries
        self.headers = headers
        self.tick = tick
        self.styles = styles
        self.ft = Font(name='Calibri', size=11)
        self.wb = Workbook()

        # type: worksheet.Worksheet
        self.sheet = self.wb.active
        self.sheet.title = 'sheet 1'

        self.cell = self.sheet.cell(row=1, column=cls.start_col)
        self.start_row_index = cls.row_margin+1
        self.end_row_index = len(self.tick) + self.start_row_index+1

        self.j = cls.row_margin + 1
        self.i = 1

        self.init_sheet()

    def init_sheet(self):
        sheet = self.sheet
        self.j = 1
        i = 2
        for val in self.headers:
            sheet.column_dimensions[colnum_string(self.j)].width = 32
            cell = sheet.cell(row=i, column=self.j)
            cell.alignment = Alignment(wrapText=True)
            if re.match(r'empty', val):
                pass
            else:
                cell.value = val
            i += 1

        cell = sheet.cell(row=1, column=1)
        cell.value = self.tick.upper()
        sheet.cell(row=1, column=2).value = 'Base year'
        for i in range(1, total_main_col-1):
            cell = sheet.cell(row=1, column=i+2)
            cell.value = i
            cell.alignment = Alignment(horizontal='center')
        sheet.cell(row=1, column=total_main_col+1).value = 'Terminal year'

        self.j += 1

    def start(self):
        cls = self.__class__
        self.j = cls.row_margin + 1
        for i, key in enumerate(self.od):
            # Table of mainly profile and last_price data
            self.i = 2
            if type(self.od[key]) is list:
                for val in self.od[key]:
                    if key != '':
                        self.make_cell(val, self.styles[i])
            else:
                assert type(self.od[key]) is float or type(self.od[key]) is int
                val = self.od[key]
                self.make_cell(val, self.styles[i])

            self.j += 1
        self.wb.save('out.xlsx')

    def make_cell(self, val, style):
        cell = self.sheet.cell(row=self.j, column=self.i)
        self.sheet.column_dimensions[colnum_string(self.i)].width = 11
        cell.alignment = Alignment(wrapText=True)
        if (type(val) is int or type(val) is float) and val == 0:
            # Suppress zero value to empty string.
            cell.value = ''
        else:
            cell.value = val
        if style == 'Comma':
            if val != 0:
                cell.style = style
                cell.number_format = '#,0.00'
        elif style == 'Percent':
            cell.style = style
            cell.number_format = '0.00%'
        elif style == 'Ratio2':
            cell.number_format = '0.00'
        elif style == 'Ratio':
            # cell.style = style
            cell.number_format = '0.0000'
        else:
            assert False
        cell.font = self.ft
        self.i += 1


class DataSet:
    def __init__(self, country, industry):
        self.path = "datacurrent"
        self.country = country
        self.industry = industry

    def get_country_tax_rates(self):
        name = 'countrytaxrates'
        result = None
        wb = load_workbook(self.path + '/{}.xlsx'.format(name))
        ws = wb[name]
        for i in range(1, ws.max_row):
            if ws['A{}'.format(i)].value == self.country:
                result = ws.cell(row=i, column=ws.max_column).value
                break
        assert result is not None
        return result

    def get_riskfree_rate(self):
        # TODO Country 10 years GBY on 13-Apr-2024 at the following websites: -
        # https://tradingeconomics.com/united-states/government-bond-yield
        # https://tradingeconomics.com/malaysia/government-bond-yield
        # https://tradingeconomics.com/china/government-bond-yield
        # https://tradingeconomics.com/hong-kong/government-bond-yield
        # https://tradingeconomics.com/taiwan/government-bond-yield
        tab = {
            'Malaysia': .03947,
            # 'United States': .0408,
            'United States': .04532,
            'Taiwan': .01565,
            'China': .02291,
            'Hong Kong': .0388, }
        assert self.country in tab
        return tab[self.country]

    def get_wacc(self):
        name = 'wacc'
        result = None
        wb = load_workbook(self.path + '/{}.xlsx'.format(name))
        ws = wb['Industry Averages']
        for i in range(20, ws.max_row):
            print(ws['A{}'.format(i)].value)
            if re.match(self.industry, ws['A{}'.format(i)].value, re.IGNORECASE):
                result = ws.cell(row=i, column=ws.max_column).value
                break
        assert result is not None
        return result

    def get_equity_risk_premium(self):
        result = None
        wb = load_workbook(self.path + '/{}.xlsx'.format('ERPs by country'))
        ws = wb['Sheet1']
        for i in range(8, ws.max_row):
            if re.match(self.country, ws['A{}'.format(i)].value):
                result = ws.cell(row=i, column=5).value
                break
        assert result is not None
        return result

    def get_currency_suffix(self, sticky_price):
        assert self.country is not None
        result = None
        with open(self.path + '/yahoo_currency_suffix.txt', encoding='utf-8') as infile:
            # https://www.gnucash.org/docs/v4/C/gnucash-help/fq-spec-yahoo.html
            suffix_index = None
            for i, field in enumerate(infile.readline().split('|')):
                if re.match('suffix', field, re.IGNORECASE):
                    suffix_index = i
                    break
            assert suffix_index is not None
            for line in infile:
                a = line[:-1].split('|')
                if re.match(a[suffix_index][1:], sticky_price):
                    result = a[suffix_index][1:]
                    break
        return result

    def get_sales_to_cap_ratio(self):
        name = 'capex'
        sales_to_cap_index = None
        result = None
        wb = load_workbook(self.path + '/{}.xlsx'.format(name))
        ws = wb['Industry Averages']
        for i in range(1, ws.max_column+1):
            a = ws.cell(row=8, column=i).value
            if re.match(r'Sales/ Invested Capital', a):
                sales_to_cap_index = i
                break
        assert sales_to_cap_index is not None
        for i in range(8, ws.max_row):
            if re.match(self.industry, ws['A{}'.format(i)].value, re.IGNORECASE):
                result = ws.cell(row=i, column=sales_to_cap_index).value
                break
        assert result is not None
        return result

    def match_sales_to_cap_ratio(self, sales_to_cap_ratio):
        name = 'capex'
        sales_to_cap_index = None
        result = None
        wb = load_workbook(self.path + '/{}.xlsx'.format(name))
        ws = wb['Industry Averages']
        for i in range(1, ws.max_column + 1):
            a = ws.cell(row=8, column=i).value
            if re.match(r'Sales/ Invested Capital', a):
                sales_to_cap_index = i
                break
        assert sales_to_cap_index is not None
        result = []
        for i in range(9, ws.max_row):
            diff = ws.cell(row=i, column=sales_to_cap_index).value - sales_to_cap_ratio
            # print(ws.cell(row=i, column=1).value, diff)
            result.append([
                ws.cell(row=i, column=1).value,
                ws.cell(row=i, column=sales_to_cap_index).value, abs(diff), ])
        return sorted(result, key=lambda e: e[2])


class DCF(Spread):
    def __init__(self, tick, country=None, industry=None, path=None):
        colour_print("Company's ticker '{}'".format(tick), bcolors.UNDERLINE)

        if country is None:
            country = 'United States'
            colour_print("Defaulting country to '{}'?".format(country), bcolors.WARNING)

        if industry is None:
            industry = 'semiconductor'
            colour_print("Defaulting industry to '{}'?".format(industry), bcolors.WARNING)

        if path is None:
            path = 'spreads'
            print("Set path to '{}'".format(path))

        self.wb = load_workbook(path + '/' + tick + '.xlsx')
        super().__init__(self.wb, tick)

        self.dataset = DataSet(country, industry)

        # Revenues, Operating Income, Interest Expense, ...

        self.sales = self.strip(self.income.match_title('Total Revenues'))
        self.forward_sales = self.trim_estimates('Revenue', nlead=8)
        self.forward_ebit = self.trim_estimates('EBIT$')

        # TODO Interest expense and Equity?
        self.ie = self.strip(self.income.match_title('Interest Expense'))
        self.book_value_equity = self.strip(self.balance.match_title('Total Equity'))
        self.book_value_debt = self.strip(self.balance.match_title('Total Debt'))
        self.current_debt = [0.]
        current_debt_not_strip = self.balance.match_title('Current Portion of Long-Term Debt', none_is_optional=True)
        if current_debt_not_strip is not None:
           self.current_debt = self.strip(current_debt_not_strip)
        self.debt = self.strip(self.balance.match_title('Total Debt'))

        cash_not_strip = self.balance.match_title('Total Cash', none_is_optional=True)
        if cash_not_strip is not None:
            self.cash = self.strip(cash_not_strip)
        else:
            self.cash = self.strip(self.balance.match_title('Cash And Equivalents'))
            assert self.cash is not None

        investment_not_strip = self.balance.match_title('Long-term Investments', none_is_optional=True)
        if investment_not_strip is not None:
            self.investments = self.strip(investment_not_strip)
        else:
            self.investments = 0

        minority_not_strip = self.income.match_title('Minority Interest', none_is_optional=True)
        if minority_not_strip is not None:
            self.minority = self.strip(minority_not_strip)
        else:
            self.minority = 0

        self.shares = self.strip(self.income.match_title('Weighted Average Diluted Shares Outstanding'))
        forward_etr_not_strip = self.trim_estimates('Effective Tax Rate', none_is_optional=True)
        if forward_etr_not_strip is not None:
            self.forward_etr = forward_etr_not_strip
        else:
            self.forward_etr = 0

        self.marginal_tax_rate = self.dataset.get_country_tax_rates()

        # Country 10 years GBY
        self.riskfree_rate = self.dataset.get_riskfree_rate()

        # Sea of change by Howard Marks
        # https://www.oaktreecapital.com/insights/memo/sea-change
        # self.riskfree_rate = .036

    def trim_estimates(self, title, nlead=8, n=4, **args):
        # Remove past annual/quarterly data from Estimates.
        est = self.estimates.match_title(title, **args)
        result = None
        if est is not None:
            excess = next((i for i in range(1, n) if est[nlead:][-i] is not None), 0)
            if excess-1 > 0:
                result = est[nlead:][:-excess+1]
            else:
                result = est[nlead:]
        return result

    def compute(self):
        d = OrderedDict()
        self.compute_revenue(d)
        self.compute_ebit(d)
        self.compute_tax(d)
        self.compute_ebt(d)
        self.compute_reinvestment(d)
        self.compute_fcff(d)
        d['empty1'] = []
        self.compute_cost_of_capital(d)
        self.compute_cumulative_df(d)
        self.compute_terminals(d)
        self.compute_return_invested_capital(d)

        headers = list(d.keys())
        excel = ExcelOut(
            self.tick, d, headers=headers,
            styles=[
                # Revenue growth rate demarcation
                'Percent', 'Comma', 'Percent', 'Comma', 'Percent', 'Comma', 'Comma', 'Comma',
                # Cost of capital demarcation
                '', 'Percent', 'Ratio', 'Comma',
                # Terminal cash flow demarcation
                '', 'Comma', 'Percent', 'Comma', 'Comma', 'Comma', 'Comma', 'Comma', 'Comma', 'Comma', 'Comma', 'Comma', 'Comma', 'Comma', 'Ratio2', 'Ratio2', 'Percent',
                # Sales to cap 
                '', 'Comma', 'Percent',
            ])
        excel.start()

    def compute_revenue(self, d):
        # Compute past
        # Compute forward forecast

        sales_growth_rate = d['Revenue growth rate'] = []
        sales = d['Revenue'] = []
        forward_sales = self.forward_sales[-4:]
        for i in range(len(forward_sales)):
            grow_rate = (forward_sales[i] - forward_sales[i-1]) / forward_sales[i-1]
            sales_growth_rate.append(grow_rate)
            sales.append(forward_sales[i])

        # Stable at year 2, 3 (optional here after), 4 and 5
        stable_growth_rate = sales_growth_rate[-1]
        cur_sales = forward_sales[-1] * (1+stable_growth_rate)
        for i in range(len(forward_sales)-1, 5):
            sales_growth_rate.append(stable_growth_rate)
            sales.append(cur_sales)
            if i < 4:
                # Otherwise, ignore the last loop
                cur_sales = cur_sales * (1+stable_growth_rate)

        # Terminal year period is based on current risk free rate based on 10 years treasury bond note yield
        term_year_per = self.riskfree_rate

        # Iterating from first year to terminal year in descending grow order, including terminal year
        for n in range(1, total_half_col):
            per = stable_growth_rate - (stable_growth_rate-term_year_per)/5 * n
            sales_growth_rate.append(per)
            cur_sales = cur_sales*(1+per)
            sales.append(cur_sales)

        # Terminal period, no grow and stagnated value
        per = stable_growth_rate - (stable_growth_rate-term_year_per)/5 * 5
        sales_growth_rate.append(per)
        cur_sales = cur_sales * (1 + per)
        sales.append(cur_sales)

    def compute_ebit(self, d):
        sales = d['Revenue']
        ebit_margin = d['EBIT margin'] = []
        ebit = d['EBIT'] = []

        for i, e in enumerate(self.forward_ebit):
            ebit_margin.append(e / sales[i])
            ebit.append(e)

        fixed_index = len(self.forward_ebit)-1
        fixed_margin = ebit_margin[-1]
        remaining_col = total_main_col - len(self.forward_ebit)
        for i in range(1, remaining_col):
            ebit_margin.append(fixed_margin)
            stable_ebit = fixed_margin * sales[fixed_index+i]
            ebit.append(stable_ebit)
        ebit_margin.append(fixed_margin)
        ebit.append(fixed_margin * sales[-1])

    def compute_tax(self, d):
        etr = d['Tax rate'] = []
        if self.forward_etr == 0:
            colour_print("Is \"{}\" a REIT company?"
                         " REIT company distributes at least 90% of its total yearly income to unit holders, the REIT itself is exempt from tax for that year, but unit holders are taxed on the distribution of income"
                         .format(self.tick), bcolors.WARNING)
            return

        assert type(self.forward_etr) is list
        for i, e in enumerate(self.forward_etr):
            if e is not None:
                etr.append(e/100)
            else:
                # TODO: Invalid ETR
                etr.append(0)

        # Previous year tax rate + (marginal tax rate - previous year tax rate) / 5
        start_tax_rate = tax_rate = etr[-1]
        for i in range(len(self.forward_etr), total_half_col):
            etr.append(tax_rate)

        for x in range(1, total_half_col):
            tax_rate = tax_rate+(self.marginal_tax_rate - start_tax_rate)/5
            etr.append(tax_rate)
        etr.append(tax_rate)

    def compute_ebt(self, d):
        ebit = d['EBIT']
        tax_rate = d['Tax rate']
        nopat = d['NOPAT'] = []
        for i, e in enumerate(self.forward_ebit):
            nopat.append(e)
            if len(tax_rate) > 0:
                nopat[-1] -= e * tax_rate[i]

        nopat_start = len(self.forward_ebit)
        for i in range(nopat_start, total_main_col):
            nopat.append(ebit[i])
            if len(tax_rate) > 0:
                nopat[-1] -= ebit[i] * tax_rate[i]

    def compute_reinvestment(self, d):
        # TODO Actual capex, D&A and changes in working capital
        # fwd_capex = self.strip(self.estimates.match_title('Capital Expenditure'))
        # fwd_dna = self.strip(self.estimates.match_title('Depreciation & Amortization'))
        # fwd change in working capital

        # Defaulting to no lag for the time being.
        # =IF(more_lag="No",(forward_2y_sales-forward_sales)/sales_to_cap,
        #   IF(years_lag=0,(forward_sales-prev_sales)/sales_to_cap,
        #   IF(years_lag=2,(forward_3y_sales-forward_2y_sales)/sales_to_cap,
        #   IF(years_lag=3,(forward_4y_sales-forward_3y_sales)/sales_to_cap,
        #   (forward_2y_sales-forward_sales)/sales_to_cap))))
        # https://pages.stern.nyu.edu/~adamodar/New_Home_Page/datafile/capex.html

        # Latest sales to book value of equity
        sales_to_cap_source = self.sales[-1] / self.book_value_equity[-1]
        print("Computed Sales to cap ratio {:.2f}".format(sales_to_cap_source))
        # TODO Asia countries not in U.S. coverage
        if True:
            print("Probable Sales to cap ratio:")
            matches = self.dataset.match_sales_to_cap_ratio(sales_to_cap_source)[:5]
            heads = ['Company', 'Sales to Cap', 'Error']
            print(tabulate(matches, headers=heads, floatfmt=".2f"), "\n")
            # Selecting mid of the 5 matches
            sales_to_cap_ratio = matches[2][1]
        else:
            sales_to_cap_ratio = sales_to_cap_source

        reinvestment = d['- Reinvestment'] = []
        reinvestment.append(None)
        sales = d['Revenue']
        for i in range(1, len(sales)-1):
            r = (sales[i+1]-sales[i]) / sales_to_cap_ratio
            reinvestment.append(r)

        # Terminal growth rate / End of ROIC * End of NOPAT
        term_growth_rate = d['Revenue growth rate'][-1]
        # TODO Cost of capital at year 10 or enter manually
        roic = .15
        nopat_end = d['NOPAT'][-1]
        reinvestment.append(term_growth_rate / roic * nopat_end)

        # Alternatively Sales to IC ratio = FCFF (computed by TIKR) + NOPAT

    def compute_fcff(self, d):
        fcff = d['FCFF'] = []
        nopat = d['NOPAT']
        reinvestment = d['- Reinvestment']
        for i in range(len(nopat)):
            if reinvestment[i] is not None:
                fcff.append(nopat[i]-reinvestment[i])
            else:
                fcff.append(None)

    def compute_cost_of_capital(self, d):
        # Cost of debt
        interest_expense = 0
        if self.ie[-1] is not None:
            interest_expense = self.ie[-1]

        debt = 0
        if self.debt[-1] is not None:
            debt = self.debt[-1]
        # I have excluded tax rate leading to lower debt number.
        pretax_cost_of_debt = 0
        if debt > 0:
            pretax_cost_of_debt = abs(interest_expense / debt)
        if self.forward_etr == 0:
            cost_of_debt = pretax_cost_of_debt
        else:
            cost_of_debt = pretax_cost_of_debt * (1-average(self.forward_etr) / 100)

        # Cost of equity
        ticker = self.tick
        if re.match(r'\d{4}$', self.tick):
            # Regex to match currency denomination express by 4 digits such as 9618.HK
            suffix = self.dataset.get_currency_suffix(self.sticky_price)
            ticker = ticker + '.{}'.format(suffix)
        yf_ticker = yf.Ticker(ticker)
        if 'beta' not in yf_ticker.info:
            assert len(self.head) > 0
            initial_query = ' '.join(self.head.split()[:-1])
            print("Waiting to query Yahoo Finance server with '{}'".format(initial_query))
            yf_ticker = yf.Ticker(get_symbol(initial_query))

        if 'beta' in yf_ticker.info:
            beta = yf_ticker.info['beta']
            print("Obtain beta:", beta)
        else:
            colour_print("Invalid beta: defaulting beta to 1.0", bcolors.WARNING)
            beta = 1.0
        mrp = self.dataset.get_equity_risk_premium()
        cost_of_equity = self.riskfree_rate + beta * mrp
        market_cap = yf_ticker.info['marketCap'] / 1e6

        total_cap = market_cap + debt
        initial_coc = market_cap/total_cap * cost_of_equity + debt/total_cap * cost_of_debt

        # Mature market ERP set to 4.5% and 0% based on U.S. CRP (Country risk premium)
        # https://www.youtube.com/watch?v=kyKfJ_7-mdg
        # https://pages.stern.nyu.edu/~adamodar/New_Home_Page/datafile/ctryprem.html
        crp = 0
        mature_market_erp = .045
        total_erp = crp + mature_market_erp
        coc = d['Cost of capital'] = [0] + [initial_coc]*5
        for i in range(1, total_half_col):
            # Prev coc - (fixed prev coc - riskfree rate + mature market risk + country risk premium)/5
            current_coc = coc[i] - (initial_coc - (self.riskfree_rate + total_erp))/5
            coc.append(current_coc)
        coc.append(self.riskfree_rate + mature_market_erp)

    def compute_cumulative_df(self, d):
        coc = d['Cost of capital']
        fcff = d['FCFF']
        cumulated_df = d['Cumulated discount factor'] = []
        pv = d['PV (FCFF)'] = []
        for i in range(total_main_col-1):
            current_coc = 0
            if coc[i] is not None:
                current_coc = coc[i]
            current_df = 1/(1+current_coc)
            if len(cumulated_df) > 0:
                current_df = cumulated_df[i-1]*(1/(1+current_coc))
            cumulated_df.append(current_df)

            current_fcff = 0
            if fcff[i] is not None:
                current_fcff = fcff[i]
            # fcff * df
            pv.append(current_fcff * cumulated_df[i])

    def compute_terminals(self, d):
        d['empty2'] = []
        d['Terminal cash flow'] = d['FCFF'][-1]
        d['Terminal cost of capital'] = d['Cost of capital'][-1]
        d['Terminal value'] = (d['Terminal cash flow'] /
                               (d['Terminal cost of capital'] - d['Revenue growth rate'][-1]))
        # print( d['Cumulated discount factor'] )
        d['PV (Terminal value)'] = d['Terminal value'] * d['Cumulated discount factor'][-1]
        d['PV (Cash flow over next 10 years)'] = sum(d['PV (FCFF)'])
        d['Sum of PV'] =  d['PV (Terminal value)'] + d['PV (Cash flow over next 10 years)']
        d['Value of operating assets'] =  d['Sum of PV']

        debt = 0
        if self.debt[-1] is not None:
            debt = self.debt[-1]
        d['- Debt'] = debt

        d['- Minority interest'] = 0
        d['+ Cash'] = self.cash[-1]
        if type(self.investments) is list:
            # type: List[float]
            non_op = 0
            if self.investments[-1] is not None:
                non_op = self.investments[-1]
            elif self.investments[-2] is not None:
                colour_print("Latest investment was not defined. Fallback to previous year", bcolors.WARNING)
                non_op = self.investments[-2]
            d['+ Non-operating assets'] = non_op
        else:
            d['+ Non-operating assets'] = 0
        d['Value of equity'] = (d['Value of operating assets']
                                - d['- Debt'] - d['- Minority interest']
                                + d['+ Cash'] + d['+ Non-operating assets'])
        d['Number of shares'] = self.shares[-1]
        d['Estimated value / share'] = d['Value of equity'] / d['Number of shares']

        ticker = yf.Ticker(self.tick)
        avg_price = (ticker.info['regularMarketDayLow'] + ticker.info['regularMarketDayHigh']) / 2.
        # d['Price'] = self.strip(self.values.match_title('Price$'))[-1]
        d['Price'] = avg_price
        d['Price as % of value'] = avg_price / d['Estimated value / share']

    def compute_return_invested_capital(self, d):
        ## Invested capital
        # TODO =IF(operating_lease="Yes",
        #     IF(rnd_expense_cap="Yes",
        #        book_value_equity + book_value_debt - cash + adjust_debt_outstanding + rnd_converter,
        #        book_value_equity + book_value_debt - cash + adjust_debt_outstanding),
        #     IF(rnd_expense_cap="Yes",
        #        book_value_equity + book_value_debt - cash + rnd_converter,
        #        book_value_equity + book_value_debt - cash))
        d['empty3'] = []
        invested_capital = d['Invested Capital'] = []
        book_value_debt = 0
        if self.book_value_debt[-1] is not None:
            book_value_debt = self.book_value_debt[-1]
        current_ic = self.book_value_equity[-1] + book_value_debt - self.cash[-1]
        prev_ic = []
        reinvestment = d['- Reinvestment']
        for i in range(total_main_col-1):
            prev_ic.append(current_ic)
            if reinvestment[i] is not None:
                current_ic += reinvestment[i]
            invested_capital.append(current_ic)

        invested_return = d['ROIC'] = []
        for i in range(total_main_col-1):
            invested_return.append(d['NOPAT'][i] / prev_ic[i])

        # TODO end of cost of capital
        # invested_return.append(d['Cost of capital'][-1])


class Ticks:
    pass
    # def __init__(self, ticks, path):
    #     self.tickers = ticks
    #     self.path = path
    #     self.modes = [
    #         Mode.Last_FCF,
    #         Mode.Avg_FCF,
    #         Mode.UFCF_5x,
    #         Mode.UFCF_8x,
    #     ]
    #
    #     # type: List[DCF]
    #     self.spreads = []
    #     for t in ticks:
    #         self.spreads.append(DCF(t, self.path, self.modes))
    #
    # def compute(self):
    #     for sp in self.spreads:
    #         sp.compute_fcf()
    #
    # def summarize(self):
    #     # https://blog.devgenius.io/how-to-easily-print-and-format-tables-in-python-18bbe2e59f5f
    #     # summarize to TV based on last FCF, avg FCF, multiple of NOPAT
    #
    #     # Set the first word as the header for our spreadsheet optionally.
    #     a = list(map(lambda x: [x.short_name() if x.head is not None else x.head,
    #                             x.tick, x.last_price], self.spreads))
    #     for i, s in enumerate(self.spreads):
    #         for j in range(len(s.sum_pvtv)):
    #             # extend the spread with sum_pvtv and poss ratio
    #             a[i].append(s.sum_pvtv[j])
    #             a[i].append(s.poss_ratio[j])
    #
    #     # Sort it to the last column which is 'Possible' ratio
    #     entries = sorted(a, key=lambda x: x[len(a[0])-1])
    #     poss_header = 'Poss. x'
    #     heads = ['Company', 'Tick', 'Last price']
    #     for i, m in enumerate(self.modes):
    #         # TODO mapping to the underlying string
    #         h = {Mode.Last_FCF: 'Last FCF',
    #              Mode.Avg_FCF: 'Avg. FCF',
    #              Mode.UFCF_5x: 'UFCF 5x',
    #              Mode.UFCF_8x: 'UFCF 8x'}[m]
    #         heads.extend((h, poss_header))
    #     print(tabulate(entries, headers=heads,
    #                    tablefmt='fancy_grid', stralign='center', numalign='center', floatfmt=".2f"))
    #
    #     # generate Excel output
    #     # styles = ['Comma', 'Comma', 'Comma']
    #     # for _ in self.modes:
    #     #     styles.extend(('Comma', 'Percent'))
    #     # excel = ExcelOut(self.tickers, entries, styles=styles, headers=heads)
    #     # excel.start()


dcf = DCF('intc', country='United States')
dcf.compute()
print("XXX", dcf)

# Damodaran main data page
# https://pages.stern.nyu.edu/~adamodar/New_Home_Page/datacurrent.html
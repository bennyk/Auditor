from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet import worksheet
# from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from typing import List, Tuple, Dict, Union
from collections import OrderedDict, namedtuple
from enum import Enum
import re
import decimal
import statistics
import datetime
from functools import partial
import math
from bcolors import bcolors, colour_print
from spread import Table

max_row = max_col = 99

Unit = namedtuple('Unit', ['value', 'symbol'])
RateType = namedtuple('Rate', ['above_avg', 'moderate_avg', 'below_avg'])

RateVerbose = {RateType.above_avg: 'above',
               RateType.moderate_avg: 'moderately at',

               # TODO "do not perform" below level
               RateType.below_avg: 'below'}


def colnum_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string


# TODO skipped first element in the list
def striped_average(l: [float], prefix):
    l = strip(l, prefix)
    return sum(l) / len(l)


def average(l: [float]):
    return sum(filter(None, l)) / len(l)


def strip(l, prefix, trim_last=False):
    if trim_last:
        return l[prefix:][:-1]
    return l[prefix:]


def strip2(l, prefix, trim_last=False):
    # TODO need to check against original strip implementation.
    if trim_last:
        return l[prefix:][:-1]

    _ = l[prefix:]
    if type(_[0]) is str:
        return list(map(lambda x: float(re.sub(r'(\d+)x', r'\1', x)), _))
    return _


def list_over_list(x, y, percent=False):
    if percent:
        # return list(map(lambda n1, n2: (n1 / n2), x, y))
        return list(map(lambda n1, n2: 0 if n1 is None else 100 * (n1 / n2), x, y))
    return list(map(lambda n1, n2: 0 if n1 is None else n1 / n2, x, y))


def list_multiply_list(x, y):
    return list(map(lambda n1, n2: n1*n2, x, y))


def list_add_list(x, y):
    a = map(lambda n1: 0 if n1 is None else n1, x)
    b = map(lambda n2: 0 if n2 is None else n2, y)
    return list(map(lambda n1, n2: n1 + n2, a, b))


def list_minus_list(x, y):
    a = map(lambda n1: 0 if n1 is None else n1, x)
    b = map(lambda n2: 0 if n2 is None else n2, y)
    return list(map(lambda n1, n2: n1 - n2, a, b))


def list_abs(x):
    return list(map(lambda n1: abs(n1), x))


def list_one(x, l):
    return [1.] * l


def cagr(l: [float]) -> float:
    # standard CAGR formula:
    # https://corporatefinanceinstitute.com/resources/valuation/what-is-cagr/
    # Otherwise we can use piecewise function to workaround complex number
    # h(x) = -|x|^(1/n) for x < 0
    # h(x) = |x|^(1/n) for x >= 0
    # alternate interpretation
    #   p = (l[len(l)-1] + abs(l[0])) / abs(l[0])
    #   https://www.exceldemy.com/how-to-calculate-cagr-in-excel-with-negative-number/?utm_source=pocket_saves

    # preceding year
    y = -1
    ri = 1/len(l)
    if l[0] < 0:
        a = l[len(l)-1] / l[0]
        if l[len(l)-1] >= 0:
            p = - abs(a) ** ri + y
        else:
            p = a ** ri + y
    else:
        p = (l[len(l)-1] / l[0]) ** ri + y
        if type(p) is complex:
            # if complex try force negation with abs
            assert l[len(l)-1] < 0
            a = l[len(l)-1] / l[0]
            p = - abs(a) ** ri + y
    assert type(p) is not complex
    return p


def zsum(a):
    return sum(list(map(lambda x: x if x is not None else 0, a)))


class Spread:
    Percent_Denominator = 100

    def __init__(self, wb: Workbook, tick: str, prof: 'Prof'):
        self.tick = tick
        self.profiler = prof
        self.tabs = []
        self.income = None
        self.balance = None
        self.cashflow = None
        self.values = None

        prefix_index = 1
        self.start_prefix = prefix_index
        self.strip = partial(strip, prefix=prefix_index)
        self.strip2 = partial(strip2, prefix=prefix_index)

        for index, name in enumerate(wb.sheetnames):
            tab = Table(wb[name])
            self.tabs.append(tab)
            if re.match(r'Income', name):
                self.income = tab
                self.start_date = self.income.tab[0][1]
                # -2 ignore year LTM
                self.end_date = self.income.tab[0][-2]
            elif re.match(r'Balance', name):
                self.balance = tab
            elif re.match(r'Cash', name):
                self.cashflow = tab
            elif re.match(r'Values', name):
                self.values = tab
            else:
                # passing Ratios
                pass

        # 2: Skip int to str in start and end years which extract two digit, year number, in alphabets
        if type(self.end_date) is datetime.datetime:
            self.end_year = int(str(self.end_date.year)[2:])
        else:
            self.end_year = int(self.end_date.split('/')[-1])

        if type(self.start_date) is datetime.datetime:
            self.start_year = int(str(self.start_date.year)[2:])
        else:
            self.start_year = int(self.start_date.split('/')[-1])
        print("Sampled from {} to {} in {} years".format(
            self.start_date, self.end_date,
            1+self.end_year-self.start_year))

        self.half_len = int((self.end_year-self.start_year+1)/2)

    def revenue(self):
        revs = self.strip(self.income.match_title('Total Revenues'))
        rev_per_share = list_over_list(revs, self.wa_diluted_shares_out())
        cagr_rev_per_share = cagr(rev_per_share)
        print("Revenue per share from {} to {} at CAGR {:.2f}% for: {}".format(
            revs[0], revs[-1],
            cagr_rev_per_share * 100, revs))
        if abs(cagr(revs) - cagr_rev_per_share) > .01:
            print("   Revenue {:.2f} in percent vs {:.2f} on per share basis".format(
                cagr_rev_per_share * 100, cagr_rev_per_share * 100))

        last_cagr_rev_per_share = cagr(rev_per_share[-2:])
        if rev_per_share[-2] == rev_per_share[-1]:
            last_cagr_rev_per_share = cagr(rev_per_share[-3:-1])

        self.profiler._collect(cagr_rev_per_share, Tag.rev_per_share, ProfMethod.CAGR,
                               val2=cagr(rev_per_share[self.half_len:]),
                               val3=last_cagr_rev_per_share)

    def epu(self):
        # EBT exclude unusual include Interest Expense, Investment Income, Income on Equity Investment,
        # Currency Exchange, Other Non Operating Income
        # Excluding M&A, Gain(Loss) sale of investments, legal settlements
        ebt_exclude_unusual = self.strip(self.income.match_title('EBT Excl. Unusual Items'))
        epu_per_share = list_over_list(ebt_exclude_unusual, self.wa_diluted_shares_out())
        cagr_epu_ratio = cagr(epu_per_share)
        print("EPU from {:.2f} to {:.2f} at CAGR {:.2f}% for: {}".format(
            epu_per_share[0], epu_per_share[-1],
            cagr_epu_ratio*Spread.Percent_Denominator, epu_per_share))

        if epu_per_share[-2] == epu_per_share[-1]:
            last_cagr_epu_ratio = cagr(epu_per_share[-3:-1])
        else:
            last_cagr_epu_ratio = cagr(epu_per_share[-2:])

        self.profiler._collect(cagr_epu_ratio, Tag.epu, ProfMethod.CAGR,
                               val2=cagr(epu_per_share[self.half_len:]),
                               val3=last_cagr_epu_ratio)

    def owner_yield(self):
        # also known as Levered FCF
        earning_not_strip = self.cashflow.match_title('Free Cash Flow$', none_is_optional=True)
        if earning_not_strip is not None:
            earning = self.strip(earning_not_strip)
        else:
            cfo = self.strip(self.cashflow.match_title('Cash from Operations$'))
            opt_acq_real_assets = self.cashflow.match_title('Acquisition of Real Estate Assets$',
                                                            none_is_optional=True)
            earning = cfo
            if opt_acq_real_assets is not None:
                acq_real_assets = self.strip(opt_acq_real_assets)
                earning = list_add_list(cfo, acq_real_assets)
        earning_yield = list_over_list(earning, self.wa_diluted_shares_out())
        avg_yield = statistics.median(earning_yield)
        print("Earning yield from {:.2f} to {:.2f} at average {:.2f}% for: {}".format(
            earning_yield[0], earning_yield[-1],
            avg_yield, earning_yield))
        self.profiler._collect(avg_yield, Tag.owner_yield, ProfMethod.Average,
                               val2=statistics.median(earning_yield[self.half_len:]),
                               val3=earning_yield[-1])

    def cfo(self):
        # aka FFO - Funds from Operations
        cfo = self.strip(self.cashflow.match_title('Cash from Operations'))
        cfo_per_share = list_over_list(cfo, self.wa_diluted_shares_out())
        cagr_cfo_per_share = cagr(cfo_per_share)
        print("FCF per share from {} to {} at CAGR {:.2f}% for: {}".format(
            cfo[0], cfo[-1],
            cagr_cfo_per_share * 100, cfo))
        if abs(cagr(cfo) - cagr_cfo_per_share) > .01:
            print("   FCF {:.2f} in percent vs {:.2f} on per share basis".format(
                cagr_cfo_per_share*100, cagr_cfo_per_share * 100))
        self.profiler.collect(cagr_cfo_per_share, cfo[-1], 'cfo_per_share', ProfMethod.CAGR)

    def _affo(self):
        cfo = self.strip(self.cashflow.match_title('Cash from Operations'))
        # Capex for real estates
        capex = self.strip(self.cashflow.match_title('Capital Expenditure'))
        affo = list_add_list(cfo, capex)

        # TODO made comparison in relation to IGBREIT's share out filing
        # share_out_filing = self.share_out_filing()
        share_out_filing = 3600
        affo_per_share = list(map(lambda f: f / share_out_filing, affo))

        # Based on IGBREIT 2021 annual: "term period" between 5.85% to 6.85%. Take the mid point.
        irr = .0635
        term_period = 0

        # reversed - Simulate in reversed order from current to far past year.
        last_term = None
        for i, a in enumerate(reversed(affo)):
            last_term = a/(1+irr)**(i+1)
            term_period += last_term
        # print("XXX", term_period, term_period/self.share_out_filing())
        last_term_period_over_shares = last_term / share_out_filing
        avg_term_period_over_shares = term_period / share_out_filing
        print("AFFO in relation to IGBREIT, at IRR {:.4f} for: {}".format(
            avg_term_period_over_shares,
            list(map(lambda x: round(x, 4), affo_per_share))
        ))
        self.profiler.collect(avg_term_period_over_shares, last_term_period_over_shares,
                              Tag.affo_per_share, ProfMethod.IRR)

    def affo(self):
        cfo = self.strip(self.cashflow.match_title('Cash from Operations'))
        # Capex for real estates
        capex = self.strip(self.cashflow.match_title('Capital Expenditure'))
        affo = list_add_list(cfo, capex)
        affo_per_share = list_over_list(affo, self.wa_diluted_shares_out())
        # use Median rather than average.
        avg_affo_per_share = statistics.median(affo_per_share)
        print("AFFO at average {:.2f}% for: {}".format(
            avg_affo_per_share*100,
            list(map(lambda x: round(x, 4), affo_per_share))
        ))
        self.profiler.collect(avg_affo_per_share, affo_per_share[-1], Tag.affo_per_share, ProfMethod.AveragePerc)

    def nav(self):
        total_asset = self.strip(self.balance.match_title('Total Assets'))
        total_liab = self.strip(self.balance.match_title('Total Liabilities'))
        nav = list_minus_list(total_asset, total_liab)
        nav_per_share = list_over_list(nav, self.wa_diluted_shares_out())
        avg_nav_per_share = cagr(nav_per_share)
        print("NAV per share at CAGR {:.2f}% for: {}".format(
            avg_nav_per_share*100,
            list(map(lambda x: round(x, 4), nav_per_share)),
        ))
        self.profiler.collect(avg_nav_per_share, nav_per_share[-1], Tag.nav_per_share, ProfMethod.CAGR)

    def tangible_book(self):
        total_equity = self.strip(self.balance.match_title('Total Equity'))
        goodwill_not_strip = self.balance.match_title('Goodwill', none_is_optional=True)
        tangible = total_equity
        if goodwill_not_strip is not None:
            goodwill = self.strip(goodwill_not_strip)
            tangible = list_minus_list(total_equity, goodwill)
        intangible_not_strip = self.balance.match_title('Other Intangibles', none_is_optional=True)
        if intangible_not_strip is not None:
            intangible = self.strip(intangible_not_strip)
            tangible = list_minus_list(tangible, intangible)
        tangible_per_share = list_over_list(tangible, self.wa_diluted_shares_out())
        avg_tangible_per_share = cagr(tangible_per_share)
        print("Tangible book per share at CAGR {:.2f}% for: {}".format(
            avg_tangible_per_share*100,
            list(map(lambda x: round(x, 4), tangible_per_share)),
        ))
        self.profiler.collect(avg_tangible_per_share, tangible_per_share[-1], Tag.tangible_per_share, ProfMethod.CAGR)

    def return_equity(self):
        net_income = self.strip(self.income.match_title('Net Income$'))
        requity = self.strip(self.balance.match_title('Total Common Equity$'))
        roce = list_over_list(net_income, requity, percent=True)
        avg_roce = average(roce)
        print("Return on Common Equity average {:.2f}% for: {}".format(
            avg_roce,
            list(map(lambda x: round(x, 2), roce))
        ))
        # TODO ROCE is not defined
        self.profiler.collect(avg_roce/100, roce[-1], Tag.ROCE, ProfMethod.AveragePerc)

    def return_invested_cap(self):
        # ROIC = (nopat - tax) / (equity + debt + cash)
        # https://www.educba.com/invested-capital-formula/
        # https://www.thebalancemoney.com/return-on-invested-capital-393587#toc-how-to-calculate-roic

        op_income = self.strip(self.income.match_title('Operating Income'))
        tax_not_strip = self.income.match_title('Income Tax Expense', none_is_optional=True)
        if tax_not_strip is not None:
            # op_income after tax calculation
            # https://www.youtube.com/watch?v=QsqzDNOt89c

            tax = self.strip(tax_not_strip)
            tax_rate = list_over_list(tax, op_income)
            # op_income * (1-tax_rate)
            op_income_after_tax = list_add_list(list_one(1, len(tax)), tax_rate)
            nopat = list_multiply_list(op_income, op_income_after_tax)
            pass
        else:
            nopat = op_income
        debt = self.strip(self.balance.match_title('Total Debt$'))
        equity = self.strip(self.balance.match_title('Total Equity$'))

        # Minus non-operating-assets
        cash = self.strip(self.cashflow.match_title('Cash from Investing$'))
        cash = list_add_list(
            self.strip(self.cashflow.match_title('Cash from Financing$')), cash)

        _1 = list_add_list(debt, equity)
        _2 = list_add_list(_1, cash)
        roic_per = list_over_list(nopat, _2, percent=True)
        avg_roic_per = average(roic_per)
        print("Return on Invested Capital average {:.2f}% for: {}".format(
            avg_roic_per,
            list(map(lambda x: round(x, 2), roic_per))
        ))
        self.profiler._collect(avg_roic_per/Spread.Percent_Denominator, Tag.ROIC, ProfMethod.AveragePerc,
                               val2=average(roic_per[self.half_len:])/Spread.Percent_Denominator,
                               val3=roic_per[-1]/Spread.Percent_Denominator)

    def net_debt_over_ebit(self):
        net_debt = self.strip(self.balance.match_title('Net Debt'))

        # Compute EBIT as Net income - Tax income expense - Interest expense.
        # Tax expense and Interest expense values from TIKR terminal have been negated.
        net_income = self.strip(self.income.match_title('Net Income$'))
        tax_expense = self.strip(self.income.match_title('Income Tax Expense$'))
        ebit = list_minus_list(net_income, tax_expense)
        interest_expense = self.strip(self.income.match_title('Interest Expense$'))
        ebit = list_minus_list(ebit, interest_expense)

        # EBIT may be more appropriate, as the Depreciation and Amortization captures
        # a portion of past capital expenditures.
        # https://corporatefinanceinstitute.com/resources/valuation/ebit-vs-ebitda/
        net_debt_over_ebit = list_over_list(net_debt, ebit)
        avg_net_debt_over_ebit = average(net_debt_over_ebit)
        print("Net debt over EBIT average {:.2f} years for: {}".format(
            avg_net_debt_over_ebit,
            list(map(lambda x: round(x, 2), net_debt_over_ebit))
        ))
        self.profiler._collect(avg_net_debt_over_ebit, Tag.net_debt_over_ebit, ProfMethod.AverageYears,
                               val2=average(net_debt_over_ebit[self.half_len:]),
                               val3=net_debt_over_ebit[-1])

        ebitda = self.strip(self.income.match_title('EBITDA$'))
        net_debt_over_ebitda = list_over_list(net_debt, ebitda)
        avg_net_debt_over_ebitda = average(net_debt_over_ebitda)
        colour_print("EBITDA: Net debt over EBITDA average {:.2f} years for: {}".format(
            avg_net_debt_over_ebitda,
            list(map(lambda x: round(x, 2), net_debt_over_ebitda))), bcolors.WARNING)

    def net_debt_over_fcf(self):
        # TODO net_debt_over_fcf
        net_debt = self.strip(self.balance.match_title('Net Debt'))
        fcf_not_strip = self.cashflow.match_title('Free Cash Flow$', none_is_optional=True)
        if fcf_not_strip is not None:
            fcf = self.strip(fcf_not_strip)
        else:
            cfo = self.strip(self.cashflow.match_title('Cash from Operations$'))
            opt_acq_real_assets = self.cashflow.match_title('Acquisition of Real Estate Assets$',
                                                            none_is_optional=True)
            fcf = cfo
            if opt_acq_real_assets is not None:
                acq_real_assets = self.strip(opt_acq_real_assets)
                fcf = list_add_list(cfo, acq_real_assets)
        net_debt_over_fcf = list_over_list(net_debt, fcf)
        avg_net_debt_over_fcf = average(net_debt_over_fcf)
        print("Net debt over FCF average {:.2f} years for: {}".format(
            avg_net_debt_over_fcf,
            list(map(lambda x: round(x, 2), net_debt_over_fcf))
        ))
        self.profiler._collect(avg_net_debt_over_fcf, Tag.net_debt_over_fcf, ProfMethod.AverageYears,
                               val2=average(net_debt_over_fcf[self.half_len:]),
                               val3=net_debt_over_fcf[-1])

    def op_margin(self):
        op_income = self.strip(self.income.match_title('Operating Income$'))
        revs = self.strip(self.income.match_title('Total Revenues$'))
        op_margins = list_over_list(op_income, revs, percent=True)
        avg_op_margins = average(op_margins)
        print("Operating margin average {:.2f}% for (numbers in percent) {}".format(
        avg_op_margins,
            list(map(lambda x: round(x, 2), op_margins))
        ))
        self.profiler._collect(avg_op_margins / Spread.Percent_Denominator, Tag.op_margin, ProfMethod.AveragePerc,
                               val2=average(op_margins[self.half_len:])/Spread.Percent_Denominator,
                               val3=op_margins[-1]/Spread.Percent_Denominator)

    def ev_over_ebit(self):
        if self.values is None:
            # TODO exception to EV over EBIT
            print("Warning: ev_over_ebit: Missing values tab.")
            return
        ev_over_ebit = self.strip2(self.values.match_title(r'LTM\s+Total\s+Enterprise\s+Value\s*/\s*EBIT$'))
        avg_ev_over_ebit = average(ev_over_ebit)
        print("EV over EBIT average {:.2f} ratio for: {}".format(
            avg_ev_over_ebit,
            list(map(lambda x: .0 if x is None else round(x, 2), ev_over_ebit))
        ))
        self.profiler._collect(avg_ev_over_ebit, Tag.ev_over_ebit, ProfMethod.ReverseRatio,
                               val2=average(ev_over_ebit[self.half_len:]),
                               val3=ev_over_ebit[-1])

    # TODO retined earnings pay in advance for one year?
    def retained_earnings_ratio(self):
        retained_earnings = self.strip(self.balance.match_title('Retained Earnings$'))
        net_income = self.strip(self.income.match_title('Net Income$'))
        retention_ratio = list_over_list(retained_earnings, net_income)

        avg_retention_ratio = average(retention_ratio)
        print("Retention ratio average {:.2f}, last {:.2f} for: {}".format(
            avg_retention_ratio,
            retention_ratio[-1],
            list(map(lambda x: round(x, 2), retention_ratio))
        ))

        # TODO
        # EPS over Retained earnings per share > 4% Buffett said will be good
        # EPS change (five years) / total retained earnings (five years).
        # https://finance.yahoo.com/news/key-metrics-retained-earnings-market-171931264.html?utm_source=pocket_reader

        self.profiler._collect(avg_retention_ratio, Tag.retained_earnings_ratio, ProfMethod.Ratio,
                               val2=average(retention_ratio[self.half_len:]),
                               val3=retention_ratio[-1])

    def market_cap_over_retained_earnings_ratio(self):
        multiples = self.strip(self.values.match_title('Multiples'))
        avail_start_year = int(multiples[0].split('/')[-1])
        start_year = int(multiples[0].split('/')[-1])
        if avail_start_year > start_year:
            # cut short the date range
            offset = avail_start_year - start_year + 1
            retained_earnings = self.balance.match_title('Retained Earnings')[offset:]
            print("Start year in 20{} has been adjusted to available year 20{}".format(
                start_year,
                avail_start_year))
        else:
            retained_earnings = self.strip(self.balance.match_title('Retained Earnings$'))

        # No exclusion in Valuation/Multiples sub
        market_cap = self.strip(self.values.match_title('Market Cap'))
        i = 0
        MC = []
        while i < len(market_cap):
            MC.append(average(market_cap[i:i+4]))
            # print(a, i, j)
            i += 4

        MC_change = MC[-1]-MC[0]
        market_over_retained = MC_change / zsum(retained_earnings)
        print("MC_change over Retained earnings ratio is {:.2f}. MC samples: {}".format(
            market_over_retained,
            list(map(lambda x: round(x, 2), MC))
        ))

        val2 = None
        if len(MC) > self.half_len:
            val2 = (MC[-1]-MC[self.half_len]) / zsum(retained_earnings[self.half_len:])

        val3 = None
        if len(MC) > 0:
            val3 = (MC[-1]-MC[-2]) / zsum(retained_earnings[-2:])

        self.profiler._collect(market_over_retained, Tag.market_cap_ov_retained_earnings, ProfMethod.ReverseRatio,
                               val2=val2,
                               val3=val3)

    def dividend_payout_ratio(self):
        div_paid = self.strip(self.cashflow.match_title('Common Dividends Paid'))
        for i, a in enumerate(div_paid):
            if a is None:
                print("W: {} does not provide dividend in year '{}".format(
                    self.tick, self.start_year+i))
        income = self.strip(self.income.match_title('Net Income'))

        # Op income is a probable replacement in the event when regular income produce negative number.
        op_income = self.strip(self.income.match_title('Operating Income'))
        net_income = []
        for i, a in enumerate(income):
            if a < 0:
                print("W: {} negative income {} in year '{}".format(
                    self.tick, a, self.start_year+i))
                net_income.append(op_income[i])
            else:
                net_income.append(income[i])

        div_payout_ratio = list_over_list(div_paid, net_income)
        # Negating div payout to positive for the math to work easier
        avg_div_payout_ratio = - average(div_payout_ratio)
        print("Dividend payout ratio at average {:.2f} ratio for: {}".format(
            avg_div_payout_ratio,
            list(map(lambda x: round(x, 2), div_payout_ratio))
        ))
        self.profiler.collect(avg_div_payout_ratio, - div_payout_ratio[-1],
                              Tag.dividend_payout_ratio, ProfMethod.Average)

    def div_yield(self):
        if self.values is None:
            # TODO exception to EV over EBIT
            print("Warning: dividend yield: Missing values tab.")
            return
        result = self.values.match_title('LTM Dividend Yield$', none_is_optional=True)
        if result is not None:
            div_yields = list(map(
                lambda z: 0 if z is None else z, self.strip2(self.values.match_title('LTM Dividend Yield$')))
            )
            avg_div_yield = average(div_yields)
            self.profiler._collect(avg_div_yield, Tag.dividend_yield, ProfMethod.Average,
                                   val2=average(div_yields[self.half_len:]),
                                   val3=div_yields[-1])
            print("Dividend yield ratio at average {:.2f} ratio for: {}".format(
                avg_div_yield,
                list(map(lambda x: round(x, 2), div_yields))
            ))
        else:
            print("No dividend yield was provided")
            self.profiler._collect(0, Tag.dividend_yield, ProfMethod.Average,
                                   val2=0, val3=0)

    def last_price(self):
        if self.values is None:
            print("Warning: last_price: Missing values tab.")
            return

        price = self.values.match_title('Price$')
        ev_over_ebit = self.strip2(self.values.match_title(r'LTM\s+Total\s+Enterprise\s+Value\s*/\s*EBIT$'))
        # TODO all div yields data?
        div_yield = None
        _ = self.values.match_title('LTM Dividend Yield$', none_is_optional=True)
        if _ is not None:
            div_yield = self.strip2(self.values.match_title('LTM Dividend Yield$'))

        market_cap = self.values.match_title('Market Cap')
        rev = self.income.match_title('Total Revenues')
        op_income = self.income.match_title('Operating Income')
        net_profit = self.income.match_title('Net Income')

        dpu = self.income.match_title('Dividends Per Share', none_is_optional=True)
        last_dpu = 0 if dpu is None or dpu[-1] is None else dpu[-1]

        # AFFO commented diff
        # price_over_affo = price[-1]/self.profiler.d[Tag.affo_per_share]['val2']

        last_div_yield = 0 if div_yield is None or div_yield[-1] is None else div_yield[-1]
        self.profiler.collect_last_price({'last_price': price[-1],
                                          'ev_over_ebit': ev_over_ebit[-1],
                                          'last_div_yield': last_div_yield,
                                          'div_yields': div_yield,
                                          'market_cap': market_cap[-1],
                                          'revenue': rev[-1],
                                          'op_income': op_income[-1],
                                          'net_profit': net_profit[-1],
                                          'dpu': last_dpu})
        # AFFO commented diff
        # 'price_over_affo': price_over_affo})

    def share_out_filing(self) -> [float]:
        # "Total Shares Out. Filing Date" is provided in Balance Sheet which computed as fully year
        # or last trailing year I think.
        # TODO Currently it is being replace by "Weighted Average Diluted Shares Outstanding"
        x = self.balance.match_title('Total Shares Out\.')
        result = list(filter(None, reversed(x[self.start_prefix:])))[0]
        return result

    def wa_diluted_shares_out(self) -> [float]:
        result = self.strip(self.income.match_title('Weighted Average Diluted Shares Outstanding'))
        return result


class ProfMethod(Enum):
    CAGR = 1
    IRR = 2
    Average = 3
    AveragePerc = 4
    AverageYears = 5
    Ratio = 6
    ReverseRatio = 7


ProfVerbose = {ProfMethod.CAGR: 'CAGR',
               ProfMethod.IRR: 'IRR',
               ProfMethod.Average: 'average',
               ProfMethod.AveragePerc: 'percent',
               ProfMethod.AverageYears: 'year',
               ProfMethod.Ratio: 'ratio',
               ProfMethod.ReverseRatio: 'ratio',
               }


class Tag(Enum):
    rev_per_share = 1
    epu = 2
    owner_yield = 13
    # AFFO commented diff
    # affo_per_share = 3
    # nav_per_share = 4
    # tangible_per_share = 14
    # TODO ROCE is optional for tabulation.
    # ROCE = 5
    ROIC = 6
    net_debt_over_ebit = 7
    # net_debt_over_fcf = 15
    ev_over_ebit = 8
    op_margin = 9
    retained_earnings_ratio = 10
    market_cap_ov_retained_earnings = 14
    # dividend_payout_ratio = 11
    dividend_yield = 12


class Prof:
    def __init__(self, name):
        self.name = name
        self.d = OrderedDict()
        self.last_price = None

        self.prof = {}

    def collect(self, val1, val3, tag, method: ProfMethod):
        # self.d[tag.__name__] = ratio
        # type: Tuple[float, ProfMethod]
        _ = OrderedDict(val1=val1, val2=val3, method=method)
        self.d[tag] = _

    def _collect(self, val1, tag, method: ProfMethod,
                 val2=None, val3=None):
        # self.d[tag.__name__] = ratio
        # type: Tuple[float, ProfMethod]
        _ = OrderedDict(val1=val1, val2=val2, val3=val3, method=method)
        self.d[tag] = _

    def collect_last_price(self, last_price):
        self.last_price = last_price

    def profile(self):
        for k, v in self.d.items():
            if k is not str:
                self.prof[k] = v


class Bucket:
    def __init__(self, name, value, method):
        self.name = name
        self.value = value
        self.method = method
        self.score = None


class ProfManager:
    # TODO report about the underlying rate?
    # AFFO and Tangible commented diffs
    Rate = {Tag.rev_per_share: {'high': .1, 'mid': .05},
            Tag.epu: {'high': .2, 'mid': .1},
            Tag.owner_yield: {'high': 3., 'mid': 1.},
            # Tag.affo_per_share: {'high': 4., 'mid': 2.},
            # Tag.nav_per_share: {'high': .08, 'mid': .05},
            # Tangible commented diff
            # Tag.tangible_per_share: {'high': .2, 'mid': .0},
            # TODO ROCE is undefined
            # Tag.ROCE: {'high': .08, 'mid': .065},
            Tag.ROIC: {'high': .5, 'mid': .3},
            Tag.net_debt_over_ebit: {'high': 5., 'mid': 8.},
            # Tag.net_debt_over_fcf: {'high': 5., 'mid': 8.},
            Tag.op_margin: {'high': .3, 'mid': .2},
            Tag.retained_earnings_ratio: {'high': 4., 'mid': 2.},
            # Tag.dividend_payout_ratio: {'high': 1.5, 'mid': 1.},
            Tag.market_cap_ov_retained_earnings: {'high': .5, 'mid': 1.},
            Tag.dividend_yield: {'high': .05, 'mid': .02},
            Tag.ev_over_ebit: {'high': 12., 'mid': 20.},
            }

    UnitRatio = (ProfMethod.AverageYears, ProfMethod.Ratio, ProfMethod.ReverseRatio)

    def __init__(self):
        self.companies = []     # type: List[Prof]
        self.company = {}
        self.metric = {}

    def create_folder(self, name):
        prof = Prof(name)
        self.companies.append(prof)
        self.company[name] = prof
        return prof

    def profile(self):
        for p in self.companies:
            # print(p.name)
            p.profile()
        self.bucketize()
        benched = self.benchmark()
        self.simulate_price(benched['comp'])
        self.output(benched)

    # Red = 'FF0000'
    # Yellow = 'FFFF00'
    # Green = '00FF00'

    # Rainbow Pastels Color Scheme
    # https://mappingmemories.ca/combfkiu2426925
    Red = 'FF9AA2'
    Yellow = 'FFDAC1'
    # Green = 'E2F0CB'
    Green = 'B5EAD7'

    # STABILO Boss Pastel colors palette | ColorsWall
    # Red = 'F6AA90'
    # Yellow = 'F8DF81'
    # Green = '9BDB07'

    def output(self, benched: Dict[str, dict]):
        # benched to access 'score' by company's name
        wrap = WorkWrap(self.companies, benched)
        wrap.start()

    def benchmark(self):
        # Grading process
        # For each metric based on rev_per_share, affo_per_share, nav_per_share, ...
        #   Rank based on above, moderate and below average.
        #       Collect the company for metric and rank level.
        # Accessible via the hash metric: metric[Tag.rev_per_share][RateType.above_avg]

        met = {}
        for k, v in self.metric.items():
            for v1 in v.values():
                if len(v1) > 0:
                    for x in v1:
                        if x.name not in met:
                            met[x.name] = 0
                        met[x.name] += x.score

        final = {RateType.above_avg: [], RateType.moderate_avg: [], RateType.below_avg: []}
        for k, v in met.items():
            if v >= 5.5:
                x = final[RateType.above_avg]
            elif v >= 4.5:
                x = final[RateType.moderate_avg]
            else:
                x = final[RateType.below_avg]
            x.append(k)

        print()
        for k, v in final.items():
            if k is RateType.above_avg:
                print("Companies that performed above average are: ", end='')
            elif k is RateType.moderate_avg:
                print("Companies that performed moderate average are: ", end='')
            else:
                print("Companies which are below average are: ", end='')
            print("{} out of {} sample".format(v, len(v)))
        print("Total companies sampled thus far is {}".format(len(self.companies)))
        return {'comp': final, 'score': met}

    def bucketize(self):
        # TODO namedtuple?

        for x in Tag:
            self.metric[x] = {RateType.above_avg: [], RateType.moderate_avg: [], RateType.below_avg: [], }

        for c in self.companies:
            for k, v in c.prof.items():
                if type(k) is not str:
                    if k in ProfManager.Rate:
                        assert k in self.metric
                        buck = self.metric[k]
                        tup = Bucket(c.name, v['val1'], v['method'])
                        if v['method'] in (ProfMethod.AverageYears, ProfMethod.ReverseRatio):
                            if v['val1'] < ProfManager.Rate[k]['high']:
                                buck[RateType.above_avg].append(tup)
                            elif v['val1'] < ProfManager.Rate[k]['mid']:
                                buck[RateType.moderate_avg].append(tup)
                            else:
                                buck[RateType.below_avg].append(tup)
                        else:
                            if v['val1'] > ProfManager.Rate[k]['high']:
                                buck[RateType.above_avg].append(tup)
                            elif v['val1'] > ProfManager.Rate[k]['mid']:
                                buck[RateType.moderate_avg].append(tup)
                            else:
                                buck[RateType.below_avg].append(tup)

        def value(val):
            return val.value

        def item(key: Bucket):
            return key.name, key.value

        def articulate(buckets: List[Bucket], key):
            values = list(map(value, buckets))
            if len(values) > 0:
                # TODO need to enumerate the index: -
                #  0 == company,
                #  1 == value of percent, years, etc.,
                #  2 == name of method see ProfMethod class
                method = buckets[0].method

                def at(k):
                    if method in ProfManager.UnitRatio:
                        return '{} at {:.2f} yrs'.format(k[0], k[1])
                    return '{} at {:.2f}%'.format(k[0], k[1] * 100)

                items = list(map(item, buckets))
                comp_at_perc = ', '.join(list(map(at, items)))

                unit = Unit(value=100, symbol='%')
                if method in ProfManager.UnitRatio:
                    unit = Unit(value=1, symbol='')

                # TODO modify current "performed above average rate"
                #  to "below the average over the last 10 years sampled, at undemanding rate"
                print("{}/{} companies sampled have performed {} average rate of {} {:.2f}{}. ".format(
                    len(buckets), len(self.companies), RateVerbose[key], ProfVerbose[method],
                    average(values) * unit.value, unit.symbol), end='')
                print("These companies are: {}".format(comp_at_perc))
                if key is RateType.above_avg:
                    # TODO apply() function?
                    for a in buckets:
                        assert a.score is None
                        # Scale it to one per metric, 10 for full points
                        a.score = 10./len(self.metric)
                elif key is RateType.moderate_avg:
                    for a in buckets:
                        assert a.score is None
                        # Scale it to half point per metric, 10 for full points
                        a.score = 5./len(self.metric)
                else:
                    for a in buckets:
                        assert a.score is None
                        a.score = 0

        # TODO need to solve for AFFO, net debt over ebit, retention ratio, div payout ratio

        for k, v in self.metric.items():
            print("Based on {}:".format(k))
            # TODO Near right but not enum hmmm RateType._fields:
            for avg_rate in RateType.above_avg, RateType.moderate_avg, RateType.below_avg:
                articulate(v[avg_rate], avg_rate)

    def simulate_price(self, benched):
        print()

        def _(cat):
            for k, v in benched.items():
                if k in cat:
                    for x in v:
                        assert x in self.company
                        company = self.company[x]
                        if company.last_price is not None:
                            print("{}'s last quote was {}: -".format(x, company.last_price['last_price']))

                            current = company.last_price['ev_over_ebit']
                            record = company.prof[Tag.ev_over_ebit]['val1']
                            diff = (current - record) / record
                            ev_over_ebit_incr = 'undemanding price'\
                                if current < record else 'premium valuation/overpriced'
                            trend = '+' if diff > 0 else ''
                            print("- Last EV over EBIT was {current:.2f},"
                                  " average was {record:.2f}, valued at {ev_over_ebit_incr} after {diff:.2f} pts"
                                  .format(ev_over_ebit_incr=ev_over_ebit_incr, trend=trend,
                                          diff=diff*100, record=record, current=current))

                            # TODO refactored the block below
                            if company.last_price['div_yields'] is not None:
                                div_yields = list(map(lambda z: 0 if z is None else z, company.last_price['div_yields']))
                                avg_div_yield = average(div_yields)
                                if avg_div_yield == 0.:
                                    continue
                                dy_incr = 100 * (company.last_price['last_div_yield'] - avg_div_yield) / avg_div_yield
                                trend = 'upside +' if dy_incr > 0 else 'downside '
                                print("- Last div yield was {:.2f} %, we see last {}{:.2f} pts"
                                      " based on average div yield of {:.2f} % ".format(
                                        company.last_price['last_div_yield']*100, trend, dy_incr, avg_div_yield*100, ))
                            else:
                                print("No dividend was reported.")

        _([RateType.above_avg, RateType.moderate_avg])
        print("\nThe following quotes were rated at below average rating though")
        _([RateType.below_avg])


class WorkWrap:

    start_col = 2
    row_margin = 1

    # TODO differential data, hmmm...
    # i += 1
    # cell = sheet.cell(row=j, column=i)
    # cell.value = 'diff ev_over_ebit'
    #
    # i += 1
    # cell = sheet.cell(row=j, column=i)
    # cell.value = 'incr div yield'

    # setting up color palette following
    gen_rule = ColorScaleRule(start_type='percentile', start_value=10, start_color=ProfManager.Red,
                              mid_type='percentile', mid_value=50, mid_color=ProfManager.Yellow,
                              end_type='percentile', end_value=90, end_color=ProfManager.Green)
    rule = {
        # net_debt_over_ebit
        Tag.net_debt_over_ebit:
            ColorScaleRule(start_type='num', start_value=1., start_color=ProfManager.Green,
                           mid_type='num', mid_value=8., mid_color=ProfManager.Yellow,
                           end_type='num', end_value=10., end_color=ProfManager.Red),
        # EV/EBIT
        Tag.ev_over_ebit:
            ColorScaleRule(start_type='num', start_value=10., start_color=ProfManager.Green,
                           mid_type='num', mid_value=15., mid_color=ProfManager.Yellow,
                           end_type='num', end_value=20., end_color=ProfManager.Red),
        Tag.market_cap_ov_retained_earnings:
            ColorScaleRule(start_type='num', start_value=.5, start_color=ProfManager.Green,
                           mid_type='num', mid_value=.8, mid_color=ProfManager.Yellow,
                           end_type='num', end_value=1.1, end_color=ProfManager.Red),
        # diff EV/EBIT
        'diff-ev_over_ebit':
            ColorScaleRule(start_type='percentile', start_value=90, start_color=ProfManager.Green,
                           mid_type='percentile', mid_value=50, mid_color=ProfManager.Yellow,
                           end_type='percentile', end_value=10, end_color=ProfManager.Red),
        'price_over_affo':
            ColorScaleRule(start_type='num', start_value=0., start_color=ProfManager.Green,
                           mid_type='num', mid_value=10., mid_color=ProfManager.Yellow,
                           end_type='num', end_value=15., end_color=ProfManager.Red),
        # Mid cap is between 200mil to 2bil.
        # https://www.bursamalaysia.com/trade/our_products_services/indices/ftse_bursa_malaysia_indices/overview
        'market_cap':
        # https://coolors.co/palette/07beb8-3dccc7-68d8d6-9ceaef-c4fff9
            ColorScaleRule(start_type='percentile', start_value=90, start_color='c4fff9',
                           mid_type='percentile', mid_value=20, mid_color='9ceaef',
                           end_type='percentile', end_value=10, end_color='3dccc7'),
    }

    def __init__(self, companies: List[Prof], benched: Dict[str, dict]):
        self.companies = companies
        # benched to access 'score' by company's name
        self.benched = benched
        self.ft = Font(name='Calibri', size=11)
        self.wb = Workbook()

        # type: worksheet.Worksheet
        self.sheet = self.wb.active
        self.sheet.title = 'sheet 1'

        self.cell = self.sheet.cell(row=1, column=WorkWrap.start_col)
        self.start_row_index = WorkWrap.row_margin+1
        self.end_row_index = len(self.companies) + self.start_row_index+1

        self.j = WorkWrap.row_margin + 1
        self.i = 2

        self.init_sheet()

    def init_sheet(self):
        # 10, 5 years and current year
        sheet = self.sheet
        cell = self.cell
        cell.value = '10 years'
        cell = self.sheet.cell(row=1, column=WorkWrap.start_col+len(Tag))
        cell.value = '5 years'
        cell = self.sheet.cell(row=1, column=WorkWrap.start_col+2*len(Tag))
        cell.value = 'Current year'

        Tag_to_long = {
            Tag.rev_per_share: 'Sales per share',
            Tag.epu: 'EPS',
            Tag.owner_yield: 'FCF per share',
            Tag.ROIC: 'ROIC',
            Tag.net_debt_over_ebit: 'Net debt /EBIT',
            Tag.ev_over_ebit: 'EV/EBIT',
            Tag.op_margin: 'Op margin',
            Tag.retained_earnings_ratio: 'Retained /Net',
            Tag.market_cap_ov_retained_earnings: 'Market /Retained',
            Tag.dividend_yield: 'Dividend yield',
        }

        # Additional 3 columns.
        for _ in range(1, 4):
            for x in Tag:
                sheet.column_dimensions[colnum_string(self.i)].width = 10
                cell = sheet.cell(row=self.j, column=self.i)
                cell.alignment = Alignment(wrapText=True)
                cell.value = Tag_to_long[x]
                self.i += 1

        # add extension header
        ext_header = ['P', 'Market Cap', 'Revenue', 'Op income', 'Net profit', 'EPU sen', 'FCF ratio',
                      'Retained ratio', 'DPU sen', 'Color']
        for x in ext_header:
            cell = sheet.cell(row=self.j, column=self.i)
            cell.value = x
            cell.alignment = Alignment(wrapText=True)
            self.i += 1

    def start(self):
        self.j += 1
        for c in self.companies:
            # Table of mainly profile and last_price data
            # print("Company", c.name)
            self.i = 1
            cell = self.sheet.cell(row=self.j, column=self.i)
            cell.value = c.name

            self.i += 1
            for ri in range(1, 4):
                self.build_lead(c, ri)
            self.build_suffix(c)

            # adding last column for color
            cell = self.sheet.cell(row=self.j, column=self.i)
            cell.value = self.benched['score'][c.name]
            cell.number_format = '0.0'
            self.sheet.conditional_formatting.add('{alpha}{start}:{alpha}{end}'.format(
                alpha=colnum_string(self.i), start=self.start_row_index, end=self.end_row_index),
                WorkWrap.gen_rule)
            self.i += 1

            self.j += 1
        self.wb.save('output.xlsx')

    def build_lead(self, com: Prof, range_index: int):
        gen_rule = WorkWrap.gen_rule
        rule = WorkWrap.rule
        lead = [
            # Last 10 years metric
            {'val': com.prof[Tag.rev_per_share]['val{}'.format(range_index)], 'rule': gen_rule},
            {'val': com.prof[Tag.epu]['val{}'.format(range_index)], 'rule': gen_rule},
            {'val': com.prof[Tag.owner_yield]['val{}'.format(range_index)], 'number': 'ratio', 'rule': gen_rule},
            # AFFO and Tangible commented diffs
            # {'val': c.prof[Tag.affo_per_share]['val{}'.format(_)], 'rule': gen_rule},
            # {'val': c.prof[Tag.tangible_per_share]['val{}'.format(_)], 'rule': gen_rule},
            # {'val': c.prof[Tag.nav_per_share]['val{}'.format(_)], 'rule': gen_rule},
            {'val': com.prof[Tag.ROIC]['val{}'.format(range_index)], 'rule': gen_rule},
            {'val': com.prof[Tag.net_debt_over_ebit]['val{}'.format(range_index)], 'number': 'ratio',
             'rule': rule[Tag.net_debt_over_ebit]},
            {'val': com.prof[Tag.ev_over_ebit]['val{}'.format(range_index)], 'number': 'ratio',
             'rule': rule[Tag.ev_over_ebit]},
            {'val': com.prof[Tag.op_margin]['val{}'.format(range_index)], 'rule': gen_rule},
            {'val': com.prof[Tag.retained_earnings_ratio]['val{}'.format(range_index)], 'number': 'ratio',
             'rule': gen_rule},
            {'val': com.prof[Tag.market_cap_ov_retained_earnings]['val{}'.format(range_index)], 'number': 'ratio',
             'rule': rule[Tag.market_cap_ov_retained_earnings]},
            {'val': com.prof[Tag.dividend_yield]['val{}'.format(range_index)], 'rule': gen_rule},
        ]

        for x in lead:
            self.build_sheet(x)

    def build_sheet(self, ent: Dict[str, Union[str, float, ColorScaleRule]]):
        # The dict arg is a key of val, rule and number to combo of float, number, color rule tuples

        cell = self.sheet.cell(row=self.j, column=self.i)
        assert 'val' in ent
        if type(ent['val']) is complex:
            # TODO Not a number
            cell.value = 'NaN'
        else:
            cell.value = ent['val']
        if 'number' in ent:
            cell.style = 'Comma'
            if ent['number'] == 'cap':
                if len(str(abs(math.floor(cell.value)))) > 3:
                    cell.number_format = "0,000.00"
                else:
                    cell.number_format = '0.00'
            elif ent['number'] == 'value2':
                cell.number_format = '0.00'
                # TODO test if we got enough decimal points
                # Decimal point can be enabled using DECIMALS TO DISPLAY in TIKR terminal
                # if 0 != decimal.Decimal(c.last_price['last_price']*10**3).as_tuple().exponent:
                #     cell.number_format = '0.000'
            else:
                # set number format to value/ratio
                cell.number_format = '0.00'
        else:
            cell.style = 'Percent'
            cell.number_format = '0.00%'
        cell.font = self.ft

        if 'rule' in ent:
            self.sheet.conditional_formatting.add('{alpha}{start}:{alpha}{end}'.format(
                alpha=colnum_string(self.i), start=self.start_row_index, end=self.end_row_index),
                ent['rule'])
        self.i += 1

    def build_suffix(self, com: Prof):
        rule = WorkWrap.rule

        # Last price data
        suffix = [
            {'val': com.last_price['last_price'], 'number': 'value2'},
            {'val': com.last_price['market_cap'], 'number': 'cap', 'rule': rule['market_cap']},
            {'val': com.last_price['revenue'], 'number': 'cap', 'rule': rule['market_cap']},
            {'val': com.last_price['op_income'], 'number': 'cap', 'rule': rule['market_cap']},
            {'val': com.last_price['net_profit'], 'number': 'cap', 'rule': rule['market_cap']},
            {'val': com.prof[Tag.epu]['val2'], 'number': 'value', 'rule': rule['market_cap']},
            {'val': com.prof[Tag.owner_yield]['val2'], 'number': 'value', 'rule': rule['market_cap']},
            {'val': com.prof[Tag.retained_earnings_ratio]['val2'], 'number': 'value', 'rule': rule['market_cap']},
            # x100 - KLSE/Bursa DPU use fractional pricing model
            {'val': com.last_price['dpu'] * 100, 'number': 'value', 'rule': rule['market_cap']},
            # P/AFFO commented diff
            # {'val': c.last_price['price_over_affo'], 'number': 'value', 'rule': rule['price_over_affo']},
        ]
        for x in suffix:
            self.build_sheet(x)

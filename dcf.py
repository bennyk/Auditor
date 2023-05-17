import tradingview
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, Alignment
from openpyxl.workbook import Workbook
from openpyxl.worksheet import worksheet
import numpy as np
from tabulate import tabulate
from typing import List
from enum import Enum

from spread import Spread
from utils import *
from bcolors import colour_print, bcolors


class Mode(Enum):
    Last_FCF = 1
    Avg_FCF = 2
    UFCF_5x = 3
    UFCF_8x = 4


class ExcelOut:
    Red = 'FF9AA2'
    Yellow = 'FFDAC1'
    # Green = 'E2F0CB'
    Green = 'B5EAD7'

    start_col = 2
    row_margin = 1

    # Setting up color palette following
    gen_rule = ColorScaleRule(start_type='percentile', start_value=10, start_color=Red,
                              mid_type='percentile', mid_value=50, mid_color=Yellow,
                              end_type='percentile', end_value=90, end_color=Green)

    def __init__(self, ticks: [str], entries, styles=None, headers=None):
        cls = self.__class__
        self.entries = entries
        self.headers = headers
        self.ticks = ticks
        self.styles = styles
        self.ft = Font(name='Calibri', size=11)
        self.wb = Workbook()

        # type: worksheet.Worksheet
        self.sheet = self.wb.active
        self.sheet.title = 'sheet 1'

        self.cell = self.sheet.cell(row=1, column=cls.start_col)
        self.start_row_index = cls.row_margin+1
        self.end_row_index = len(self.ticks) + self.start_row_index+1

        self.j = cls.row_margin + 1
        self.i = 1

        self.init_sheet()

    def init_sheet(self):
        sheet = self.sheet
        self.j = 1
        self.i = 2
        for e in self.headers:
            sheet.column_dimensions[colnum_string(self.i)].width = 10.5
            cell = sheet.cell(row=self.j, column=self.i)
            cell.alignment = Alignment(wrapText=True)
            cell.value = e
            self.i += 1
        self.j += 1

    def start(self):
        cls = self.__class__
        self.j = cls.row_margin + 1
        for ent in self.entries:
            # Table of mainly profile and last_price data
            # print("Company", ent)
            self.i = 1
            for i, e in enumerate(ent):
                self.make_cell(e, self.styles[i])

            for i, s in enumerate(self.styles, 1):
                if s == 'Percent':
                    self.sheet.conditional_formatting.add('{alpha}{start}:{alpha}{end}'.format(
                        alpha=colnum_string(i), start=self.start_row_index, end=self.end_row_index),
                        cls.gen_rule)
            self.j += 1
        self.wb.save('dcf_out.xlsx')

    def make_cell(self, e, style):
        cell = self.sheet.cell(row=self.j, column=self.i)
        # sheet.column_dimensions[colnum_string(self.i)].width = 10
        cell.alignment = Alignment(wrapText=True)
        cell.value = e
        if style == 'Comma':
            cell.style = 'Comma'
            cell.number_format = '0.00'
        else:
            cell.style = 'Percent'
            cell.number_format = '0.00%'
        cell.font = self.ft
        self.i += 1


class DCF(Spread):
    def __init__(self, tick, path, modes: [Mode]):
        self.wb = load_workbook(path + '/' + tick + '.xlsx')
        super().__init__(self.wb, tick)

        # TODO approximate WACC based on highest transaction recorded that investors willing to pay
        self.wacc = .1

        # Forecast in 2027
        self.tgr = .0212
        # self.tgr = .0595
        self.term_dr = 1/(self.wacc-self.tgr)
        self.sum_pvtv = []
        self.last_price = None
        self.poss_ratio = None
        self.modes = modes

    def wa_diluted_shares_out(self) -> list:
        shares_out = self.strip(self.income.match_title('Weighted Average Diluted Shares Outstanding'))

        # Find the first item in list that is not None, replace None to last item.
        last_item = next((x for x in shares_out if x is not None))
        result = list(map(lambda x: x if x is not None else last_item, shares_out))
        return result

    def fcf(self):
        # also known as Levered FCF
        earning_not_strip = self.cashflow.match_title('Free Cash Flow$', none_is_optional=True)
        if earning_not_strip is not None:
            earning = self.strip(earning_not_strip)
        else:
            cfo = self.strip(self.cashflow.match_title('Cash from Operations$'))
            opt_acq_real_assets = self.cashflow.match_title('Acquisition of Real Estate Assets$',
                                                            none_is_optional=True)
            earning = cfo
            if opt_acq_real_assets is not None:
                acq_real_assets = self.strip(opt_acq_real_assets)
                earning = list_add_list(cfo, acq_real_assets)
        return list_over_list(earning, self.wa_diluted_shares_out())

    def tv_ev_ov_ebit(self):
        ev_over_ebit = self.strip2(self.values.match_title('LTM Total Enterprise Value / EBIT$'))
        half_len = round(len(ev_over_ebit) / 2.)
        ev_over_ebit_half = ev_over_ebit[half_len:]
        return average(ev_over_ebit_half) * 5.

    # TODO tv_multiple_nopat maybe obsoleted by tv_multiple_nopat_* bind functions
    def tv_multiple_nopat(self, **kwargs):
        ni = self.strip(self.income.match_title('Net Income$'))
        tax = self.strip(self.income.match_title('Income Tax Expense$'))
        # ie = self.strip(self.income.match_title('Interest Expense$'))
        nopat = list_add_list(ni, tax)
        # ebit = list_add_list(ebit, ie)
        ni_per_share = list_over_list(nopat[self.half_len:],
                                      self.wa_diluted_shares_out()[self.half_len:])
        return average(ni_per_share) * 5. * (1+self.tgr) * self.term_dr

        # hmm, not preferred when comparing company. The preferred is EBITDA
        # ebitda = self.strip(self.income.match_title('EBITDA'))
        # ebitda_per_share = list_over_list(ebitda[self.half_len:],
        #                                   self.wa_diluted_shares_out()[self.half_len:])
        # return average(ebitda_per_share)*(1+self.tgr) * self.term_dr

    def tv_last_fcf(self, **kwargs):
        return kwargs['fcf'][-1] * (1+self.tgr) * self.term_dr

    def tv_avg_fcf(self, **kwargs):
        # tv = average(fcf)*(1+tgr) * term_dr
        return average(kwargs['fcf'][self.half_len:]) * (1+self.tgr) * self.term_dr

    def compute_fcf(self):
        print(self.tick)
        self.last_price = self.values.match_title('Price$')[-1]
        print("last price", self.last_price)
        fcf = self.fcf()
        nper = len(fcf)

        # https://www.investopedia.com/terms/d/dcf.asp
        dr = [1/(1+self.wacc) ** (i+1) for i in range(nper)]
        pv = [fcf[i] * dr[i] for i in range(nper)]

        print('nper', nper)
        print('cagr of fcf', '{:.1f}%'.format(cagr(fcf) * 100))
        print('avg of fcf', '{:.1f}%'.format(average(fcf) * 100))
        print('term_dr', '{:.2f}'.format(self.term_dr))
        print('fcf', np.around(fcf, decimals=2))
        print('dr', np.around(dr, decimals=2))
        print('pv', np.around(pv, decimals=2))
        print("sum of pv {:.2f}".format(sum(pv)))

        # https://www.investopedia.com/terms/t/terminalvalue.asp
        tv = []
        last_dr = []
        pvtv = []
        pv_ov_pvtv = []
        self.poss_ratio = []

        # TODO mapping to the underlying implementation
        calc_fcf = list(map(lambda x: {
            Mode.Last_FCF: self.tv_last_fcf,
            Mode.Avg_FCF: self.tv_avg_fcf,
            Mode.UFCF_5x: bind(UFCF(self), tv_multiple_ufcf_5x),
            Mode.UFCF_8x: bind(UFCF(self), tv_multiple_ufcf_8x),
        }[x], self.modes))
        for m in calc_fcf:
            _tv = m(fcf=fcf)
            _dr = 1 / (1+self.wacc) ** nper
            _pvtv = _dr * _tv
            _sum_pvtv = sum(pv) + _pvtv
            _pv_ov_pvtv = sum(pv)/_pvtv if _pvtv != 0 else float('nan')

            # Possible ratio of sum_pvtv vs last price, assuming buying at purchase price
            # https://www.investopedia.com/ask/answers/how-do-you-calculate-percentage-gain-or-loss-investment
            _poss_ratio = (_sum_pvtv-self.last_price) / self.last_price

            tv.append(_tv)
            last_dr.append(_dr)
            pvtv.append(_pvtv)
            pv_ov_pvtv.append(_pv_ov_pvtv)
            self.sum_pvtv.append(_sum_pvtv)
            self.poss_ratio.append(_poss_ratio)

        print('tv\t\t\t', np.around(tv, decimals=2))
        print('dr\t\t\t', np.around(last_dr, decimals=2))
        print('pvtv\t\t', np.around(pvtv, decimals=2))
        print('sum_pvtv\t', np.around(self.sum_pvtv, decimals=2))
        print('poss_ratio\t\t', np.around(self.poss_ratio, decimals=2))
        print('pv_ov_pvtv\t\t', np.around(pv_ov_pvtv, decimals=2))
        print()

        # adbe
        # fcf = [15.6, 18.72, 21.53, 24.76, 28.47, 30.6, 32.9, 35.36, 38.02, 40.87, 708.47]


# https://stackoverflow.com/questions/1015307/python-bind-an-unbound-method
class UFCF:
    def __init__(self, src: DCF):
        # ni = src.strip(src.income.match_title('Net Income$'))
        ufcf = src.strip(src.income.match_title('EBITDA$'))
        tax = src.strip(src.income.match_title('Income Tax Expense$'))
        # ie = self.strip(self.income.match_title('Interest Expense$'))
        ufcf = list_add_list(ufcf, tax)

        capex = src.strip(src.cashflow.match_title('Capital Expenditure$'))
        ufcf = list_add_list(ufcf, capex)

        _ = src.cashflow.match_title(
            'Memo: Change in Net Working Capital$', none_is_optional=True)
        if _ is not None:
            work_cap = src.strip(_)
            ufcf = list_add_list(ufcf, work_cap)
        else:
            colour_print('Missing NWC entry', bcolors.WARNING)

        # Ignore debt issuance/repaid for DCF?
        # ChatGPT:
        # In general, UFCF is more appropriate for comparing the cash-generating ability of different companies,
        # while LFCF is more appropriate for estimating the cash flow available to equity investors after accounting
        # for the company's debt obligations.
        if False:
            # TODO assuming percent debt repaid?
            # debt_issued = src.strip(src.cashflow.match_title('Total Debt Issued$'))
            debt_repaid = src.strip(src.cashflow.match_title('Total Debt Repaid$'))
            # earning = list_add_list(earning, debt_issued)
            ufcf = list_add_list(ufcf, debt_repaid)

        # ebit = list_add_list(ebit, ie)
        self.ufcf_per_share = list_over_list(
            ufcf[src.half_len:], src.wa_diluted_shares_out()[src.half_len:])
        # ufcf, src.wa_diluted_shares_out())
        # TODO option for half_len?
        self.tgr = src.tgr
        self.term_dr = src.term_dr


def bind(self, func, as_name=None):
    """
    Bind the function *func* to *instance*, with either provided name *as_name*
    or the existing name of *func*. The provided *func* should accept the
    instance as the first argument, i.e. "self".
    """
    if as_name is None:
        as_name = func.__name__
    bound_method = func.__get__(self, self.__class__)
    setattr(self, as_name, bound_method)
    return bound_method


# TODO refactor?
def tv_multiple_ufcf_5x(src: UFCF, **kwargs):
    return average(src.ufcf_per_share) * 5. * (1+src.tgr) * src.term_dr


def tv_multiple_ufcf_8x(src: UFCF, **kwargs):
    return average(src.ufcf_per_share) * 8. * (1+src.tgr) * src.term_dr


class Ticks:
    def __init__(self, ticks, path):
        self.tickers = ticks
        self.path = path
        self.modes = [
            Mode.Last_FCF,
            Mode.Avg_FCF,
            Mode.UFCF_5x,
            Mode.UFCF_8x,
        ]

        # type: List[DCF]
        self.spreads = []
        for t in ticks:
            self.spreads.append(DCF(t, self.path, self.modes))

    def compute(self):
        for sp in self.spreads:
            sp.compute_fcf()

    def summarize(self):
        # https://blog.devgenius.io/how-to-easily-print-and-format-tables-in-python-18bbe2e59f5f
        # summarize to TV based on last FCF, avg FCF, multiple of NOPAT

        # Set the first word as the header for our spreadsheet optionally.
        a = list(map(lambda x: [x.head.split()[0] if x.head is not None else x.head,
                                x.tick, x.last_price], self.spreads))
        for i, s in enumerate(self.spreads):
            for j in range(len(s.sum_pvtv)):
                # extend the spread with sum_pvtv and poss ratio
                a[i].append(s.sum_pvtv[j])
                a[i].append(s.poss_ratio[j])

        # Sort it to the last column which is 'Possible' ratio
        entries = sorted(a, key=lambda x: x[len(a[0])-1])
        poss_header = 'Poss. x'
        heads = ['Company', 'Tick', 'Last price']
        for i, m in enumerate(self.modes):
            # TODO mapping to the underlying string
            h = {Mode.Last_FCF: 'Last FCF',
                 Mode.Avg_FCF: 'Avg. FCF',
                 Mode.UFCF_5x: 'UFCF 5x',
                 Mode.UFCF_8x: 'UFCF 8x'}[m]
            heads.extend((h, poss_header))
        print(tabulate(entries, headers=heads,
                       tablefmt='fancy_grid', stralign='center', numalign='center', floatfmt=".2f"))

        # generate Excel output
        styles = ['Comma', 'Comma', 'Comma']
        for _ in self.modes:
            styles.extend(('Comma', 'Percent'))
        excel = ExcelOut(self.tickers, entries, styles=styles, headers=heads)
        excel.start()


if __name__ == '__main__':
    # tickers = tradingview.TradingView().fetch()
    tickers = []
    t = Ticks(tickers, 'spreads')
    t.compute()
    t.summarize()

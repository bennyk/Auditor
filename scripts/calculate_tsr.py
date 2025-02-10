import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.formatting.rule import FormatObject, DataBar, Rule
from openpyxl.utils.dataframe import dataframe_to_rows
from bcolors import colour_print, bcolors
from utils import colnum_string


def load_worksheet(path, name, sheet_name):
    wb = load_workbook(f"{path}/{name}.xlsx")
    return wb, wb[sheet_name]


def extract_data(ws, target_labels):
    max_row, max_column = ws.max_row+1, ws.max_column+1
    data = {label: [] for label in target_labels}

    for i in range(1, max_row):
        row_label = ws[f'A{i}'].value
        if row_label in target_labels:
            for col in range(2, max_column):
                value = ws.cell(row=i, column=col).value
                if row_label == 'Price Close' and value is not None:
                    # Data was extracted from TIKR terminal
                    value = float(re.sub(r'MYR\s+', '', value)) if isinstance(value, str) else value
                data[row_label].append(value)
    return data


def calculate_tsr(header, price_close, dps, years=5):
    purchased_price = price_close[-(years + 1)]
    if purchased_price is None:
        raise IndexError(f"Purchased price was invalid {years+1} years ago")

    tsr_values = {}
    for i in range(years, 0, -1):
        current_price = price_close[-i]
        dividend_paid = dps[-i]
        tsr = (current_price - purchased_price + dividend_paid) / purchased_price
        # tsr_values[f'Year {years-i+1}'] = tsr
        m = re.search(r'(\d+)$', header[-i])
        if m is not None:
            tsr_values[f"'{m.group(1)}"] = tsr
        else:
            tsr_values[f"{header[-i]}"] = tsr

    return tsr_values


def write_tsr_to_excel(out_wb, tsr_values, row, name, years):
    ws = out_wb.active
    ws.title = "TSR Report"

    ws.cell(row=row, column=1).value = name
    j = 2
    for year, tsr in tsr_values.items():
        ws.cell(row=1, column=j).value = year
        ws.cell(row=row, column=j).value = tsr
        ws.cell(row=row, column=j).number_format = FORMAT_PERCENTAGE_00
        j += 1

    # TODO Openpyxl doesn't support conditional formatting yet
    # apply_conditional_formatting(ws, row, years)


def apply_conditional_formatting(ws, row, years):
    """
    Applies conditional formatting with data bars to an Excel file using openpyxl,
    coloring negative percentages with a red bar and positive percentages with a green bar.
    """

    # Determine the range of cells to apply formatting to (excluding headers)
    start_row = 2
    end_row = row+1
    start_col = 1
    end_col = years+1

    for col_idx in range(start_col, end_col + 1):
        col = colnum_string(col_idx)
        cell_range = f"{col}{start_row}:{col}{end_row}"

        # Define data bar formatting rules
        # Define the parameters for the positive data bar
        first_green = FormatObject(type='min')
        second_green = FormatObject(type='max')
        data_bar_green = DataBar(cfvo=[first_green, second_green], color="00FF00", showValue=None, minLength=None, maxLength=None) # Green

        # Create the rule for positive values
        rule_green = Rule(type='dataBar', dataBar=data_bar_green)

        # Apply the rule to the desired range
        ws.conditional_formatting.add(cell_range, rule_green)

        # Define the parameters for the negative data bar
        first_red = FormatObject(type='min')
        second_red = FormatObject(type='max')
        data_bar_red = DataBar(cfvo=[first_red, second_red], color="FF0000", showValue=None, minLength=None, maxLength=None) # Red

        # Create the rule for negative  values
        rule_red = Rule(type='dataBar', dataBar=data_bar_red)

        # Apply the rule to the desired range
        ws.conditional_formatting.add(cell_range, rule_red)


def main():
    out_wb = Workbook()
    path = '../spreads'
    sheet_name = 'Income'
    row = 2
    for name in ['kipreit', 'igbreit', 'klcc', 'sunreit']:
        wb, ws = load_worksheet(path, name, sheet_name)
        print(f"Loaded worksheet: {name}")
        data = extract_data(ws, ['Income Statement | TIKR.com', 'Dividends per share', 'Price Close'])
        if 'Dividends per share' in data and 'Price Close' in data:
            # years = 4
            # years = 5
            years = 9
            try:
                tsr_values = calculate_tsr(
                    data['Income Statement | TIKR.com'],
                    data['Price Close'],
                    data['Dividends per share'],
                    years=years, )

                for year, tsr in tsr_values.items():
                    print(f"{year}: TSR = {tsr:.2%}")

                write_tsr_to_excel(out_wb, tsr_values, row=row, name=name, years=years)
                row += 1
                print("TSR values written to Excel successfully.")
            except IndexError as msg:
                colour_print(f"Skipping: {msg}", bcolors.WARNING)
    out_wb.save(f"tsr_report.xlsx")


if __name__ == "__main__":
    main()

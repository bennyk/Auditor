import re
from typing import List

from openpyxl import load_workbook, Workbook
from openpyxl.styles.numbers import FORMAT_NUMBER_00, FORMAT_PERCENTAGE_00
from utils import colnum_string, list_over_list

path = '../spreads'


class CAGR:
    def __init__(self, ws, j, start_idx):
        self.ws = ws
        self.j = j
        self.start_idx = start_idx

    def calculate(self, idx):
        self.ws.cell(row=idx, column=self.j).value = \
            (f"=(({colnum_string(self.j-1)}{idx}/{colnum_string(self.start_idx)}{idx})"
             f"^(1/COLUMNS({colnum_string(self.start_idx)}{idx}:{colnum_string(self.j-1)}{idx}))-1)")
        self.ws.cell(row=idx, column=self.j).number_format = FORMAT_PERCENTAGE_00


class ExcelSheet:
    def __init__(self):
        self.data = {}
        self.row_idx = 0
        pass

    def extract_data(self, ws, target_labels, tag):
        max_row, max_column = ws.max_row + 1, ws.max_column + 1
        data = {label: [] for label in target_labels}

        for i in range(1, max_row):
            row_label = ws[f'A{i}'].value
            if row_label in target_labels:
                for col in range(2, max_column):
                    value = ws.cell(row=i, column=col).value
                    if row_label == 'Price Close' and value is not None:
                        # Data was extracted from TIKR terminal
                        value = float(re.sub(r'MYR\s+', '', value)) if isinstance(value, str) else value
                    data[row_label].append(value)
        self.data[tag] = data
        pass

    def parse_statement(self, name):
        wb = load_workbook(f"{path}/{name}.xlsx")
        income_sheet = wb['Income']
        balance_sheet = wb['Balance']
        cash_flow = wb['Cash']

        for row in range(1, income_sheet.max_row+1):
            self.extract_data(income_sheet, ['Income Statement | TIKR.com',
                                             r'Total Revenues',
                                             r'Operating Income',
                                             r'Net Income',
                                             r'EBT Excl. Unusual Items',
                                             r'Weighted Average Diluted Shares Outstanding',
                                             r'Market Cap',
                                             r'Price Close',
                                             r'Dividends per share',
                                             ], 'Income')

        for row in range(1, cash_flow.max_row+1):
            self.extract_data(cash_flow, ['Cash Flow Statement | TIKR.com',
                                          r'Net Income',
                                          r'Total Depreciation, Depletion & Amortization',
                                          r'Total Asset Writedown',
                                          r'Provision and Write-off of Bad Debts',
                                          r'Acquisition of Real Estate Assets',
                                          ], 'Cash')

        for row in range(1, balance_sheet.max_row + 1):
            self.extract_data(balance_sheet, ['Balance Sheet | TIKR.com',
                                              r'Total Real Estate Assets',
                                              r'Total Debt',
                                              r'Total Assets',
                                              ], "Balance")
        pass

    def parse_header_year(self, head):
        tikr_header = None
        if head == "Income":
            tikr_header = 'Income Statement | TIKR.com'
        elif head == "Balance":
            tikr_header = 'Balance Sheet | TIKR.com'
        elif head == "Cash":
            tikr_header = 'Cash Flow Statement | TIKR.com'
        else:
            raise "Invalid header"

        m = re.search(r'(\d{2})$', self.data[head][tikr_header][0])
        assert m is not None
        return int(m.group(1))

    def add_row_idx(self):
        self.row_idx += 1
        return self.row_idx

    def calculate_cagr(self, j, start_idx, idx, ws):
        ws.cell(row=idx, column=j).value = \
            (f"=(({colnum_string(j-1)}{idx}/{colnum_string(start_idx)}{idx})"
             f"^(1/COLUMNS({colnum_string(start_idx)}{idx}:{colnum_string(j-1)}{idx}))-1)")
        ws.cell(row=idx, column=j).number_format = FORMAT_PERCENTAGE_00

    def write(self, out_wb, title):
        ws = out_wb.create_sheet(title)
        self.row_idx = 0

        WADS = r'Weighted Average Diluted Shares Outstanding'
        ws.column_dimensions["A"].width = 25

        j = 2
        income_items_idx = self.add_row_idx()
        total_revenues_idx = self.add_row_idx()
        revenues_growth_idx = self.add_row_idx()
        op_income_idx = self.add_row_idx()
        op_income_changes_idx = self.add_row_idx()
        op_margin_idx = self.add_row_idx()
        net_income_idx = self.add_row_idx()
        adj_net_income_idx = self.add_row_idx()
        adj_net_margin_idx = self.add_row_idx()
        epu_idx = self.add_row_idx()
        epu_sen_idx = self.add_row_idx()
        epu_sen_growth_idx = self.add_row_idx()
        market_cap_idx = self.add_row_idx()
        price_close_idx = self.add_row_idx()
        shares_outstanding_idx = self.add_row_idx()
        shares_outstanding_growth_idx = self.add_row_idx()
        per_idx = self.add_row_idx()
        dps_idx = self.add_row_idx()
        dps_sen_idx = self.add_row_idx()
        div_payout_idx = self.add_row_idx()
        div_yield_idx = self.add_row_idx()
        tsr_per_idx = self.add_row_idx()
        yield_on_cost = self.add_row_idx()

        ws.cell(row=income_items_idx, column=1).value = "Income items / end of year"
        ws.cell(row=total_revenues_idx, column=1).value = "Total sales"
        ws.cell(row=revenues_growth_idx, column=1).value = "  % Sales change YoY"
        ws.cell(row=op_income_idx, column=1).value = "Operating Income"
        ws.cell(row=op_income_changes_idx, column=1).value = "  % Change YoY"
        ws.cell(row=op_margin_idx, column=1).value = "  % Op. Margins"
        ws.cell(row=net_income_idx, column=1).value = "Net income"
        ws.cell(row=adj_net_income_idx, column=1).value = "Adj. net income"
        ws.cell(row=adj_net_margin_idx, column=1).value = "  % Adj. net margin"
        ws.cell(row=epu_idx, column=1).value = "Adj. EPU"
        ws.cell(row=epu_sen_idx, column=1).value = "Adj. EPU (sen)"
        ws.cell(row=epu_sen_growth_idx, column=1).value = "  % EPU change YoY"
        ws.cell(row=market_cap_idx, column=1).value = "Market Cap"
        ws.cell(row=price_close_idx, column=1).value = "Price Close"
        ws.cell(row=shares_outstanding_idx, column=1).value = "Shares outstanding"
        ws.cell(row=shares_outstanding_growth_idx, column=1).value = "  % Change YoY"
        ws.cell(row=per_idx, column=1).value = "PER"
        ws.cell(row=dps_idx, column=1).value = "Dividends per share"
        ws.cell(row=dps_sen_idx, column=1).value = "Dividends per share (sen)"
        ws.cell(row=div_payout_idx, column=1).value = "  % Payout"
        ws.cell(row=div_yield_idx, column=1).value = "  % Dividends yield"
        ws.cell(row=tsr_per_idx, column=1).value = "  % TSR"
        ws.cell(row=yield_on_cost, column=1).value = "  % YoC"

        first_ffo_col = None
        data = self.data["Income"]
        for i in range(len(data[r'Income Statement | TIKR.com'])):
            ws.cell(row=income_items_idx, column=j).value = data[r'Income Statement | TIKR.com'][i]

            ws.cell(row=total_revenues_idx, column=j).value = data[r'Total Revenues'][i]
            ws.cell(row=total_revenues_idx, column=j).number_format = FORMAT_NUMBER_00

            if i > 0 and data[r'Total Revenues'][i-1] is not None:
                ws.cell(row=revenues_growth_idx, column=j).value =\
                    f"={colnum_string(j)}{total_revenues_idx}/{colnum_string(j-1)}{total_revenues_idx}-1"
                ws.cell(row=revenues_growth_idx, column=j).number_format = FORMAT_PERCENTAGE_00

            ws.cell(row=op_income_idx, column=j).value = data[r'Operating Income'][i]
            ws.cell(row=op_income_idx, column=j).number_format = FORMAT_NUMBER_00

            if i >= 1:
                ws.cell(row=op_income_changes_idx, column=j).value =\
                    f"={colnum_string(j)}{op_income_idx}/{colnum_string(j-1)}{op_income_idx}-1"
                ws.cell(row=op_income_changes_idx, column=j).number_format = FORMAT_PERCENTAGE_00

            ws.cell(row=op_margin_idx, column=j).value = f"={colnum_string(j)}{op_income_idx}/{colnum_string(j)}{total_revenues_idx}"
            ws.cell(row=op_margin_idx, column=j).number_format = FORMAT_PERCENTAGE_00

            ws.cell(row=net_income_idx, column=j).value = data[r'Net Income'][i]
            ws.cell(row=net_income_idx, column=j).number_format = FORMAT_NUMBER_00

            ws.cell(row=adj_net_income_idx, column=j).value = data[r'EBT Excl. Unusual Items'][i]
            ws.cell(row=adj_net_income_idx, column=j).number_format = FORMAT_NUMBER_00

            ws.cell(row=adj_net_margin_idx, column=j).value = \
                f"={colnum_string(j)}{adj_net_income_idx}/{colnum_string(j)}{total_revenues_idx}"
            ws.cell(row=adj_net_margin_idx, column=j).number_format = FORMAT_PERCENTAGE_00

            if data[WADS][i] is not None:
                ws.cell(row=epu_idx, column=j).value = f"={colnum_string(j)}{adj_net_income_idx}/{colnum_string(j)}{shares_outstanding_idx}"
                ws.cell(row=epu_idx, column=j).number_format = '0.0000'

                ws.cell(row=epu_sen_idx, column=j).value = f"=100*{colnum_string(j)}{epu_idx}"
                ws.cell(row=epu_sen_idx, column=j).number_format = FORMAT_NUMBER_00

                if i > 0 and data[WADS][i-1] is not None:
                    ws.cell(row=epu_sen_growth_idx, column=j).value = \
                        f"={colnum_string(j)}{epu_sen_idx}/{colnum_string(j-1)}{epu_sen_idx}-1"
                    ws.cell(row=epu_sen_growth_idx, column=j).number_format = FORMAT_PERCENTAGE_00

            ws.cell(row=market_cap_idx, column=j).value = data[r'Market Cap'][i]
            ws.cell(row=market_cap_idx, column=j).number_format = FORMAT_NUMBER_00

            ws.cell(row=price_close_idx, column=j).value = data[r'Price Close'][i]
            ws.cell(row=price_close_idx, column=j).number_format = FORMAT_NUMBER_00

            ws.cell(row=shares_outstanding_idx, column=j).value = data[WADS][i]
            ws.cell(row=shares_outstanding_idx, column=j).number_format = FORMAT_NUMBER_00

            if i > 1 and data[WADS][i-1] is not None and data[WADS][i-1] != data[WADS][i]:
                ws.cell(row=shares_outstanding_growth_idx, column=j).value =\
                    f"=({colnum_string(j)}{shares_outstanding_idx}/{colnum_string(j-1)}{shares_outstanding_idx}-1)"
                ws.cell(row=shares_outstanding_growth_idx, column=j).number_format = FORMAT_PERCENTAGE_00

            if data[WADS][i] is not None:
                ws.cell(row=per_idx, column=j).value = f"={colnum_string(j)}{price_close_idx}/{colnum_string(j)}{epu_idx}"
                ws.cell(row=per_idx, column=j).number_format = FORMAT_NUMBER_00

            ws.cell(row=dps_idx, column=j).value = data[r'Dividends per share'][i] if data[r'Dividends per share'][i] is not None else ''
            ws.cell(row=dps_idx, column=j).number_format = '0.0000'

            ws.cell(row=dps_sen_idx, column=j).value = f"=100*{colnum_string(j)}{dps_idx}"
            ws.cell(row=dps_sen_idx, column=j).number_format = FORMAT_NUMBER_00

            if data[WADS][i] is not None:
                ws.cell(row=div_payout_idx, column=j).value =\
                    f"={colnum_string(j)}{dps_idx}/{colnum_string(j)}{epu_idx}"
                ws.cell(row=div_payout_idx, column=j).number_format = FORMAT_PERCENTAGE_00

                ws.cell(row=div_yield_idx, column=j).value = f"={colnum_string(j)}{dps_idx}/{colnum_string(j)}{price_close_idx}"
                ws.cell(row=div_yield_idx, column=j).number_format = FORMAT_PERCENTAGE_00

                if first_ffo_col is None:
                    first_ffo_col = j
                else:
                    ws.cell(row=tsr_per_idx, column=j).value =\
                        (f"=({colnum_string(j)}{price_close_idx} - {colnum_string(first_ffo_col)}{price_close_idx}"
                         f"+ {colnum_string(j)}{dps_idx}) / {colnum_string(first_ffo_col)}{price_close_idx}")
                    ws.cell(row=tsr_per_idx, column=j).number_format = FORMAT_PERCENTAGE_00

                # print(f"{self.data['Balance']['Total Real Estate Assets'][i-2]}")
                ws.cell(row=yield_on_cost, column=j).value =\
                    f"={colnum_string(j)}{op_income_idx}/{self.data['Balance']['Total Real Estate Assets'][i-2]}"
                ws.cell(row=yield_on_cost, column=j).number_format = FORMAT_PERCENTAGE_00
            j += 1

        # CAGR return
        # Close value over initial value power to 1/periods. Formulaic below
        #   =(close value/initial value)^(1/COLUMNS(initial:close))-1
        # TODO temp fix for short term than 10 years
        # start_idx = j-11
        start_idx = j-9
        ws.cell(row=1, column=j).value = "CAGR"

        cagr_cal = CAGR(ws, j, start_idx)
        cagr_cal.calculate(total_revenues_idx)
        cagr_cal.calculate(adj_net_income_idx)
        cagr_cal.calculate(epu_idx)
        cagr_cal.calculate(dps_idx)

        # Income header year minus Cash header year. Years offset adjustment to IPO since pre listing.
        a = self.parse_header_year("Income")
        b = self.parse_header_year("Cash")
        j = 2 + b - a

        self.row_idx += 1
        cash_items_idx = self.add_row_idx()
        ffo_idx = self.add_row_idx()
        ffo_per_share_idx = self.add_row_idx()
        p_over_ffo_idx = self.add_row_idx()
        ffo_return_idx = self.add_row_idx()

        self.row_idx += 1
        affo_idx = self.add_row_idx()
        affo_per_share_idx = self.add_row_idx()
        p_over_affo_idx = self.add_row_idx()
        affo_return_idx = self.add_row_idx()

        ws.cell(row=cash_items_idx, column=1).value = "Cash items / end of year"
        ws.cell(row=ffo_idx, column=1).value = "FFO"
        ws.cell(row=ffo_per_share_idx, column=1).value = "FFO per share"
        ws.cell(row=p_over_ffo_idx, column=1).value = "P/FFO per share"
        ws.cell(row=ffo_return_idx, column=1).value = " % FFO per share return"

        ws.cell(row=affo_idx, column=1).value = "AFFO"
        ws.cell(row=affo_per_share_idx, column=1).value = "AFFO per share"
        ws.cell(row=p_over_affo_idx, column=1).value = "P/AFFO per share"
        ws.cell(row=affo_return_idx, column=1).value = "  % AFFO per share return"

        first_ffo_col = None
        first_affo_col = None
        data = self.data["Cash"]
        for i in range(len(data[r'Cash Flow Statement | TIKR.com'])):
            ws.cell(row=cash_items_idx, column=j).value = data[r'Cash Flow Statement | TIKR.com'][i]

            ffo = data['Net Income'][i]
            ffo += data['Total Depreciation, Depletion & Amortization'][i]
            if len(data['Total Asset Writedown']) > 0:
                if data['Total Asset Writedown'][i] is not None:
                    ffo += data['Total Asset Writedown'][i]

                if data['Provision and Write-off of Bad Debts'][i] is not None:
                    ffo += data['Provision and Write-off of Bad Debts'][i]

            ws.cell(row=ffo_idx, column=j).value = ffo
            ws.cell(row=ffo_idx, column=j).number_format = FORMAT_NUMBER_00

            ws.cell(row=ffo_per_share_idx, column=j).value = \
                f"={colnum_string(j)}{ffo_idx} / {colnum_string(j)}{shares_outstanding_idx}"
            ws.cell(row=ffo_per_share_idx, column=j).number_format = '0.0000'

            ws.cell(row=p_over_ffo_idx, column=j).value = f"={colnum_string(j)}{price_close_idx} / {colnum_string(j)}{ffo_per_share_idx}"
            ws.cell(row=p_over_ffo_idx, column=j).number_format = FORMAT_NUMBER_00

            if first_ffo_col is None:
                first_ffo_col = j
            else:
                ws.cell(row=ffo_return_idx, column=j).value =\
                    f"={colnum_string(j)}{ffo_per_share_idx} / {colnum_string(first_ffo_col)}{ffo_per_share_idx}-1"
                ws.cell(row=ffo_return_idx, column=j).number_format = FORMAT_PERCENTAGE_00

            affo = ffo
            affo += data['Acquisition of Real Estate Assets'][i]
            ws.cell(row=affo_idx, column=j).value = affo
            ws.cell(row=affo_idx, column=j).number_format = FORMAT_NUMBER_00

            ws.cell(row=affo_per_share_idx, column=j).value =\
                f"={colnum_string(j)}{affo_idx} / {colnum_string(j)}{shares_outstanding_idx}"
            ws.cell(row=affo_per_share_idx, column=j).number_format = '0.0000'

            ws.cell(row=p_over_affo_idx, column=j).value =\
                f"={colnum_string(j)}{price_close_idx} / {colnum_string(j)}{affo_per_share_idx}"
            ws.cell(row=p_over_affo_idx, column=j).number_format = FORMAT_NUMBER_00

            if first_affo_col is None:
                first_affo_col = j
            else:
                ws.cell(row=affo_return_idx, column=j).value = \
                    f"={colnum_string(j)}{affo_per_share_idx} / {colnum_string(first_affo_col)}{affo_per_share_idx}-1"
                ws.cell(row=affo_return_idx, column=j).number_format = FORMAT_PERCENTAGE_00
            j += 1

        a = self.parse_header_year("Income")
        b = self.parse_header_year("Balance")
        j = 2 + b - a
        self.row_idx += 1
        balance_items_idx = self.add_row_idx()
        debt_to_assets_idx = self.add_row_idx()

        ws.cell(row=balance_items_idx, column=1).value = "Balance items / end of year"
        ws.cell(row=debt_to_assets_idx, column=1).value = "Debt to Assets %"
        data = self.data["Balance"]
        for i in range(len(data[r'Balance Sheet | TIKR.com'])):
            ws.cell(row=balance_items_idx, column=j).value = data[r'Balance Sheet | TIKR.com'][i]

            ws.cell(row=debt_to_assets_idx, column=j).value = data["Total Debt"][i] / data["Total Assets"][i]
            ws.cell(row=debt_to_assets_idx, column=j).number_format = FORMAT_PERCENTAGE_00
            j += 1


def main():
    sheet = ExcelSheet()
    out_wb = Workbook()
    del out_wb['Sheet']

    for name in ['kipreit', 'igbreit', 'klcc', 'sunreit', 'axreit']:
        print(name)
        sheet.parse_statement(name)
        sheet.write(out_wb, name)
    out_wb.save(f"xyz_report.xlsx")


if __name__ == "__main__":
    main()


# KIPREIT TODOs
# 1. Add line revenue growth % and EPU growth %
# 2. P/FFO and P/AFFO
# 3. Pick up end-of-year items when tabulating shares outstanding (WADSO).
#    TIKR uses an average of outstanding shares instead, leading to a pessimistic/overvaluation when trading.
# 4. Does the calculated DPU (1.544 sen) not match the reported 1.66 sen?
# 5. YoC and % payout
# 6. Use original EPS



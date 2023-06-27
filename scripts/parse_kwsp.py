from datetime import datetime, timedelta
from openpyxl.workbook import Workbook
from openpyxl.worksheet import worksheet
import numpy as np
from enum import Enum

from utils import *

class KWSP_Mode(Enum):
    Current = 2
    Long = 1


class Work:
    def __init__(self, mode):
        self.result = {}
        self.start_column = 1
        self.start_offset = self.start_column+1
        self.wb = Workbook()

        # End year will be updated at complete parsing
        self.end_year = None
        self.start_year = {KWSP_Mode.Current: 2013, KWSP_Mode.Long: 1960}[mode]

    def start(self):
        start_date = datetime(self.start_year, 1, 1)
        date = start_date
        sheet = self.wb.active
        sheet.cell(row=1, column=self.start_column, value='Date')
        sheet.cell(row=1, column=self.start_offset, value='KWSP')
        _ = self.fill(date)
        self.end_year = _.year

        sheet.cell(row=1, column=self.start_offset+1, value='MY Inflation')
        self.fill(date, add_col=1)

    def fill(self, date, add_col=0):
        start_ordinal = date.toordinal()
        current_date = datetime.now()
        sheet = self.wb.active
        while date <= current_date:
            i = date.toordinal() - start_ordinal + 2
            if add_col == 0:
                cell = sheet.cell(row=i, column=self.start_column)
                cell.value = date
                cell.number_format = 'DD/MM/YY'
                self.result[date.toordinal()] = {'cell': cell}

            # inserting initial percent
            offset = self.start_offset+add_col
            cell = sheet.cell(row=i, column=offset)
            if i > 2:
                cell.value = "={}{}".format(colnum_string(offset), i-1)
                cell.number_format = '0.00%'
            date += timedelta(days=1)
        return date


class HistoricalChart(Work):
    def __init__(self, mode):
        super().__init__(mode)
        self.mode = mode
        self.div_rate = []
        self.filepath = {KWSP_Mode.Current: 'parsed_kwsp_div',
                         KWSP_Mode.Long: 'parsed_kwsp_div_long'}[mode] + '.xlsx'

    def start(self):
        super().start()
        self.parse_kwsp()
        self.parse_inflation()
        if self.mode == KWSP_Mode.Current:
            self.parse_wsj()
        self.save()

    def parse_kwsp(self):
        with open("data/kwsp-dividend-rates-raw.txt", 'r') as f:
            a = None
            extrapolate = None
            while True:
                line = f.readline()
                if not line:
                    break
                if re.match(r'19|20', line):
                    if a is not None:
                        self.div_rate.append(a)
                        if extrapolate is not None:
                            for i in range(extrapolate):
                                self.div_rate.append(a)
                    extrapolate = None
                    a = [line.strip()]
                    m = re.match(r'(\d+)\s*[-–]\s*(\d+)', line)
                    if m is not None:
                        extrapolate = int(m.group(2)) - int(m.group(1))
                    continue

                if re.match(r'Dividend|Year', line):
                    if a is not None:
                        self.div_rate.append(a)
                    a = [line.strip()]
                    continue

                if re.match(r'[-–]|(?:\d+(?:(\.\d+))?)$', line):
                    if re.match(r'[-–]$', line):
                        a.append(0)
                    else:
                        a.append(float(line.strip()) / 100.0)
                else:
                    if re.match(r'[()\w]+', line):
                        a.append(line.strip())
                # print(line, end='')

        sheet = self.wb.active
        striped_div_rate = self.div_rate[2:]
        s = 1.0
        div_span = {KWSP_Mode.Current: striped_div_rate[:10],
                    KWSP_Mode.Long: striped_div_rate}[self.mode]
        for x in self.cal_div_rate(div_span):
            # Dividend usually announce in March averagely.
            c = self.result[datetime(x[0], 3, 1).toordinal()]['cell']
            # cell = sheet.cell(row=c.row, column=c.column+offset)
            cell = sheet.cell(row=c.row, column=self.start_offset)
            s += float(x[1])
            cell.value = s-1
            cell.number_format = '0.00%'
            print(s, x)

    @staticmethod
    def cal_div_rate(d_rate):
        # print(div_rate)
        # starting year in the reversed order as parsed.
        title = d_rate[-1][0].split()[0]
        start_year = int(title)
        print("starting year", start_year)
        print("end year", d_rate[0][0])

        # strip out the headers including EPF to Shariah owners.
        _ = list(map(lambda x: x[2], d_rate))
        _.reverse()
        print("div rate", np.around(_, decimals=4))
        num_years = len(_)
        print("num of years", num_years)
        s = 1.0
        for i in range(len(_)):
            s += _[i]
            yield start_year+i, _[i]

        price_return = HistoricalChart.cagr_price_return(s, 1., num_years)
        print("cagr {:.2f}%".format(price_return*100))
        print("number of multiple {:.1f}x".format(s-1))
        print()

    def parse_inflation(self):
        def compute_historical(file, index):
            sheet = self.wb.active

            Open = 1
            Close = 1
            lines = file.readlines()
            init_open = None
            last_closed = .0
            for line in lines[1:]:
                a = line.strip().split(',')
                # print(a)
                parsed_date = datetime.strptime(a[0], "%Y-%m-%d")
                if parsed_date.toordinal() in self.result:
                    c = self.result[parsed_date.toordinal()]['cell']
                    # type: worksheet.Worksheet
                    cell = sheet.cell(row=c.row, column=c.column + index)
                    # TODO Inflation-adjusted return = (1 + Stock Return) / (1 + Inflation) - 1
                    if init_open is None:
                        init_open = float(a[Open])/100
                    current_close = float(a[Close])/100
                    last_closed += current_close
                    cell.value = last_closed
                    # cell.style = 'Percent'
                    cell.number_format = '0.00%'
                    print(parsed_date, current_close, last_closed)
                    pass
            price_return = HistoricalChart.cagr_price_return(last_closed, init_open,
                                                             self.end_year-self.start_year)
            print("cagr {:.2f}%".format(price_return * 100))
            print("number of multiple {:.1f}x".format(last_closed/init_open))

        fname = "malaysia inflation max FPCPITOTLZGMYS.csv"
        path = "data/{}".format(fname)
        with open(path, 'r') as f:
            print(fname)
            compute_historical(f, 2)
            print()

    def parse_wsj(self):
        def compute_historical(file, index):
            sheet = self.wb.active
            name = file.name.split('.csv')[0].split('-')[1].upper()
            add_col = 1
            cell = sheet.cell(row=1, column=add_col+index+1)
            cell.value = name

            Open = 1
            Close = 4
            lines = file.readlines()
            init_open = float(lines[1:][-1].strip().split(', ')[Open])
            last_closed = None
            for line in reversed(lines[1:]):
                a = line.strip().split(', ')
                # print(a)
                try:
                    parsed_date = datetime.strptime(a[0], "%m/%d/%y")
                    if parsed_date.toordinal() in self.result:
                        c = self.result[parsed_date.toordinal()]['cell']
                        # type: worksheet.Worksheet
                        cell = sheet.cell(row=c.row, column=c.column+index+add_col)
                        cell.value = float(a[Close]) / init_open - 1
                        # cell.style = 'Percent'
                        cell.number_format = '0.00%'
                        last_closed = float(a[Close])
                except ValueError:
                    pass
            price_return = HistoricalChart.cagr_price_return(last_closed, init_open,
                                                             self.end_year-self.start_year)
            print("cagr {:.2f}%".format(price_return * 100))
            print("number of multiple {:.1f}x".format(last_closed/init_open))

        for i, fname in enumerate([
            'wsj-klci-historical.csv',
            'wsj-malaysia-treasury10y-historical.csv',
            'wsj-sgol-historical.csv',
            'wsj-treasury10y-historical.csv',
            'wsj-s&p500-historical.csv',
            'wsj-brk.b-historical.csv',
            'wsj-soxx-historical.csv',
            'wsj-aapl-historical.csv',
            'wsj-tsla-historical.csv',
            'wsj-nvda-historical.csv',
            'wsj-intc-historical.csv',
        ], 2):
            path = "data/{}".format(fname)
            with open(path) as f:
                print(fname)
                compute_historical(f, i)
                print()

    def save(self):
        self.wb.save(self.filepath)

    @staticmethod
    def cagr_price_return(last_closed, init_open, num_years):
        price_return = (last_closed/init_open-1.)
        return (1+price_return)**(1/num_years)-1.


if __name__ == '__main__':
    for m in [KWSP_Mode.Long, KWSP_Mode.Current]:
        work = HistoricalChart(m)
        work.start()

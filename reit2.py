from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.worksheet import worksheet
# from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from typing import List, Tuple
from collections import OrderedDict, namedtuple
from enum import Enum
import re
import decimal
import statistics

max_row = max_col = 99

Unit = namedtuple('Unit', ['value', 'symbol'])
RateType = namedtuple('Rate', ['above_avg', 'moderate_avg', 'below_avg'])

RateVerbose = {RateType.above_avg: 'above',
               RateType.moderate_avg: 'moderately at',

               # TODO "do not perform" below level
               RateType.below_avg: 'below'}


def colnum_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string


# TODO skipped first element in the list
def striped_average(l: [float]):
    l = strip(l)
    return sum(l) / len(l)


def average(l: [float]):
    return sum(filter(None, l)) / len(l)


def strip(l, trim_last=False):
    if trim_last:
        return l[1:][:-1]
    return l[1:]


def strip2(l, trim_last=False):
    # TODO need to check against original strip implementation.
    if trim_last:
        return l[1:][:-1]

    _ = l[1:]
    if type(_[0]) is str:
        return list(map(lambda x: float(re.sub(r'(\d+)x', r'\1', x)), _))
    return _


def list_over_list(x, y, percent=False):
    if percent:
        # return list(map(lambda n1, n2: (n1 / n2), x, y))
        return list(map(lambda n1, n2: 0 if n1 is None else 100 * (n1 / n2), x, y))
    return list(map(lambda n1, n2: 0 if n1 is None else n1 / n2, x, y))


def list_add_list(x, y):
    a = map(lambda n1: 0 if n1 is None else n1, x)
    b = map(lambda n2: 0 if n2 is None else n2, y)
    return list(map(lambda n1, n2: n1 + n2, a, b))


def list_minus_list(x, y):
    a = map(lambda n1: 0 if n1 is None else n1, x)
    b = map(lambda n2: 0 if n2 is None else n2, y)
    return list(map(lambda n1, n2: n1 - n2, a, b))


def cagr(l: [float]) -> float:
    if l[len(l)-1] < 0:
        p = abs(l[len(l) - 1])
        print("Warning: experimenting with patched number in CAGR")
        return (p / (p+l[0])) ** (1 / len(l)) - 1

    if l[0] < 0:
        p = abs(l[0])
        print("Warning: experimenting with patched number in CAGR")
        return (p + l[len(l)-1] / p) ** (1 / len(l)) - 1

    return (l[len(l) - 1] / l[0]) ** (1 / len(l)) - 1


class Table:
    col_limit = 0

    def __init__(self, sheet_ranges):
        self.date_range = []
        last_limit = 0
        try:
            # configure spreadsheet based on number of cols.
            for j in range(2, max_col):
                c0 = "{}{}".format(colnum_string(j), 1)
                self.date_range.append(sheet_ranges[c0].value)
                if re.match(r'LTM$', sheet_ranges[c0].value):
                    Table.col_limit = j + 1
                    break
                last_limit = j+1
        except TypeError:
            c0 = "{}{}".format(colnum_string(last_limit), 1)
            if sheet_ranges[c0].value is None:
                Table.col_limit = last_limit
            else:
                Table.col_limit = last_limit+1

        self.tab = []
        for i in range(1, max_row):
            c0 = "{}{}".format(colnum_string(1), i)
            if sheet_ranges[c0].value is None:
                break
            r = []
            for j in range(1, Table.col_limit):
                c1 = "{}{}".format(colnum_string(j), i)
                r.append(sheet_ranges[c1].value)
            self.tab.append(r)

    def match_title(self, reg, none_is_optional=False):
        result = None
        for _ in self.tab:
            if re.match(reg, _[0].strip()):
                result = _
                break
        if not none_is_optional:
            assert result is not None
        return result


class Spread:
    def __init__(self, wb, tick, prof: 'Prof'):
        self.tick = tick
        self.profiler = prof
        self.tabs = []
        self.income = None
        self.balance = None
        self.cashflow = None
        self.values = None
        for index, name in enumerate(wb.sheetnames):
            tab = Table(wb[name])
            self.tabs.append(tab)
            if re.match(r'Income', name):
                self.income = tab
                self.start_date = self.income.tab[0][1]
                # -2 ignore year LTM
                self.end_date = self.income.tab[0][-2]
            elif re.match(r'Balance', name):
                self.balance = tab
            elif re.match(r'Cash', name):
                self.cashflow = tab
            elif re.match(r'Values', name):
                self.values = tab
            else:
                # passing Ratios
                pass

        self.end_year = int(self.end_date.split('/')[-1])
        self.start_year = int(self.start_date.split('/')[-1])
        print("Sampled from {} to {} in {} years".format(
            self.start_date, self.end_date,
            1+self.end_year-self.start_year))

    def revenue(self):
        revs = strip(self.income.match_title('Total Revenues'))
        rev_per_share = list(map(lambda f: f / self.share_out_filing(), revs))
        cagr_rev_per_share = cagr(rev_per_share)
        print("Revenue per share from {} to {} at CAGR {:.2f}% for: {}".format(
            revs[0], revs[-1],
            cagr_rev_per_share * 100, revs))
        if abs(cagr(revs) - cagr_rev_per_share) > .01:
            print("   Revenue {:.2f} in percent vs {:.2f} on per share basis".format(
                cagr_rev_per_share * 100, cagr_rev_per_share * 100))

        self.profiler.collect(cagr_rev_per_share, None, Tag.rev_per_share, ProfMethod.CAGR)

    def epu(self):
        ebt_exclude_unusual = strip(self.income.match_title('EBT Excl. Unusual Items'))
        shares_outstanding = strip(self.income.match_title('Weighted Average Diluted Shares Outstanding'))
        epu = list_over_list(ebt_exclude_unusual, shares_outstanding)
        cagr_epu = cagr(epu)
        print("EPU from {:.2f} to {:.2f} at CAGR {:.2f}% for: {}".format(
            epu[0]*100, epu[-1]*100,
            cagr_epu*100, epu))
        self.profiler.collect(cagr_epu, epu[-1]*100, Tag.epu, ProfMethod.CAGR)

    def cfo(self):
        # aka FFO - Funds from Operations
        cfo = strip(self.cashflow.match_title('Cash from Operations'))
        cfo_per_share = list(map(lambda f: f / self.share_out_filing(), cfo))
        cagr_cfo_per_share = cagr(cfo_per_share)
        print("FCF per share from {} to {} at CAGR {:.2f}% for: {}".format(
            cfo[0], cfo[-1],
            cagr_cfo_per_share * 100, cfo))
        if abs(cagr(cfo) - cagr_cfo_per_share) > .01:
            print("   FCF {:.2f} in percent vs {:.2f} on per share basis".format(
                cagr_cfo_per_share*100, cagr_cfo_per_share * 100))
        self.profiler.collect(cagr_cfo_per_share, cfo[-1], 'cfo_per_share', ProfMethod.CAGR)

    def _affo(self):
        cfo = strip(self.cashflow.match_title('Cash from Operations'))
        # Capex for real estates
        capex = strip(self.cashflow.match_title('Capital Expenditure'))
        affo = list_add_list(cfo, capex)

        # TODO made comparison in relation to IGBREIT's share out filing
        # share_out_filing = self.share_out_filing()
        share_out_filing = 3600
        affo_per_share = list(map(lambda f: f / share_out_filing, affo))

        # Based on IGBREIT 2021 annual: "term period" between 5.85% to 6.85%. Take the mid point.
        irr = .0635
        term_period = 0

        # reversed - Simulate in reversed order from current to far past year.
        last_term = None
        for i, a in enumerate(reversed(affo)):
            last_term = a/(1+irr)**(i+1)
            term_period += last_term
        # print("XXX", term_period, term_period/self.share_out_filing())
        last_term_period_over_shares = last_term / share_out_filing
        avg_term_period_over_shares = term_period / share_out_filing
        print("AFFO in relation to IGBREIT, at IRR {:.4f} for: {}".format(
            avg_term_period_over_shares,
            list(map(lambda x: round(x, 4), affo_per_share))
        ))
        self.profiler.collect(avg_term_period_over_shares, last_term_period_over_shares,
                              Tag.affo_per_share, ProfMethod.IRR)

    def affo(self):
        cfo = strip(self.cashflow.match_title('Cash from Operations'))
        # Capex for real estates
        capex = strip(self.cashflow.match_title('Capital Expenditure'))
        affo = list_add_list(cfo, capex)
        affo_per_share = list(map(lambda f: f / self.share_out_filing(), affo))
        # use Median rather than average.
        avg_affo_per_share = statistics.median(affo_per_share)
        print("AFFO at average {:.2f}% for: {}".format(
            avg_affo_per_share*100,
            list(map(lambda x: round(x, 4), affo_per_share))
        ))
        self.profiler.collect(avg_affo_per_share, affo_per_share[-1], Tag.affo_per_share, ProfMethod.AveragePerc)

    def nav(self):
        total_asset = strip(self.balance.match_title('Total Assets'))
        total_liab = strip(self.balance.match_title('Total Liabilities'))
        nav = list_minus_list(total_asset, total_liab)
        nav_per_share = list(map(lambda f: f / self.share_out_filing(), nav))
        avg_nav_per_share = cagr(nav_per_share)
        print("NAV per share at CAGR {:.2f}% for: {}".format(
            avg_nav_per_share*100,
            list(map(lambda x: round(x, 4), nav_per_share)),
        ))
        self.profiler.collect(avg_nav_per_share, nav_per_share[-1], Tag.nav_per_share, ProfMethod.CAGR)

    def return_equity(self):
        net_income = strip(self.income.match_title('Net Income$'))
        requity = strip(self.balance.match_title('Total Common Equity$'))
        roce = list_over_list(net_income, requity, percent=True)
        avg_roce = average(roce)
        print("Return on Common Equity average {:.2f}% for: {}".format(
            avg_roce,
            list(map(lambda x: round(x, 2), roce))
        ))
        # TODO ROCE is not defined
        self.profiler.collect(avg_roce/100, roce[-1], Tag.ROCE, ProfMethod.AveragePerc)

    def return_invested_cap(self):
        # ROIC = (nopat - tax) / (equity + debt + cash)
        # https://www.educba.com/invested-capital-formula/
        op_income = strip(self.income.match_title('Operating Income'))
        tax_not_strip = self.income.match_title('Income Tax Expense', none_is_optional=True)
        if tax_not_strip is not None:
            tax = strip(tax_not_strip)
            nopat = list_minus_list(op_income, tax)
        else:
            nopat = op_income
        debt = strip(self.balance.match_title('Total Debt$'))
        equity = strip(self.balance.match_title('Total Equity$'))
        cash = strip(self.cashflow.match_title('Cash from Investing$'))
        cash = list_add_list(
            strip(self.cashflow.match_title('Cash from Financing$')), cash)

        _1 = list_add_list(debt, equity)
        _2 = list_add_list(_1, cash)
        roic_per = list_over_list(nopat, _2, percent=True)
        avg_roic_per = average(roic_per)
        print("Return on Invested Capital average {:.2f}% for: {}".format(
            avg_roic_per,
            list(map(lambda x: round(x, 2), roic_per))
        ))
        self.profiler.collect(avg_roic_per/100, roic_per[-1]/100, Tag.ROIC, ProfMethod.AveragePerc)

    def net_debt_over_ebit(self):
        net_debt = strip(self.balance.match_title('Net Debt'))
        ebit = strip(self.income.match_title('Operating Income$'))
        # ebitda = self.match_title('EBITDA$')
        net_debt_over_ebit = list_over_list(net_debt, ebit)
        avg_net_debt_over_ebit = average(net_debt_over_ebit)
        print("Net debt over EBIT average {:.2f} years for: {}".format(
            avg_net_debt_over_ebit,
            list(map(lambda x: round(x, 2), net_debt_over_ebit))
        ))
        self.profiler.collect(avg_net_debt_over_ebit, net_debt_over_ebit[-1],
                              Tag.net_debt_over_ebit, ProfMethod.AverageYears)

    def ebit_margin(self):
        ebits = strip(self.income.match_title('Operating Income$'))
        revs = strip(self.income.match_title('Total Revenues$'))
        ebit_margins = list_over_list(ebits, revs, percent=True)
        avg_ebit_margins = average(ebit_margins)
        print("EBIT margin average {:.2f}% for (numbers in percent) {}".format(
            avg_ebit_margins,
            list(map(lambda x: round(x, 2), ebit_margins))
        ))
        self.profiler.collect(avg_ebit_margins/100, ebit_margins[-1],
                              Tag.ebit_margin, ProfMethod.AveragePerc)

    def ev_over_ebit(self):
        if self.values is None:
            # TODO exception to EV over EBIT
            print("Warning: ev_over_ebit: Missing values tab.")
            return
        ev_over_ebit = strip2(self.values.match_title('LTM Total Enterprise Value / EBIT$'))
        avg_ev_over_ebit = average(ev_over_ebit)
        print("EV over EBIT average {:.2f} ratio for: {}".format(
            avg_ev_over_ebit,
            list(map(lambda x: .0 if x is None else round(x, 2), ev_over_ebit))
        ))
        self.profiler.collect(avg_ev_over_ebit, ev_over_ebit[-1],
                              Tag.ev_over_ebit, ProfMethod.ReverseRatio)

    # TODO retined earnings pay in advance for one year?
    def retained_earnings_ratio(self):
        _ = strip(self.balance.match_title('Retained Earnings$'), trim_last=True)
        # skip the latest annual for retained earnings only
        retained_earnings = _[:-1]
        # TODO some Company such as IGBREIT does not provide Retained Earnings forecast
        # print("XXX retained earnings", len(retained_earnings), retained_earnings)
        _ = strip(self.income.match_title('Net Income$'), trim_last=True)
        # skip the first year annual for net income only
        net_income = _[1:]
        # print("XXX net income", len(net_income), net_income)
        retention_ratio = list_over_list(retained_earnings, net_income)
        # div_paid = strip(self.cashflow.match_title('Common Dividends Paid'))
        # retention_ratio = list_add_list(net_income, div_paid)

        last_retention_ratio = retention_ratio[-1]
        # retention ratio is measured by carry out from the previous annual report,
        # while adding net income to the current annual report

        avg_retention_ratio = average(retention_ratio)
        print("Retention ratio last {:.2f}, average {:.2f} for: {}".format(
            last_retention_ratio,
            avg_retention_ratio,
            list(map(lambda x: round(x, 2), retention_ratio))
        ))
        self.profiler.collect(avg_retention_ratio, last_retention_ratio,
                              Tag.retained_earnings_ratio, ProfMethod.Ratio)

    def dividend_payout_ratio(self):
        div_paid = strip(self.cashflow.match_title('Common Dividends Paid'))
        for i, a in enumerate(div_paid):
            if a is None:
                print("W: {} does not provide dividend in year '{}".format(
                    self.tick, self.start_year+i))
        income = strip(self.income.match_title('Net Income'))

        # Op income is a probable replacement in the event when regular income produce negative number.
        op_income = strip(self.income.match_title('Operating Income'))
        net_income = []
        for i, a in enumerate(income):
            if a < 0:
                print("W: {} negative income {} in year '{}".format(
                    self.tick, a, self.start_year+i))
                net_income.append(op_income[i])
            else:
                net_income.append(income[i])

        div_payout_ratio = list_over_list(div_paid, net_income)
        # Negating div payout to positive for the math to work easier
        avg_div_payout_ratio = - average(div_payout_ratio)
        print("Dividend payout ratio at average {:.2f} ratio for: {}".format(
            avg_div_payout_ratio,
            list(map(lambda x: round(x, 2), div_payout_ratio))
        ))
        self.profiler.collect(avg_div_payout_ratio, - div_payout_ratio[-1],
                              Tag.dividend_payout_ratio, ProfMethod.Average)

    def div_yield(self):
        if self.values is None:
            # TODO exception to EV over EBIT
            print("Warning: dividend yield: Missing values tab.")
            return
        self.values.match_title('LTM Dividend Yield$')
        div_yields = list(map(
            lambda z: 0 if z is None else z, strip2(self.values.match_title('LTM Dividend Yield$')))
        )
        avg_div_yield = average(div_yields)
        self.profiler.collect(avg_div_yield, div_yields[-1], Tag.dividend_yield, ProfMethod.Average)

    def last_price(self):
        if self.values is None:
            print("Warning: last_price: Missing values tab.")
            return

        price = self.values.match_title('Price$')
        ev_over_ebit = strip2(self.values.match_title('LTM Total Enterprise Value / EBIT$'))
        # TODO all div yields data?
        div_yield = strip2(self.values.match_title('LTM Dividend Yield$'))

        market_cap = self.values.match_title('Market Cap')
        rev = self.income.match_title('Total Revenues')
        op_income = self.income.match_title('Operating Income')
        net_profit = self.income.match_title('Net Income')

        dpu = self.income.match_title('Dividends Per Share')
        last_dpu = 0 if dpu[-1] is None else dpu[-1]

        price_over_affo = price[-1]/self.profiler.d[Tag.affo_per_share]['val2']

        last_div_yield = 0 if div_yield[-1] is None else div_yield[-1]
        self.profiler.collect_last_price({'last_price': price[-1],
                                          'ev_over_ebit': ev_over_ebit[-1],
                                          'last_div_yield': last_div_yield,
                                          'div_yields': div_yield,
                                          'market_cap': market_cap[-1],
                                          'revenue': rev[-1],
                                          'op_income': op_income[-1],
                                          'net_profit': net_profit[-1],
                                          'dpu': last_dpu,
                                          'price_over_affo': price_over_affo})

    def share_out_filing(self) -> float:
        x = self.balance.match_title('Total Shares Out\.')
        result = list(filter(None, reversed(x[1:])))[0]
        return result


class ProfMethod(Enum):
    CAGR = 1
    IRR = 2
    Average = 3
    AveragePerc = 4
    AverageYears = 5
    Ratio = 6
    ReverseRatio = 7


ProfVerbose = {ProfMethod.CAGR: 'CAGR',
               ProfMethod.IRR: 'IRR',
               ProfMethod.Average: 'average',
               ProfMethod.AveragePerc: 'percent',
               ProfMethod.AverageYears: 'year',
               ProfMethod.Ratio: 'ratio',
               ProfMethod.ReverseRatio: 'ratio',
               }


class Tag(Enum):
    rev_per_share = 1
    epu = 2
    affo_per_share = 3
    nav_per_share = 4
    # TODO ROCE is optional for tabulation.
    # ROCE = 5
    ROIC = 6
    net_debt_over_ebit = 7
    ev_over_ebit = 8
    ebit_margin = 9
    retained_earnings_ratio = 10
    # dividend_payout_ratio = 11
    dividend_yield = 12


class Prof:
    def __init__(self, name):
        self.name = name
        self.d = OrderedDict()
        self.last_price = None

        self.prof = {}

    def collect(self, val1, val2, tag, method: ProfMethod):
        # self.d[tag.__name__] = ratio
        # type: Tuple[float, ProfMethod]
        _ = OrderedDict(val1=val1, val2=val2, method=method)
        self.d[tag] = _

    def collect_last_price(self, last_price):
        self.last_price = last_price

    def profile(self):
        for k, v in self.d.items():
            if k is not str:
                self.prof[k] = v


class Bucket:
    def __init__(self, name, value, method):
        self.name = name
        self.value = value
        self.method = method
        self.score = None


class ProfManager:
    # TODO report about the underlying rate?
    Rate = {Tag.rev_per_share: {'high': .08, 'mid': .04},
            Tag.epu: {'high': .1, 'mid': .0},
            Tag.affo_per_share: {'high': .1, 'mid': .07},
            Tag.nav_per_share: {'high': .08, 'mid': .05},
            # TODO ROCE is undefined
            # Tag.ROCE: {'high': .08, 'mid': .065},
            Tag.ROIC: {'high': .06, 'mid': .05},
            Tag.net_debt_over_ebit: {'high': 5., 'mid': 8.},
            Tag.ebit_margin: {'high': .7, 'mid': .6},
            Tag.retained_earnings_ratio: {'high': 5., 'mid': .0},
            # Tag.dividend_payout_ratio: {'high': 1.5, 'mid': 1.},
            Tag.dividend_yield: {'high': .07, 'mid': .05},
            Tag.ev_over_ebit: {'high': 16., 'mid': 18.},
            }

    UnitRatio = (ProfMethod.AverageYears, ProfMethod.Ratio, ProfMethod.ReverseRatio)

    def __init__(self):
        self.companies = []     # type: List[Prof]
        self.company = {}
        self.metric = {}

    def create_folder(self, name):
        prof = Prof(name)
        self.companies.append(prof)
        self.company[name] = prof
        return prof

    def profile(self):
        for p in self.companies:
            # print(p.name)
            p.profile()
        self.bucketize()
        benched = self.benchmark()
        self.simulate_price(benched['comp'])
        self.output(benched)

    # Red = 'FF0000'
    # Yellow = 'FFFF00'
    # Green = '00FF00'

    # Rainbow Pastels Color Scheme
    # https://mappingmemories.ca/combfkiu2426925
    Red = 'FF9AA2'
    Yellow = 'FFDAC1'
    # Green = 'E2F0CB'
    Green = 'B5EAD7'

    # STABILO Boss Pastel colors palette | ColorsWall
    # Red = 'F6AA90'
    # Yellow = 'F8DF81'
    # Green = '9BDB07'

    def output(self, benched):
        ft = Font(name='Calibri', size=11)

        wb = Workbook()
        sheet = wb.active   # type: worksheet.Worksheet
        sheet.title = 'sheet 1'
        cell = sheet.cell(row=1, column=1)

        j = 1
        i = 2
        for x in Tag:
            cell = sheet.cell(row=j, column=i)
            cell.value = x.name
            i += 1

        # Ext to data based on Tag.
        ext_header = ['P', 'Market Cap', 'Revenue', 'Op income', 'Net profit', 'EPU sen', 'DPU sen',
                      'Price over AFFO',
                      'EV over EBIT', 'Dividend yield', 'ROIC', 'Net debt over EBIT', 'color']
        for x in range(len(ext_header)):
            cell = sheet.cell(row=j, column=i+x)
            cell.value = ext_header[x]
        i += len(ext_header)

        # TODO differential data, hmmm...
        # i += 1
        # cell = sheet.cell(row=j, column=i)
        # cell.value = 'diff ev_over_ebit'
        #
        # i += 1
        # cell = sheet.cell(row=j, column=i)
        # cell.value = 'incr div yield'

        # setting up color palette following
        gen_rule = ColorScaleRule(start_type='percentile', start_value=10, start_color=ProfManager.Red,
                                  mid_type='percentile', mid_value=50, mid_color=ProfManager.Yellow,
                                  end_type='percentile', end_value=90, end_color=ProfManager.Green)
        rule = {
            # net_debt_over_ebit
            Tag.net_debt_over_ebit:
                ColorScaleRule(start_type='num', start_value=1., start_color=ProfManager.Green,
                               mid_type='num', mid_value=8., mid_color=ProfManager.Yellow,
                               end_type='num', end_value=10., end_color=ProfManager.Red),
            # EV/EBIT
            Tag.ev_over_ebit:
                ColorScaleRule(start_type='num', start_value=15., start_color=ProfManager.Green,
                               mid_type='num', mid_value=18., mid_color=ProfManager.Yellow,
                               end_type='num', end_value=25., end_color=ProfManager.Red),
            # diff EV/EBIT
            'diff-ev_over_ebit':
                ColorScaleRule(start_type='percentile', start_value=90, start_color=ProfManager.Green,
                               mid_type='percentile', mid_value=50, mid_color=ProfManager.Yellow,
                               end_type='percentile', end_value=10, end_color=ProfManager.Red),
            'price_over_affo':
                ColorScaleRule(start_type='num', start_value=0., start_color=ProfManager.Green,
                               mid_type='num', mid_value=10., mid_color=ProfManager.Yellow,
                               end_type='num', end_value=15., end_color=ProfManager.Red),
            # Mid cap is between 200mil to 2bil.
            # https://www.bursamalaysia.com/trade/our_products_services/indices/ftse_bursa_malaysia_indices/overview
            'market_cap':
            # https://coolors.co/palette/07beb8-3dccc7-68d8d6-9ceaef-c4fff9
                ColorScaleRule(start_type='percentile', start_value=90, start_color='c4fff9',
                               mid_type='percentile', mid_value=20, mid_color='9ceaef',
                               end_type='percentile', end_value=10, end_color='3dccc7'),
        }
        start_row_index = 1
        end_row_index = len(self.companies) + 1

        j += 1
        for c in self.companies:
            # Table of mainly profile and last_price data
            lead = [
                {'val': c.prof[Tag.rev_per_share]['val1'], 'rule': gen_rule},
                {'val': c.prof[Tag.epu]['val1'], 'rule': gen_rule},
                {'val': c.prof[Tag.affo_per_share]['val1'], 'rule': gen_rule},
                {'val': c.prof[Tag.nav_per_share]['val1'], 'rule': gen_rule},
                {'val': c.prof[Tag.ROIC]['val1'], 'rule': gen_rule},
                {'val': c.prof[Tag.net_debt_over_ebit]['val1'], 'number': 'ratio',
                 'rule': rule[Tag.net_debt_over_ebit]},
                {'val': c.prof[Tag.ev_over_ebit]['val1'], 'number': 'ratio',
                 'rule': rule[Tag.ev_over_ebit]},
                {'val': c.prof[Tag.ebit_margin]['val1'], 'rule': gen_rule},
                {'val': c.prof[Tag.retained_earnings_ratio]['val1'], 'number': 'ratio', 'rule': gen_rule},
                {'val': c.prof[Tag.dividend_yield]['val1'], 'rule': gen_rule},
                # value2 to increase decimal point
                {'val': c.last_price['last_price'], 'number': 'value2'},
                {'val': c.last_price['market_cap'], 'number': 'cap', 'rule': rule['market_cap']},
                {'val': c.last_price['revenue'], 'number': 'value', 'rule': rule['market_cap']},
                {'val': c.last_price['op_income'], 'number': 'value', 'rule': rule['market_cap']},
                {'val': c.last_price['net_profit'], 'number': 'value', 'rule': rule['market_cap']},
                {'val': c.prof[Tag.epu]['val2'], 'number': 'value', 'rule': rule['market_cap']},
                # x100 - KLSE/Bursa DPU use fractional pricing model
                {'val': c.last_price['dpu']*100, 'number': 'value', 'rule': rule['market_cap']},
                {'val': c.last_price['price_over_affo'], 'number': 'value', 'rule': rule['price_over_affo']},
                {'val': c.prof[Tag.ev_over_ebit]['val2'], 'number': 'ratio',
                 'rule': rule[Tag.ev_over_ebit]},
                {'val': c.prof[Tag.dividend_yield]['val2'], 'rule': gen_rule},
                {'val': c.prof[Tag.ROIC]['val2'], 'rule': gen_rule},
                {'val': c.prof[Tag.net_debt_over_ebit]['val2'], 'number': 'ratio',
                 'rule': rule[Tag.net_debt_over_ebit]},
            ]

            i = 1
            cell = sheet.cell(row=j, column=i)
            cell.value = c.name

            i += 1
            for x in lead:
                cell = sheet.cell(row=j, column=i)
                assert 'val' in x
                cell.value = x['val']
                if 'number' in x:
                    cell.style = 'Comma'
                    if x['number'] == 'cap':
                        cell.number_format = '0,00'
                    elif x['number'] == 'value2':
                        cell.number_format = '0.00'
                        # test if we got enough decimal points
                        # Decimal point can be enabled using DECIMALS TO DISPLAY in TIKR terminal
                        if 0 != decimal.Decimal(c.last_price['last_price']*10**3).as_tuple().exponent:
                            cell.number_format = '0.000'
                    else:
                        # set number format to value/ratio
                        cell.number_format = '0.00'
                else:
                    cell.style = 'Percent'
                    cell.number_format = '0.00%'
                cell.font = ft

                if 'rule' in x:
                    sheet.conditional_formatting.add('{alpha}{start}:{alpha}{end}'.format(
                        alpha=colnum_string(i), start=start_row_index, end=end_row_index),
                        x['rule'])
                i += 1

            # adding last column for color
            cell = sheet.cell(row=j, column=i)
            cell.value = benched['score'][c.name]
            cell.number_format = '0.0'
            sheet.conditional_formatting.add('{alpha}{start}:{alpha}{end}'.format(
                alpha=colnum_string(i), start=start_row_index, end=end_row_index), gen_rule)
            i += 1

            j += 1
        wb.save('output.xlsx')

    def benchmark(self):
        # Grading process
        # For each metric based on rev_per_share, affo_per_share, nav_per_share, ...
        #   Rank based on above, moderate and below average.
        #       Collect the company for metric and rank level.
        # Accessible via the hash metric: metric[Tag.rev_per_share][RateType.above_avg]

        met = {}
        for k, v in self.metric.items():
            for v1 in v.values():
                if len(v1) > 0:
                    for x in v1:
                        if x.name not in met:
                            met[x.name] = 0
                        met[x.name] += x.score

        final = {RateType.above_avg: [], RateType.moderate_avg: [], RateType.below_avg: []}
        for k, v in met.items():
            if v >= 5.5:
                x = final[RateType.above_avg]
            elif v >= 4.5:
                x = final[RateType.moderate_avg]
            else:
                x = final[RateType.below_avg]
            x.append(k)

        print()
        for k, v in final.items():
            if k is RateType.above_avg:
                print("Companies that performed above average are: ", end='')
            elif k is RateType.moderate_avg:
                print("Companies that performed moderate average are: ", end='')
            else:
                print("Companies which are below average are: ", end='')
            print("{} out of {} sample".format(v, len(v)))
        print("Total companies sampled thus far is {}".format(len(self.companies)))
        return {'comp': final, 'score': met}

    def bucketize(self):
        # TODO namedtuple?

        for x in Tag:
            self.metric[x] = {RateType.above_avg: [], RateType.moderate_avg: [], RateType.below_avg: [], }

        for c in self.companies:
            for k, v in c.prof.items():
                if type(k) is not str:
                    if k in ProfManager.Rate:
                        assert k in self.metric
                        buck = self.metric[k]
                        tup = Bucket(c.name, v['val1'], v['method'])
                        if v['method'] in (ProfMethod.AverageYears, ProfMethod.ReverseRatio):
                            if v['val1'] < ProfManager.Rate[k]['high']:
                                buck[RateType.above_avg].append(tup)
                            elif v['val1'] < ProfManager.Rate[k]['mid']:
                                buck[RateType.moderate_avg].append(tup)
                            else:
                                buck[RateType.below_avg].append(tup)
                        else:
                            if v['val1'] > ProfManager.Rate[k]['high']:
                                buck[RateType.above_avg].append(tup)
                            elif v['val1'] > ProfManager.Rate[k]['mid']:
                                buck[RateType.moderate_avg].append(tup)
                            else:
                                buck[RateType.below_avg].append(tup)

        def value(val):
            return val.value

        def item(key: Bucket):
            return key.name, key.value

        def articulate(buckets: List[Bucket], key):
            values = list(map(value, buckets))
            if len(values) > 0:
                # TODO need to enumerate the index: -
                #  0 == company,
                #  1 == value of percent, years, etc.,
                #  2 == name of method see ProfMethod class
                method = buckets[0].method

                def at(k):
                    if method in ProfManager.UnitRatio:
                        return '{} at {:.2f} yrs'.format(k[0], k[1])
                    return '{} at {:.2f}%'.format(k[0], k[1] * 100)

                items = list(map(item, buckets))
                comp_at_perc = ', '.join(list(map(at, items)))

                unit = Unit(value=100, symbol='%')
                if method in ProfManager.UnitRatio:
                    unit = Unit(value=1, symbol='')

                # TODO modify current "performed above average rate"
                #  to "below the average over the last 10 years sampled, at undemanding rate"
                print("{}/{} companies sampled have performed {} average rate of {} {:.2f}{}. ".format(
                    len(buckets), len(self.companies), RateVerbose[key], ProfVerbose[method],
                    average(values) * unit.value, unit.symbol), end='')
                if key is RateType.above_avg:
                    print("These companies are: {}".format(comp_at_perc))
                    # TODO apply() function?
                    for a in buckets:
                        assert a.score is None
                        # Scale it to one per metric, 10 for full points
                        a.score = 10./len(self.metric)
                elif key is RateType.moderate_avg:
                    print("These companies are: {}".format(comp_at_perc))
                    for a in buckets:
                        assert a.score is None
                        # Scale it to half point per metric, 10 for full points
                        a.score = 5./len(self.metric)
                else:
                    for a in buckets:
                        assert a.score is None
                        a.score = 0
                    print()

        # TODO need to solve for AFFO, net debt over ebit, retention ratio, div payout ratio

        for k, v in self.metric.items():
            print("Based on {}:".format(k))
            # TODO Near right but not enum hmmm RateType._fields:
            for avg_rate in RateType.above_avg, RateType.moderate_avg, RateType.below_avg:
                articulate(v[avg_rate], avg_rate)

    def simulate_price(self, benched):
        print()

        def _(cat):
            for k, v in benched.items():
                if k in cat:
                    for x in v:
                        assert x in self.company
                        company = self.company[x]
                        if company.last_price is not None:
                            print("{}'s last quote was {}: -".format(x, company.last_price['last_price']))

                            current = company.last_price['ev_over_ebit']
                            record = company.prof[Tag.ev_over_ebit]['val1']
                            diff = (current - record) / record
                            ev_over_ebit_incr = 'undemanding price'\
                                if current < record else 'premium valuation/overpriced'
                            trend = '+' if diff > 0 else ''
                            print("- Last EV over EBIT was {current:.2f},"
                                  " average was {record:.2f}, valued at {ev_over_ebit_incr} after {diff:.2f} pts"
                                  .format(ev_over_ebit_incr=ev_over_ebit_incr, trend=trend,
                                          diff=diff*100, record=record, current=current))

                            # TODO refactored the block below
                            div_yields = list(map(lambda z: 0 if z is None else z, company.last_price['div_yields']))
                            avg_div_yield = average(div_yields)
                            dy_incr = 100 * (company.last_price['last_div_yield'] - avg_div_yield) / avg_div_yield
                            trend = 'upside +' if dy_incr > 0 else 'downside '
                            print("- Last div yield was {:.2f} %, we see last {}{:.2f} pts"
                                  " based on average div yield of {:.2f} % ".format(
                                    company.last_price['last_div_yield']*100, trend, dy_incr, avg_div_yield*100, ))

        _([RateType.above_avg, RateType.moderate_avg])
        print("\nThe following quotes were rated at below average rating though")
        _([RateType.below_avg])


def main():
    path = "C:/Users/benny/iCloudDrive/Documents/Bursa Malaysia Energy Infrastructure, Equipment & Services Companies"
    prof = ProfManager()

    tickers = ['deleum',
               'dialog', 'yinson', 'armada', 'dayang', 'coastal',
               'velesto', 'saprng', 'mhb', 'waseong',
               #'icon'
               't7global', 'penergy', 'perdana', 'uzma', 'carimin']

    # TODO Adding TODO may need to fix AHP.
    # tickers = ['carimin']
    for c in tickers:
        print('Ticker {}'.format(c))
        wb = load_workbook(path+'/' + c + '.xlsx')
        pf = prof.create_folder(c)
        t = Spread(wb, c, pf)
        t.revenue()
        t.epu()
        # t.cfo()
        t.affo()
        t.nav()
        # t.return_equity()
        t.return_invested_cap()
        t.net_debt_over_ebit()
        t.retained_earnings_ratio()
        t.ebit_margin()
        t.ev_over_ebit()
        # t.dividend_payout_ratio()
        t.div_yield()
        t.last_price()
        print()

    prof.profile()


if __name__ == '__main__':
    main()

from calculator import *
from bcolors import colour_print, bcolors
import yfinance as yf
from openpyxl import Workbook, worksheet
from collections import OrderedDict

one_hundred = 100


class SpreadOut:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.create_sheet('Header')
        self.od = OrderedDict()
        self.column = 1

        # removing initial sheet
        ws = self.wb.active
        self.wb.remove(ws)

    def save(self):
        self.wb.save('{}.xlsx'.format("dividend-out"))


class _Spread:

    def __init__(self, spread: SpreadOut, ticker):
        self.od = spread.od
        self.ws = spread.ws
        self.column = spread.column
        self.row = 2

        # Set row to ticker name
        self.ws.cell(row=1, column=self.column, value=ticker)

    def write(self, title, ratio=None, bold=False):
        if title not in self.od:
            # Init pre-config template
            self.ws.cell(row=self.row, column=1, value=title)
            if bold:
                self.ws.cell(row=self.row, column=1).font = Font(bold=True)
            self.od[title] = {"row": self.row, "ratio": None}
            self.ws.column_dimensions[colnum_string(1)].width = 36

            self.row += 1

        if ratio is not None:
            print("{}: {:.2f}%".format(title, ratio * one_hundred))
            assert title in self.od
            od = self.od[title]
            od["ratio"] = ratio
            cell = self.ws.cell(row=od["row"], column=self.column)
            cell.value = ratio
            cell.number_format = '0.00%'


class Dividend(Spread):
    def __init__(self, tick, spread, path='spreads'):
        colour_print("Company's ticker '{}'".format(tick), bcolors.UNDERLINE)

        self.wb: Workbook = load_workbook(path + '/' + tick + '.xlsx')
        super().__init__(self.wb, tick)

        self.spread = _Spread(spread, tick)
        self.shares_out = self.income.match_title('Weighted Average Diluted Shares Outstanding')
        self.div_paid = self.cashflow.match_title('Common Dividends Paid')
        self.ticker = yf.Ticker(get_symbol(self.tick))
        self.mcap = self.ticker.info['marketCap'] / 1e6

    def trim_estimates(self, title, **kwargs):
        # **kwargs: Passthrough to allow none_is_optional, optional argument
        result = None
        for i, a in enumerate(self.estimates.date_range):
            if re.match(r'.*\bE$', a):
                # Assuming table sorted in 'A' (actual) to 'E' (estimates)
                tab = self.estimates.match_title(title, **kwargs)
                if tab is not None:
                    result = []
                    for t, d in zip(tab[i+1:], self.estimates.date_range[i:]):
                        if t is not None:
                            result.append((t, d))
                    # result = [(t, d) for t, d in zip(tab[i:], self.estimates.date_range[i:])
                    #           if t is not None]
                break
            elif re.match(r'.*\bA$', a):
                pass
            else:
                assert False
        return result

    def add_subtitle(self, subtitle):
        colour_print(subtitle, bcolors.UNDERLINE)
        self.spread.write(subtitle, bold=True)

    # Dividend yield
    def _dividend_yield_ttm(self):
        # TTM dividend yield
        # LTM (last twelve month) has already normalized it to TTM
        div_paid = self.cashflow.match_title('Common Dividends Paid')
        div0 = abs(div_paid[-1]) / self.shares_out[-1]
        div_yield = (div0 * self.shares_out[-1]) / self.mcap
        return div_yield

    def _dividend_yield_ntm(self):
        # Forward dividend yield
        # dps = self.trim_estimates('Dividend Per Share')
        # shares_out = self.shares_out[-1]
        # TODO annually, semi quarter, and quarterly granularity
        # d1 = dps[1]/2. * shares_out
        # d2 = dps[2]/2. * shares_out
        # d = d1 + d2
        # a = d / self.mcap
        # print(a)
        # return a
        # TODO might miss out on dilution?
        return self.strip(self.values.match_title('NTM Dividend Yield'))[-1]

    def _dividend_yield_past_years(self):
        # 4 years dividend yield
        market_cap = self.strip(self.values.match_title('Market Cap'))
        i = 0
        MC = []
        while i < len(market_cap):
            MC.append(average(market_cap[i:i+4]))
            # print(a, i, j)
            i += 4

        div_yield = []
        for i in range(1, 5):
            if self.div_paid[-i] is None:
                colour_print("No dividend was paid in year {}".format(i), bcolors.WARNING)
                div_yield.append(0)
                continue
            dps = abs(self.div_paid[-i]) / self.shares_out[-i]
            a = (dps * self.shares_out[-i]) / MC[-i]
            div_yield.append(a)
        return average(div_yield)

    def _earnings_yield_ttm(self):
        # Earnings yield
        earnings = self.income.match_title('Net Income')
        asset_writedown = self.income.match_title('Asset Writedown')
        eps0 = (earnings[-1] - asset_writedown[-1]) / self.shares_out[-1]
        earning_yield = (eps0 * self.shares_out[-1]) / self.mcap
        return earning_yield

    def _earnings_yield_ntm(self):
        # earnings = self.trim_estimates('Net Income Normalized')
        # eps0 = abs(earnings[1]) / self.shares_out[-1]
        eps0 = self.values.match_title('NTM Normalized Earnings Per Share')
        earning_yield = (eps0[-1] * self.shares_out[-1]) / self.mcap
        return earning_yield

    def _yield_on_cost(self, r0, r1=-1):
        # YOC
        # For example, if an investor purchased a stock five years ago for $20,
        # and its current dividend is $1.50 per share,
        # then the YOC for that stock would be 7.5%.
        price = self.strip(self.values.match_title('Price$'))
        i = 0
        P = []
        while i < len(price):
            P.append(average(price[i:i + 4]))
            # print(a, i, j)
            i += 4

        div_paid = self.cashflow.match_title('Common Dividends Paid')
        dps = abs(average(div_paid[r0:r1])) / average(self.shares_out[r0:r1])
        yoc = dps / average(P[r0:r1])
        return yoc

    # Dividend growth
    def _dps_growth_ntm(self):
        # DPS growth ntm
        dps = self.strip(self.values.match_title('NTM Dividend / Share'))
        dps_growth = average(dps[-4:]) / average(dps[-8:-4]) - 1
        return dps_growth

    def _dps_growth_ltm(self):
        # DPS growth ltm
        dps = self.strip(self.values.match_title('LTM Dividend Per Share'))
        dps_growth = average(dps[-4:]) / average(dps[-8:-4]) - 1
        return dps_growth

    def _dps_growth_next_years(self):
        dps, years = zip(*self.trim_estimates('Dividend Per Share'))
        return cagr(dps), years

    def _dps_growth_past_years(self, n_year):
        dps = self.strip(self.income.match_title('Dividends per share'))
        trim = 0
        # for i in range(len(dps[-n_year:])):
        for i in range(n_year):
            if dps[-n_year+i] is None:
                trim += 1
        dps_growth = dps[-n_year+trim:]
        years = self.income.date_range[-n_year+trim:]
        if trim > 0:
            colour_print("Short of {} past year".format(trim), bcolors.WARNING)
        return cagr(dps_growth), years

    def _cash_dividend_payout_ratio_ltm(self):
        div_paid = self.strip(self.cashflow.match_title('Common Dividends Paid'))
        net_income = self.strip(self.income.match_title('Net Income$'))
        ffo = net_income[-1]
        a = self.cashflow.match_title('Total Asset Writedown', none_is_optional=True)
        if a is not None:
            asset_writedown = self.strip(a)
            ffo += asset_writedown[-1]

        a = self.cashflow.match_title('Total Depreciation', none_is_optional=True)
        if a is not None:
            depreciation = self.strip(a)
            ffo += depreciation[-1]

        a = self.cashflow.match_title('\\(Gain\\) Loss On Sale of Asset', none_is_optional=True)
        if a is not None:
            gain_loss_asset_sales = self.strip(a)
            if gain_loss_asset_sales[-1] is not None:
                ffo += gain_loss_asset_sales[-1]
        return -div_paid[-1] / ffo

    def _dividend_payout_ratio_ltm(self):
        dps = self.strip(self.income.match_title('Dividends per share'))
        eps = self.strip(self.income.match_title('Diluted EPS'))
        return dps[-1] / eps[-1]

    def _dividend_payout_ratio_1y(self):
        dps = self.trim_estimates('Dividend Per Share')
        eps = self.trim_estimates('EPS Normalized')
        if len(eps) == 0:
            colour_print("No payout estimation was provided", bcolors.WARNING)
            return 0.

        return dps[0][0] / eps[0][0]

    def _cash_flow_payout_ratio(self):
        div_paid = self.strip(self.cashflow.match_title('Common Dividends Paid'))
        cfo = self.strip(self.cashflow.match_title('Cash from Operations'))
        acq_reit = self.strip(self.cashflow.match_title('Acquisition of Real Estate Assets'))
        # TODO Normal stock use Free Cash Flow
        # Special for REIT Acq.
        return -div_paid[-1] / (cfo[-1] - acq_reit[-1])

    def _cash_flow_payout_ratio_1y(self):
        _ = self.trim_estimates('Dividend Per Share')
        dps = _[0][0]
        _ = self.income.match_title('Weighted Average Diluted Shares Outstanding')
        shares_out = _[-1]
        _ = self.trim_estimates('Cash From Operations')
        cfo = _[0][0]
        _ = self.trim_estimates('Capital Expenditure')
        capex = _[0][0]
        result = (dps * shares_out) / (cfo + capex)
        return result

    def _cash_flow_yield_to_dividend_yield(self):
        cfo = self.strip(self.cashflow.match_title('Cash from Operations'))
        acq_reit = self.strip(self.cashflow.match_title('Acquisition of Real Estate Assets'))
        # shares_out = self.income.match_title('Weighted Average Diluted Shares Outstanding')
        fcf_yield = (cfo[-1] + acq_reit[-1]) / self.mcap
        div_yield = self.strip(self.values.match_title('LTM Dividend Yield'))[-1]
        return fcf_yield / div_yield

    def _dividend_yield_to_dividend_payout(self):
        div_yield = self.strip(self.values.match_title('LTM Dividend Yield'))[-1]
        div_payout = self._dividend_payout_ratio_ltm()
        return div_yield / div_payout

    def _dividend_payout_ratio_next_years(self, year_end):
        past_dps = self.strip(self.income.match_title('Dividends per share'))
        dps, years = zip(*self.trim_estimates('Dividend Per Share'))
        return cagr([past_dps[-1]] + list(dps[:year_end]))

    def _price_over_nav(self):
        book_value = self.balance.match_title('Tangible Book Value')
        return self.mcap / book_value[-1]

    def _liability_to_asset_ratio(self):
        total_liability = self.balance.match_title('Total Liabilities')
        total_assets = self.balance.match_title('Total Assets')
        return total_liability[-1] / total_assets[-1]

    def _debt_to_equity_ratio(self):
        debt = self.balance.match_title('Total Debt')
        equity = self.balance.match_title('Total Equity')
        return debt[-1] / equity[-1]

    def _cash_to_short_term_debt(self):
        cash = self.strip(self.balance.match_title('Cash And Equivalents'))
        current_debt = self.strip(self.balance.match_title('Current Portion of Long-Term Debt'))
        return cash[-1] / current_debt[-1]

    def compute_dividend_yield(self):
        self.add_subtitle("Dividend Yield")
        self.spread.write("4 year average dividend yield past years", ratio=self._dividend_yield_past_years())
        self.spread.write("Dividend yield (TTM)", ratio=self._dividend_yield_ttm())
        self.spread.write("Dividend yield (FWD)", ratio=self._dividend_yield_ntm())
        self.spread.write("1 year yield on cost", ratio=self._yield_on_cost(-2))
        self.spread.write("3 year yield on cost", ratio=self._yield_on_cost(-4))
        self.spread.write("5 year yield on cost", ratio=self._yield_on_cost(-6))
        self.spread.write("Earnings yield (TTM)", ratio=self._earnings_yield_ttm())
        self.spread.write("Earnings yield (NTM)", ratio=self._earnings_yield_ntm())

        self.spread.row += 1
        print()

    def compute_dps_growth(self):
        self.add_subtitle("DPS Growth")
        self.spread.write("DPS growth (FWD)", ratio=self._dps_growth_ntm())
        self.spread.write("DPS growth (LTM)", ratio=self._dps_growth_ltm())

        dps, years = self._dps_growth_next_years()
        # self.spread.write("DPS growth rate for next {} years since {} (CAGR)".format(
        self.spread.write("DPS growth rate for next {} years (CAGR)".format(
            len(years), ), ratio=dps)

        # TODO past growth rate
        dps, years = self._dps_growth_past_years(3)
        self.spread.write("DPS growth rate for past {} years (CAGR)"
                          .format(len(years), ), ratio=dps)

        dps, years = self._dps_growth_past_years(5)
        self.spread.write("DPS growth rate for past {} years (CAGR)"
                          .format(len(years), ), dps)
        try:
            dps, years = self._dps_growth_past_years(9)
            self.spread.write("DPS growth rate for past {} years (CAGR)"
                              .format(len(years), ), dps)
        except TypeError:
            pass
        self.spread.row += 1
        print()

    def compute_dividend_safety(self):
        self.add_subtitle("Dividend Safety")
        self.spread.write("Cash dividend payout ratio (LTM)", ratio=self._cash_dividend_payout_ratio_ltm())
        self.spread.write("Dividend payout ratio (LTM)", ratio=self._dividend_payout_ratio_ltm())
        self.spread.write("Dividend payout ratio 1y", ratio=self._dividend_payout_ratio_1y())
        self.spread.write("Cash flow dividend payout ratio (LTM)", ratio=self._cash_flow_payout_ratio())
        self.spread.write("Cash flow dividend payout ratio 1y", ratio=self._cash_flow_payout_ratio_1y())
        self.spread.write("Free cash flow yield to dividend yield ratio (LTM)",
                          self._cash_flow_yield_to_dividend_yield())
        self.spread.write("Dividend yield ratio to dividend payout ratio (LTM)",
                          self._dividend_yield_to_dividend_payout())
        self.spread.row += 1
        print()

    def compute_consistency(self):
        pass

    def compute_dividend_estimates(self):
        self.add_subtitle("Dividend Estimates")
        self.spread.write("Dividend payout growth next 3y", self._dividend_payout_ratio_next_years(3))
        self.spread.write("Dividend payout growth next 2y", self._dividend_payout_ratio_next_years(2))
        self.spread.row += 1
        print()

    def compute_metrics(self):
        # NAV = Asset - Liabilities / number of outstanding shares
        colour_print("Other metrics", bcolors.UNDERLINE)
        print("Price over NAV: {:.2f}".format(self._price_over_nav()))
        print("Liability to asset ratio: {:.2f}".format(self._liability_to_asset_ratio()))
        print("Debt to equity ratio: {:.2f}".format(self._debt_to_equity_ratio()))
        print("Cash to short term debt ratio: {:.2f}".format(self._cash_to_short_term_debt()))
        print()

    def compute(self):
        self.compute_dividend_safety()
        self.compute_dps_growth()
        self.compute_dividend_yield()
        # self.compute_consistency()
        self.compute_dividend_estimates()
        # self.compute_metrics()


spread = SpreadOut()
# for co in ['kipreit', 'igbreit', 'klcc', 'ytlreit', 'xzl']:
cos = ['kipreit', 'igbreit', 'klcc', 'ytlreit',
       'sunreit', 'pavreit', 'axreit', 'clmt',
       'sentral', 'alaqar', 'uoareit', 'hektar',
       'alsreit', ]
cos = ['sunreit']
# arreit, atrium, amfirst, twrreit

for co in cos:
    spread.column += 1
    data = Dividend(co, spread)
    data.compute()
spread.save()

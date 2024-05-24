from selenium import webdriver
import chromedriver_binary
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import MoveTargetOutOfBoundsException
from openpyxl import Workbook, worksheet
from bs4 import BeautifulSoup
import pyperclip
import re
import time
from tkinter import *
import math

from table import Page, fix_currency, fix_percent
from bcolors import colour_print, bcolors
import bootstrap

global driver


class WaitCapture:
    def __init__(self):
        self.ticker = None

    def run(self):
        print("Waiting to prompt ticker")
        # https://stackoverflow.com/questions/110923/how-do-i-close-a-tkinter-window
        looper = Tk()
        Button(looper, text="Run auto", command=looper.destroy).pack()
        Button(looper, text="Exit", command=exit).pack()
        looper.mainloop()
        # elem = driver.find_element(By.ID, "input-168")
        elem = driver.find_element(By.CLASS_NAME, "v-chip__content")
        assert elem.text is not None
        print("Found ticker", elem.text)
        self.ticker = elem.text.lower()


class MainTable:
    def __init__(self, ticker):
        self.ticker = ticker

    def open(self, title, start_offset=None, period_offset=None):
        print("waiting to load {title}".format(title=title))
        time.sleep(5)

        driver.find_element(By.PARTIAL_LINK_TEXT, "{title}".format(title=title)).click()
        # TODO 5 secs sleep might not work in Valuation spread.
        time.sleep(8)

        if start_offset is not None:
            assert period_offset is not None
            print("sliding year range to max years available")
            driver.find_element(By.CLASS_NAME, "v-slider__thumb")
            print("found slider thumb")
            move = ActionChains(driver)

            # https://stackoverflow.com/questions/40485157/how-to-move-range-input-using-selenium-in-python
            key = "aria-valuemax"
            elem = driver.find_element(By.CSS_SELECTOR, "div[{}]".format(key))
            avail_period = int(elem.get_dom_attribute('{}'.format(key)))

            offset = period_offset*start_offset/avail_period

            print("period_offset {} start_offset {} avail_period {} offset {:.2f}".format(
                period_offset, start_offset, avail_period, offset))
            try:
                move.click_and_hold(elem).move_by_offset(offset, 0).release().perform()
            except MoveTargetOutOfBoundsException:
                # TODO relative offset position
                # fpgroup
                # period_offset 58 start_offset -1050 avail_period 36 offset -1691.67
                # period_offset 50 start_offset -1050 avail_period 36 offset -1458.33
                # gtronic
                # period_offset 58 start_offset -1050 avail_period 69 offset -882.61
                colour_print("move target out of bound exception. Trying to retract the offset", bcolors.WARNING)
                fixed = 50
                x = fixed*start_offset/avail_period
                print("period_offset {} start_offset {} avail_period {} offset {:.2f}".format(
                    fixed, start_offset, avail_period, x))
                move.click_and_hold(elem).move_by_offset(x, 0).release().perform()
            time.sleep(5)
        else:
            print("Skip start offset")

    def run(self):
        print("Waiting to prompt header dialog")
        time.sleep(3)
        header_containers = driver.find_element(By.XPATH,
            "//div[contains(@class, 'container') and contains(@class, 'header')]")
        line = header_containers.text.split('\n')
        assert len(line) > 6
        header = line[0]
        sticky_price = line[6]

        # Parse the first line text only at the time being.
        clip = Clipboard(header, sticky_price)

        # max case for full span
        # offset = 15*start_offset/years# self.open('Financials', start_offset=0)

        # Based on initial number of years setting + fudge factor
        # offset = 5*start_offset/years
        self.open('Financials', start_offset=-1050, period_offset=5)
        clip.run(selection=["Income Statement", "Balance Sheet", "Cash Flow Statement"])

        # No change to start offset. Period offset based on 10 years in quarterly period
        self.open('Valuation', start_offset=-1050, period_offset=58)
        clip.run(selection="Values")

        self.open('Estimates')
        clip.run(selection="Estimates")

        clip.save(self.ticker)
        print("Saved to {}".format(self.ticker))


class Clipboard:
    def __init__(self, header, sticky_price):
        self.wb = Workbook()
        ws = self.wb.create_sheet('Header')
        # set cell to header
        ws.cell(row=1, column=1, value=header)
        ws.cell(row=2, column=1, value=sticky_price)

        # removing initial sheet
        ws = self.wb.active
        self.wb.remove(ws)

    def run(self, selection=None):
        # TIKR's support decided to remove "Copy Table" button to comply with data provider or negotiate further
        if type(selection) is list:
            for t in selection:
                self.select(t)
                self.write_excel(t)
                # self.copy_table(t)
                # self.paste(t, pref_num_format='0.00')
        elif type(selection) is str:
            # skipping selection dialog
            t = selection
            self.write_excel(t)
            # self.copy_table(t)
            # self.paste(t, pref_num_format='0.00')

    def select(self, title):
        # https://stackoverflow.com/questions/21713280/find-div-element-by-multiple-class-names
        txt = "//*[contains(text(), '{title}')]".format(title=title)
        driver.find_element(By.XPATH, txt).click()
        # driver.find_element(By.XPATH, "//*[text()='Income Statement']").click()
        # driver.find_element(By.XPATH, "//*[text()='Balance Sheet']").click()
        # driver.find_element(By.XPATH, "//*[contains(text(), 'Cash Flow Statement')]").click()

    def copy_table(self, title):
        print("clicking 'Copy Table to Clipboard' on '{title}' table".format(title=title))
        # alternative to contains function
        # elem = driver.find_element(By.XPATH, "//button[contains(@class, 'v-btn v-btn--icon v-btn--round theme--light v-size--default primaryAction--text')]")

        # Based on dark theme setting
        # v-btn v-btn--icon v-btn--round theme--dark v-size--default primaryAction--text
        elem = driver.find_element(By.XPATH, "//button[@class='v-btn v-btn--icon v-btn--round theme--light v-size--default primaryAction--text']")
        ActionChains(driver).click(elem).perform()
        print("copied table", elem.text)
        time.sleep(3)

    def paste(self, title, pref_num_format=None):
        # https://stackoverflow.com/questions/62527396/real-time-copying-and-pasting-to-excel
        # https://python-forum.io/thread-26979.html
        clipped = pyperclip.paste()
        clipped = clipped.split('\r\n')
        clipped = [item.split('\t') for item in clipped]

        # worksheet.worksheet.Worksheet
        # type: worksheet
        first_word = title.split()[0]
        ws = self.wb.create_sheet(first_word)
        # created sheet may intro the title automatically
        # ws.title = "Income Statement"

        for row, row_data in enumerate(clipped, start=1):
            for col, cell_val in enumerate(row_data, start=1):
                # print(cell_data)
                number_flag = False
                per_flag = False
                if re.match(Page.re_numerical, cell_val):
                    try:
                        if re.match(Page.re_percent, cell_val):
                            cell_val = fix_percent(cell_val)
                            per_flag = True
                        else:
                            # assuming it is a number until it caught exception
                            cell_val = fix_currency(cell_val)
                            number_flag = True
                    except ValueError:
                        pass
                cell = ws.cell(row=row, column=col, value=cell_val)
                # https://stackoverflow.com/questions/12387212/openpyxl-setting-number-format
                if number_flag:
                    if pref_num_format is not None:
                        cell.number_format = pref_num_format
                        if len(str(abs(math.floor(cell_val)))) > 3:
                            # Numerical number larger than 3 digits add prefix
                            cell.number_format = "0,00" + pref_num_format
                    else:
                        cell.number_format = '0,00'
                elif per_flag:
                    cell.number_format = '0.00%'

    def write_excel(self, title):
        soup = BeautifulSoup(driver.page_source, 'lxml')
        a = [x for x in soup.find_all('table')]
        if len(a) == 0:
            colour_print("Empty table: \"{}\"".format(title), bcolors.WARNING)
            return

        bootstrap.Out = open('console.log', 'w')
        with bootstrap.Out as out:
            page = Page(a[0])
            page.parse_top()
            first_word = title.split()[0]
            ws = self.wb.create_sheet(first_word)
            pref_num_format = '0.00'
            for row, row_data in enumerate(page.data, start=1):
                for col, cell_val in enumerate(row_data, start=1):
                    number_flag = False
                    per_flag = False
                    val = cell_val
                    if type(cell_val) is float:
                        number_flag = True
                    elif type(cell_val) is str:
                        if re.match(Page.re_percent, val):
                            per_flag = True
                            val = fix_percent(val)
                    cell = ws.cell(row=row, column=col, value=val)
                    if number_flag:
                        if pref_num_format is not None:
                            cell.number_format = pref_num_format
                            if len(str(abs(math.floor(cell_val)))) > 3:
                                # Numerical number larger than 3 digits add prefix
                                cell.number_format = "0,00" + pref_num_format
                        else:
                            cell.number_format = '0,00'
                    elif per_flag:
                        cell.number_format = '0.00%'

    def save(self, ticker):
        self.wb.save('{}.xlsx'.format(ticker))


def run_main():
    global driver

    # bootstrap functions
    driver = webdriver.Chrome()
    session_id = driver.session_id
    executor_url = driver.command_executor._url
    # driver.get("https://app.tikr.com/stock/estimates?ref=iwd7tf")
    driver.get("https://app.tikr.com/login")

    print("Session id:", session_id)
    print("Exec URL:", executor_url)
    print("Current cookies:", driver.get_cookies())

    driver.maximize_window()

    # TODO input ID?
    input_id = 'input-13'
    elem = driver.find_element(By.ID, input_id)
    with open('meow.txt') as f:
        if elem is not None:
            # elem.clear()
            elem.send_keys(f.readline())
            elem.send_keys(Keys.RETURN)

        elem = driver.find_element(By.ID, "input-16")
        if elem is not None:
            elem.send_keys(f.readline())
            elem.send_keys(Keys.RETURN)

    # https://www.selenium.dev/documentation/webdriver/waits/
    print("wait until username")
    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, input_id))
    )
    print("got it")

    def repeat():
        # search the text manually on Search entry manually
        capture = WaitCapture()
        capture.run()

        main = MainTable(capture.ticker)
        # run the pending chain of actions
        main.run()

    # loop forever until stop interruption
    while True:
        repeat()
        # time.sleep(3600)


if __name__ == '__main__':
    run_main()

# Footnotes
# - Selenium manpage.
#   https://selenium-python.readthedocs.io/installation.html
# - If iframe can't be solved hmmm...
#   https://www.guru99.com/handling-iframes-selenium.html
# - UI testing
#   https://success.outsystems.com/documentation/how_to_guides/devops/how_to_do_ui_testing_with_selenium/#using-selenium-for-testing-the-ui-of-your-apps
# - Requests: https://www.geeksforgeeks.org/get-post-requests-using-python/?ref=lbp

# TODO spread shorter than 10 years old

import functools
from operator import add, sub, mul, truediv
from openpyxl import Workbook, load_workbook, worksheet
from openpyxl.styles import Font, Alignment
import re
from utils import *
from collections import OrderedDict
from functools import cached_property
from datetime import datetime
from spread import *
from utils import excel_to_decimal

total_main_col = 12
total_half_col = int(total_main_col / 2)
total_elem = 10
total_half_elem = int(total_elem / 2)
prev_year_offset = 1
start_year_offset = 2
next_year_offset = 3
half_base_offset = total_half_elem + 2


class ExcelWriter:
    start_col = 2
    row_margin = 1

    def __init__(self, tick: str):
        cls = self.__class__
        self.tick = tick
        # self.styles = styles
        self.ft = Font(name='Calibri', size=11)
        self.wb = Workbook()

        # type: worksheet.Worksheet
        self.sheet = self.wb.active
        self.sheet.title = 'sheet 1'

        self.cell = self.sheet.cell(row=1, column=cls.start_col)
        self.start_row_index = cls.row_margin+1
        self.end_row_index = len(self.tick) + self.start_row_index+1

        self.dict = None
        self.init_sheet()

    def init_sheet(self):
        sheet = self.sheet
        cell = sheet.cell(row=1, column=1)
        cell.value = self.tick.upper()
        sheet.cell(row=1, column=2).value = 'Base year'
        for i in range(1, total_main_col-1):
            cell = sheet.cell(row=1, column=i+2)
            cell.value = i+datetime.datetime.now().year-1
            cell.alignment = Alignment(horizontal='center')
        sheet.cell(row=1, column=total_main_col+1).value = 'Terminal year'

    def create_dict(self):
        self.dict = ExcelDict(self)
        return self.dict


def excel_calc(wb, cell):
    sheet = wb.active
    string = ''
    if cell.value[0] == '=':
        # print("XXX", cell.value[0])
        line = cell.value[1:]
        for x in re.findall(r'\w+|\S', line):
            m = re.match(r'([A-Za-z]+)(\d+)', x)
            if m is not None:
                cell = sheet.cell(
                    row=int(m.group(2)),
                    column=excel_to_decimal(m.group(1)))
                string += ' ' + str(cell.value)
            else:
                string += ' ' + x
    return calculate(string)


class ExcelDict:
    def __init__(self, excel: ExcelWriter):
        self.excel = excel

    def create_array(self, key, row, style='Comma'):
        start_col = 1
        sheet = self.excel.sheet
        sheet.column_dimensions[colnum_string(start_col)].width = 32
        cell = sheet.cell(row=row, column=start_col)
        cell.alignment = Alignment(wrapText=True)
        # TODO create array?
        cell.value = key
        return ExcelArray(row, self.excel, style=style)

    def add_label(self, label, row):
        sheet = self.excel.sheet
        cell = sheet.cell(row=row, column=1)
        cell.value = label

    def get(self, key):
        empty_count = 0
        result = None
        sheet = self.excel.sheet
        for j in range(2, 99):
            cell = sheet.cell(row=j, column=1)
            if cell.value is None:
                # TODO empty row tolerance
                if empty_count > 1:
                    break
                empty_count += 1
            else:
                if key == cell.value:
                    result = ExcelArray(j, self.excel)
                    break
        return result
    # Skipping accessors: getitem, setitem, delitem, iter, len

    def get_expr(self, cell):
        sheet = self.excel.sheet
        m = re.match(r'([A-Z]+)(\d+)', cell)
        a = excel_to_decimal(m.group(1))
        b = sheet.cell(column=a, row=int(m.group(2)))
        return b.value

    def set(self, key, val, row, style='Comma'):
        sheet = self.excel.sheet
        cell = sheet.cell(row=row, column=1)
        cell.value = key

        cell = sheet.cell(row=row, column=2)
        cell.value = '={}'.format(val)
        cell.style = style
        if cell.style == 'Percent':
            cell.number_format = '0.00%'
        else:
            cell.number_format = '#,0.00'
        cell.font = Font(name='Calibri', size=11)


class ExcelArray:
    def __init__(self, row: int, excel: ExcelWriter, style: str='Comma'):
        self.excel = excel
        self.style = style

        self.i = 2
        self.j = row

    def append(self, val):
        sheet = self.excel.sheet
        cell = sheet.cell(row=self.j, column=self.i)
        sheet.column_dimensions[colnum_string(self.i)].width = 11
        cell.alignment = Alignment(wrapText=True)
        if (type(val) is int or type(val) is float) and val == 0:
            # Suppress zero value to empty string.
            cell.value = ''
        else:
            cell.value = val
        if self.style == 'Comma':
            if val != 0:
                cell.style = self.style
                cell.number_format = '#,0.00'
        elif self.style == 'Percent':
            cell.style = self.style
            cell.number_format = '0.00%'
        elif self.style == 'Ratio2':
            cell.number_format = '0.00'
        elif self.style == 'Ratio':
            # cell.style = style
            cell.number_format = '0.0000'
        else:
            if cell.style == "Normal":
                pass
            else:
                assert False
        cell.font = self.excel.ft
        self.i += 1

    def last(self):
        return "{}{}".format(colnum_string(total_main_col+1), self.j)

    def second_last(self):
        return "{}{}".format(colnum_string(total_main_col), self.j)

    def value(self):
        return "{}{}".format(colnum_string(self.i), self.j)

    def start(self):
        return colnum_string(next_year_offset)

    def end(self):
        return colnum_string(total_main_col)


def tokenize_string(input_string):
    """
    Tokenize a string containing operators and numbers with optional spaces.

    Args:
        input_string (str): The input string to be tokenized.

    Returns:
        list: A list of tokens.
    """
    # Removal absolute reference, locked reference: $[A-Z]$\d+
    clean_str = re.sub(r'\$', '', input_string)

    m = re.match(r'(\w+)\(([A-Z]+\d+):([A-Z]+\d+)\)$', clean_str)
    if m is not None:
        return [(None, m)]

    pattern = r'\s*(?:([\+\-\*/\(\)])|([A-Z]*\d+(?:\.\d+)?)?)?'
    # TODO The regex was generated by Perplexity
    # pattern = r'([\+\-\*\(\)])|[A-Z]+[0-9]|\d+(\.\d+)?|\$[A-Z]+\$\d+'
    # TODO Trying to support Excel $[A-Z]$\d+, absolute reference, locked reference
    #  =K6+(0.25-$G$6)/5
    # pattern = r'\s*(?:([\+\-\*\(\)])|(\$?[A-Z]*\$?\d+(?:\.\d+)?)?)?'

    # for tok, cell in re.findall(pattern, input_string):
    #     print(tok, cell)
    tokens = re.findall(pattern, clean_str)

    # Flatten the list of tuples and remove empty strings
    m = [(token, cell) for token, cell in tokens if token or cell]
    return m


class Calculator:
    @classmethod
    def precedence(cls, op):
        if op in ('+', '-'):
            return 1
        if op in ('*', '/'):
            return 2
        return 0

    @classmethod
    def apply_operation(cls, operators, values):
        operator = operators.pop()
        right = values.pop()
        left = values.pop()
        if operator == '+':
            values.append(left + right)
        elif operator == '-':
            values.append(left - right)
        elif operator == '*':
            values.append(left * right)
        elif operator == '/':
            values.append(left / right)

    @classmethod
    def evaluate(cls, tokens):
        values = []
        operators = []
        i = 0

        while i < len(tokens):
            token = tokens[i]
            if type(token) is float:
                values.append(token)
            elif token == '(':  # If the token is '(', push it to the operators stack
                operators.append(token)
            elif token == ')':  # If the token is ')', solve the entire brace
                while operators and operators[-1] != '(':
                    cls.apply_operation(operators, values)
                operators.pop()  # Remove the '('
            else:  # The token is an operator
                while operators and cls.precedence(operators[-1]) >= cls.precedence(token):
                    cls.apply_operation(operators, values)
                operators.append(token)

            i += 1

        while operators:
            cls.apply_operation(operators, values)

        return values[-1]

    def evaluate_cell(self, cell, d):
        value = None
        if re.match(r'#[A-Z]+\d+', cell):
            expression = d.get_expr(cell[1:])
            if type(expression) not in (float, int):
                value = self.evaluate_expression(expression, d)
            else:
                value = expression
        assert value is not None
        return value

    def evaluate_match(self, cell, d):
        # TODO Currently support SUM function only.
        assert len(cell.groups(0)) == 3
        assert re.match('sum', cell.group(1), re.IGNORECASE)

        start = cell.group(2)
        end = cell.group(3)
        m = re.match(r'([A-Z]+)(\d+)', start)
        a, z = excel_to_decimal(m.group(1)), m.group(2)
        b = excel_to_decimal(re.match(r'([A-Z]+)', end).group(1)) + 1
        total = 0
        for i in range(a, b):
            value = self.evaluate_cell('#' + colnum_string(i)+z, d)
            total += value
        return total

    def evaluate_expression(self, expression, d):
        string = None
        if re.match(r'=', expression):
            string = expression[1:]
        assert string is not None

        # Tokenize the input expression
        tokens = []
        for tok, cell in tokenize_string(string):
            if tok is not None and tok != '':
                tokens.append(tok)
            else:
                if type(cell) is str and re.match(r'[0-9.]', cell):
                    tokens.append(float(cell))
                elif type(cell) in (float, int):
                    assert False
                else:
                    value = None
                    if type(cell) is re.Match:
                        assert cell.group(1) == 'SUM'
                        value = self.evaluate_match(cell, d)
                    elif re.match(r'[A-Z]+\d+', cell):
                        # Cell colnum and decimal
                        value = self.evaluate_cell('#'+cell, d)

                    assert value is not None
                    if type(value) in (float, int):
                        tokens.append(float(value))
                    elif re.match(r'\d+', value):
                        tokens.append(float(value))
                    else:
                        assert False
        return self.evaluate(tokens)


def calculate(cell, d):
    calc = Calculator()
    return calc.evaluate_cell(cell, d)

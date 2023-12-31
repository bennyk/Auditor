from openpyxl import Workbook, worksheet
from functools import partial
import re
import datetime

from utils import strip, strip2, colnum_string

max_row = max_col = 99


class Spread:
    Percent_Denominator = 100

    def __init__(self, wb: Workbook, tick: str):
        self.tick = tick
        self.tabs = []
        self.income = None
        self.balance = None
        self.cashflow = None
        self.values = None
        self.estimates = None
        self.head = None

        prefix_index = 1
        self.start_prefix = prefix_index
        self.strip = partial(strip, prefix=prefix_index)
        self.strip2 = partial(strip2, prefix=prefix_index)

        for index, name in enumerate(wb.sheetnames):
            if name == 'Header':
                ws = wb[name]
                self.head = ws['A1'].value
                continue

            tab = Table(wb[name])
            self.tabs.append(tab)
            if re.match(r'Income', name):
                self.income = tab
                self.start_date = self.income.tab[0][1]
                # -2 ignore year LTM
                self.end_date = self.income.tab[0][-2]
            elif re.match(r'Balance', name):
                self.balance = tab
            elif re.match(r'Cash', name):
                self.cashflow = tab
            elif re.match(r'Values', name):
                self.values = tab
            elif re.match(r'Estimates', name):
                self.estimates = tab
                # TODO Removing two expected entry and CAGR data not available for TIKR Pro
                tab.remove(-3)
            else:
                # passing Ratios
                pass

        # 2: Skip int to str in start and end years which extract two digit, year number, in alphabets
        if type(self.end_date) is datetime.datetime:
            self.end_year = int(str(self.end_date.year)[2:])
        else:
            _ = self.end_date.split('/')[-1]
            self.end_year = int(_.strip('EA'))

        if type(self.start_date) is datetime.datetime:
            self.start_year = int(str(self.start_date.year)[2:])
        else:
            _ = self.start_date.split('/')[-1]
            self.start_year = int(_.strip('EA'))
        print("Sampled from {} to {} in {} years".format(
            self.start_date, self.end_date,
            1+self.end_year-self.start_year))

        self.half_len = int((self.end_year-self.start_year+1)/2)

    def short_name(self):
        a = re.sub(r'The', '', self.head)
        return a.split()[0]

    def share_out_filing(self) -> [float]:
        # "Total Shares Out. Filing Date" is provided in Balance Sheet which computed as fully year
        # or last trailing year I think.
        # TODO Currently it is being replace by "Weighted Average Diluted Shares Outstanding"
        x = self.balance.match_title('Total Shares Out\.')
        result = list(filter(None, reversed(x[self.start_prefix:])))[0]
        return result

    def wa_diluted_shares_out(self) -> [float]:
        shares_out = self.strip(self.income.match_title('Weighted Average Diluted Shares Outstanding'))

        # Find the first item in list that is not None, replace None to last item.
        last_item = next((x for x in shares_out if x is not None))
        result = list(map(lambda x: x if x is not None else last_item, shares_out))
        return result


class Table:
    col_limit = 0

    def __init__(self, sheet_ranges: worksheet):
        self.date_range = []
        last_limit = 0
        try:
            # configure spreadsheet based on number of cols.
            for j in range(2, max_col):
                c0 = "{}{}".format(colnum_string(j), 1)
                if type(sheet_ranges[c0].value) is datetime.datetime:
                    # type: datetime.datetime
                    _ = sheet_ranges[c0].value
                    self.date_range.append('{}/{}/{}'.format(_.month, _.day, _.year))
                else:
                    self.date_range.append(sheet_ranges[c0].value)

                if re.match(r'LTM$', self.date_range[-1]):
                    Table.col_limit = j + 1
                    break
                last_limit = j+1
        except TypeError:
            c0 = "{}{}".format(colnum_string(last_limit), 1)
            if sheet_ranges[c0].value is None:
                Table.col_limit = last_limit
            else:
                Table.col_limit = last_limit+1

        self.tab = []
        for i in range(1, sheet_ranges.max_row+1):
            r = []
            for j in range(1, Table.col_limit):
                c1 = "{}{}".format(colnum_string(j), i)
                a = sheet_ranges[c1].value
                if type(sheet_ranges[c1].value) is str:
                    if re.match(r'[-–]|(?:[,\d]+(?:(\.\d+))?)x$', a):
                        # Match multiple such as n.nnx where n is a digit and x prefix is char
                        a = a.replace(',', '')
                        a = float(a.replace('x', ''))
                r.append(a)
            self.tab.append(r)

    def match_title(self, reg, none_is_optional=False, flags=0):
        result = None
        for _ in self.tab:
            if _[0] is not None and re.match(reg, _[0].strip(), flags):
                result = _
                break
        if not none_is_optional:
            assert result is not None
        return result

    def remove(self, offset):
        self.tab = [row[:offset] for row in self.tab]

import re
import pandas as pd

from spread import Spread
from utils import *
from bcolors import colour_print, bcolors

from openpyxl import Workbook, load_workbook, worksheet
from openpyxl.styles import Font, Alignment
from collections import OrderedDict
from typing import List
from enum import IntEnum
import yfinance as yf
from tabulate import tabulate
from calculator import ExcelWriter
from datetime import datetime

total_main_col = 12
total_half_col = int(total_main_col / 2)
total_elem = 10
total_half_elem = int(total_elem / 2)
prev_year_offset = 1
start_year_offset = 2
next_year_offset = 3
half_base_offset = total_half_elem + 2


class RowIndex(IntEnum):
    """Excel row index"""
    year = 1
    sales_growth_rate = 2
    sales = 3
    ebit_margin = 4
    ebit = 5
    tax_rate = 6
    nopat = 7
    reinvestment = 8
    fcff = 9
    # 10: empty row
    cost_of_capital = 11
    cumulated_df = 12
    pv = 13
    end_of_roll_number = 14

    returns = 32
    invested_capital = 33
    roic = 34

    trade_year = 36
    trade_sales_growth_rate = 37
    trade_sales = 38
    trade_ebit = 39
    trade_interest_expense = 40
    trade_eps_proj = 41
    trade_adr_ratio = 42
    trade_adr_convert = 43
    trade_pe_ratio = 44
    trade_price_target = 45


class DataSet:
    def __init__(self, country, industry):
        self.path = "datacurrent"
        self.country = country
        self.industry = industry

    def get_country_tax_rates(self):
        name = 'countrytaxrates'
        result = None
        wb = load_workbook(self.path + '/{}.xlsx'.format(name))
        ws = wb[name]
        for i in range(1, ws.max_row):
            if ws['A{}'.format(i)].value == self.country:
                result = ws.cell(row=i, column=ws.max_column).value
                break
        assert result is not None
        return result

    def get_riskfree_rate(self):
        # TODO Country 10 years GBY on 13-Apr-2024 at the following websites: -
        # https://tradingeconomics.com/united-states/government-bond-yield
        # https://tradingeconomics.com/malaysia/government-bond-yield
        # https://tradingeconomics.com/china/government-bond-yield
        # https://tradingeconomics.com/hong-kong/government-bond-yield
        # https://tradingeconomics.com/taiwan/government-bond-yield
        tab = {
            'Malaysia': .03947,
            # 'United States': .0408,
            'United States': .04532,
            'Taiwan': .01565,
            'China': .02291,
            'Hong Kong': .0388, }
        assert self.country in tab
        return tab[self.country]

    def get_wacc(self):
        name = 'wacc'
        result = None
        wb = load_workbook(self.path + '/{}.xlsx'.format(name))
        ws = wb['Industry Averages']
        for i in range(20, ws.max_row):
            print(ws['A{}'.format(i)].value)
            if re.match(self.industry, ws['A{}'.format(i)].value, re.IGNORECASE):
                result = ws.cell(row=i, column=ws.max_column).value
                break
        assert result is not None
        return result

    def get_equity_risk_premium(self):
        result = None
        wb = load_workbook(self.path + '/{}.xlsx'.format('ERPs by country'))
        ws = wb['Sheet1']
        for i in range(8, ws.max_row):
            if re.match(self.country, ws['A{}'.format(i)].value):
                result = ws.cell(row=i, column=total_half_elem).value
                break
        assert result is not None
        return result

    def get_currency_suffix(self, sticky_price):
        assert self.country is not None
        result = None
        with open(self.path + '/yahoo_currency_suffix.txt', encoding='utf-8') as infile:
            # https://www.gnucash.org/docs/v4/C/gnucash-help/fq-spec-yahoo.html
            suffix_index = None
            for i, field in enumerate(infile.readline().split('|')):
                if re.match('suffix', field, re.IGNORECASE):
                    suffix_index = i
                    break
            assert suffix_index is not None
            for line in infile:
                a = line[:-1].split('|')
                if re.match(a[suffix_index][1:], sticky_price):
                    result = a[suffix_index][1:]
                    break
        return result

    def get_sales_to_cap_ratio(self):
        name = 'capex'
        sales_to_cap_index = None
        result = None
        wb = load_workbook(self.path + '/{}.xlsx'.format(name))
        ws = wb['Industry Averages']
        for i in range(1, ws.max_column+1):
            a = ws.cell(row=8, column=i).value
            if re.match(r'Sales/ Invested Capital', a):
                sales_to_cap_index = i
                break
        assert sales_to_cap_index is not None
        for i in range(8, ws.max_row):
            if re.match(self.industry, ws['A{}'.format(i)].value, re.IGNORECASE):
                result = ws.cell(row=i, column=sales_to_cap_index).value
                break
        assert result is not None
        return result

    def match_sales_to_cap_ratio(self, sales_to_cap_ratio):
        name = 'capex'
        sales_to_cap_index = None
        result = None
        wb = load_workbook(self.path + '/{}.xlsx'.format(name))
        ws = wb['Industry Averages']
        for i in range(1, ws.max_column + 1):
            a = ws.cell(row=8, column=i).value
            if re.match(r'Sales/ Invested Capital', a):
                sales_to_cap_index = i
                break
        assert sales_to_cap_index is not None
        result = []
        for i in range(9, ws.max_row):
            diff = ws.cell(row=i, column=sales_to_cap_index).value - sales_to_cap_ratio
            # print(ws.cell(row=i, column=1).value, diff)
            result.append([
                ws.cell(row=i, column=1).value,
                ws.cell(row=i, column=sales_to_cap_index).value, abs(diff), ])
        return sorted(result, key=lambda e: e[2])


class DCF(Spread):
    def __init__(self, tick, country=None, industry=None, path=None):
        colour_print("Company's ticker '{}'".format(tick), bcolors.UNDERLINE)

        if country is None:
            country = 'United States'
            colour_print("Defaulting country to '{}'?".format(country), bcolors.WARNING)

        if industry is None:
            industry = 'semiconductor'
            colour_print("Defaulting industry to '{}'?".format(industry), bcolors.WARNING)

        if path is None:
            path = 'spreads'
            print("Set path to '{}'".format(path))

        self.wb = load_workbook(path + '/' + tick + '.xlsx')
        super().__init__(self.wb, tick)

        self.excel = ExcelWriter(tick)
        self.dataset = DataSet(country, industry)
        self.cached_ticker = None

        # Revenues, Operating Income, Interest Expense, ...

        self.sales = self.strip(self.income.match_title('Total Revenues'))
        self.forward_sales = self.trim_estimates('Revenue')
        self.forward_ebit = self.trim_estimates('EBIT$')
        self.forward_interest = self.trim_estimates('Interest Expense')

        self.ie = self.strip(self.income.match_title('Interest Expense'))
        self.book_value_equity = self.strip(self.balance.match_title('Total Equity'))
        self.book_value_debt = self.strip(self.balance.match_title('Total Debt'))
        self.current_debt = [0.]
        current_debt_not_strip = self.balance.match_title('Current Portion of Long-Term Debt', none_is_optional=True)
        if current_debt_not_strip is not None:
            self.current_debt = self.strip(current_debt_not_strip)
        self.debt = self.strip(self.balance.match_title('Total Debt'))

        cash_not_strip = self.balance.match_title('Total Cash', none_is_optional=True)
        if cash_not_strip is not None:
            self.cash = self.strip(cash_not_strip)
        else:
            self.cash = self.strip(self.balance.match_title('Cash And Equivalents'))
            assert self.cash is not None

        investment_not_strip = self.balance.match_title('Long-term Investments', none_is_optional=True)
        if investment_not_strip is not None:
            self.investments = self.strip(investment_not_strip)
        else:
            self.investments = 0

        minority_not_strip = self.income.match_title('Minority Interest', none_is_optional=True)
        if minority_not_strip is not None:
            self.minority = self.strip(minority_not_strip)
        else:
            self.minority = 0

        self.shares = self.strip(self.income.match_title('Weighted Average Diluted Shares Outstanding'))
        forward_etr_not_strip = self.trim_estimates('Effective Tax Rate', none_is_optional=True)
        if forward_etr_not_strip is not None:
            self.forward_etr = forward_etr_not_strip
        else:
            self.forward_etr = 0

        self.marginal_tax_rate = self.dataset.get_country_tax_rates()

        # Country 10 years GBY
        self.riskfree_rate = self.dataset.get_riskfree_rate()

        # Sea of change by Howard Marks
        # https://www.oaktreecapital.com/insights/memo/sea-change
        # self.riskfree_rate = .036

    def trim_estimates(self, title, **kwargs):
        # **kwargs: Passthrough to allow none_is_optional, optional argument
        result = None
        for i, a in enumerate(self.estimates.date_range):
            if re.match(r'.*\bE$', a):
                # Assuming table sorted in 'A' (actual) to 'E' (estimates)
                tab = self.estimates.match_title(title, **kwargs)
                if tab is not None:
                    # Prune out trailing Nones
                    result = [t for t in tab[i:] if t is not None]
                break
                # assert False
            elif re.match(r'.*\bA$', a):
                pass
            else:
                assert False
        return result

    def compute(self):
        d = self.excel.create_dict()
        self.compute_revenue(d)
        self.compute_ebit(d)
        self.compute_tax(d)
        self.compute_ebt(d)
        self.compute_reinvestment(d)
        self.compute_fcff(d)
        self.compute_cost_of_capital(d)
        self.compute_cumulative_df(d)
        self.compute_terminals(d)
        self.compute_return_invested_capital(d)
        self.compute_trade(d)

        self.excel.wb.save('aaa.xlsx')

    def compute_revenue(self, d):
        # Compute past
        # Compute forward forecast

        sales_growth_rate = d.create_array('Revenue growth rate', RowIndex.sales_growth_rate, style='Percent')
        sales = d.create_array('Revenue', RowIndex.sales)
        forward_sales = self.forward_sales
        forward_end = len(self.forward_sales)+1
        for i in range(len(forward_sales)):
            if i != 0:
                grow_rate_cell = ("=({start_year}{sales_row} - {prev_year}{sales_row})"
                                  "/ {prev_year}{sales_row}").format(
                    start_year=colnum_string(i+start_year_offset),
                    prev_year=colnum_string(i+prev_year_offset),
                    sales_row=RowIndex.sales)
                sales_growth_rate.append(grow_rate_cell)
            else:
                sales_growth_rate.append(0)
            sales.append(forward_sales[i])

        # Stable at year 2, 3 (optional here after), 4 and 5
        stable_growth_rate = "={}{row}".format(
            colnum_string(forward_end), row=RowIndex.sales_growth_rate)
        current_sales = "={}{sales_row}*(1+{}{sales_grate_row})".format(
            colnum_string(forward_end),
            colnum_string(len(forward_sales)+start_year_offset),
            sales_row=RowIndex.sales, sales_grate_row=RowIndex.sales_growth_rate)
        for i in range(len(forward_sales)-1, forward_end):
            sales_growth_rate.append(stable_growth_rate)
            stable_growth_rate = "={}{row}".format(
                colnum_string(forward_end), row=RowIndex.sales_growth_rate)
            sales.append(current_sales)
            if i < forward_end:
                # Otherwise, ignore the last loop
                current_sales = "={}{sales_row}*(1+{}{sales_grate_row})".format(
                    colnum_string(len(forward_sales)+i-1),
                    colnum_string(len(forward_sales)+i),
                    sales_row=RowIndex.sales, sales_grate_row=RowIndex.sales_growth_rate)

        # Terminal year period is based on current riskfree rate based on 10 years treasury bond note yield
        term_year_per = self.riskfree_rate

        # Iterating from first year to terminal year in descending grow order, including terminal year
        for n in range(1, total_half_col):
            per = "={fwd_sales_plus1}{row} - ({fwd_sales_plus1}{row}-{term_year_per})/{half_elem} * {n}".format(
                fwd_sales_plus1=colnum_string(total_half_col+n), n=n,
                row=RowIndex.sales_growth_rate, term_year_per=term_year_per, half_elem=total_half_elem)
            sales_growth_rate.append(per)
            current_sales = "={}{sales_row}*(1+{}{sales_grate_row})".format(
                colnum_string(total_half_col + n), colnum_string(half_base_offset+n),
                sales_row=RowIndex.sales, sales_grate_row=RowIndex.sales_growth_rate)
            sales.append(current_sales)

        # Terminal period, no grow and stagnated value
        per = "={fwd_sales_plus1}{row} - ({fwd_sales_plus1}{row}-{term_year_per})/{half_elem} * {half_elem}".format(
            fwd_sales_plus1=colnum_string(total_main_col),
            row=RowIndex.sales_growth_rate, term_year_per=term_year_per, half_elem=total_half_elem)
        sales_growth_rate.append(per)
        current_sales = "={}{sales_row}*(1+{}{sales_grate_row})".format(
            colnum_string(total_main_col), colnum_string(total_main_col+1),
            sales_row=RowIndex.sales, sales_grate_row=RowIndex.sales_growth_rate)
        sales.append(current_sales)

    def compute_ebit(self, d):
        ebit_margin = d.create_array('EBIT margin', RowIndex.ebit_margin, style='Percent')
        ebit = d.create_array('EBIT', RowIndex.ebit)

        for i, e in enumerate(self.forward_ebit):
            margin_template = "={start_year}{ebit_row} / {start_year}{sales_row}".format(
                start_year=colnum_string(i+start_year_offset),
                ebit_row=RowIndex.ebit, sales_row=RowIndex.sales
            )
            ebit_margin.append(margin_template)
            ebit.append(e)

        fixed_margin = '={}{ebit_margin_row}'.format(
            colnum_string(total_half_col-1), ebit_margin_row=RowIndex.ebit_margin)
        remaining_col = total_main_col - len(self.forward_ebit)
        for i in range(1, remaining_col):
            ebit_margin.append(fixed_margin)
            stable_ebit = "={col}{ebit_margin_row}*{col}{sales_row}".format(
                col=colnum_string(i+total_half_col-1),
                ebit_margin_row=RowIndex.ebit_margin, sales_row=RowIndex.sales
            )
            ebit.append(stable_ebit)
        ebit_margin.append(fixed_margin)
        ebit_formula = "={col}{ebit_margin_row}*{col}{sales_row}".format(
            ebit_margin_row=RowIndex.ebit_margin, sales_row=RowIndex.sales,
            col=colnum_string(total_main_col + 1))
        ebit.append(ebit_formula)

    def compute_tax(self, d):
        tax_col = colnum_string(half_base_offset)
        etr = d.create_array('Tax rate', RowIndex.tax_rate, style='Percent')
        if self.forward_etr == 0:
            colour_print("Is \"{}\" a REIT company?"
                         " REIT company distributes at least 90% of its total yearly income to unit holders, the REIT itself is exempt from tax for that year, but unit holders are taxed on the distribution of income"
                         .format(self.tick), bcolors.WARNING)
            return

        assert type(self.forward_etr) is list
        for i, e in enumerate(self.forward_etr):
            if e is not None:
                etr.append(e/100)
            else:
                # TODO: Invalid ETR
                etr.append(0)

        start_tax_rate = "{}{}".format(colnum_string(i+start_year_offset), RowIndex.tax_rate)
        for i in range(len(self.forward_etr), total_half_col):
            etr.append("={}".format(start_tax_rate))

        # Previous year tax rate + (marginal tax rate - previous year tax rate) / 5
        tax_cell = "${tax_col}${tax_rate}".format(tax_col=tax_col, tax_rate=RowIndex.tax_rate)
        for i in range(1, total_half_col):
            tax_rate_cell = "={half_col}{tax_rate}+({marginal_tax_rate}-{tax_cell})/{half_elem}".format(
                half_col=colnum_string(i + total_half_col), tax_rate=RowIndex.tax_rate,
                marginal_tax_rate=self.marginal_tax_rate, tax_cell=tax_cell, half_elem=total_half_elem)
            etr.append(tax_rate_cell)
        tax_rate_cell = "={}{}".format(colnum_string(i+half_base_offset), RowIndex.tax_rate)
        etr.append(tax_rate_cell)

    def compute_ebt(self, d):
        nopat = d.create_array('NOPAT', RowIndex.nopat)
        for i in range(len(self.forward_ebit)):
            nopat_cell = ("={start_year_offset}{ebit_index}"
                          "*(1-{start_year_offset}{tax_rate_index})").format(
                start_year_offset=colnum_string(i+start_year_offset),
                ebit_index=RowIndex.ebit, tax_rate_index=RowIndex.tax_rate)
            nopat.append(nopat_cell)

        nopat_start = len(self.forward_ebit)
        for i in range(nopat_start, total_main_col):
            nopat_cell = "={start_year}{ebit_index}*(1-{start_year}{tax_rate_index})".format(
                start_year=colnum_string(i+start_year_offset),
                ebit_index=RowIndex.ebit, tax_rate_index=RowIndex.tax_rate)
            nopat.append(nopat_cell)

    def compute_reinvestment(self, d):
        # TODO Actual capex, D&A and changes in working capital
        # fwd_capex = self.strip(self.estimates.match_title('Capital Expenditure'))
        # fwd_dna = self.strip(self.estimates.match_title('Depreciation & Amortization'))
        # fwd change in working capital

        # Defaulting to no lag for the time being.
        # =IF(more_lag="No",(forward_2y_sales-forward_sales)/sales_to_cap,
        #   IF(years_lag=0,(forward_sales-prev_sales)/sales_to_cap,
        #   IF(years_lag=2,(forward_3y_sales-forward_2y_sales)/sales_to_cap,
        #   IF(years_lag=3,(forward_4y_sales-forward_3y_sales)/sales_to_cap,
        #   (forward_2y_sales-forward_sales)/sales_to_cap))))
        # https://pages.stern.nyu.edu/~adamodar/New_Home_Page/datafile/capex.html

        # Latest sales to book value of equity
        sales_to_cap_source = self.sales[-1] / self.book_value_equity[-1]
        print("Computed Sales to cap ratio {:.2f}".format(sales_to_cap_source))
        # TODO Asia countries not in U.S. coverage
        if True:
            print("Probable Sales to cap ratio:")
            matches = self.dataset.match_sales_to_cap_ratio(sales_to_cap_source)[:total_half_elem]
            heads = ['Company', 'Sales to Cap', 'Error']
            print(tabulate(matches, headers=heads, floatfmt=".2f"), "\n")
            # Selecting mid of the 5 matches
            sales_to_cap_ratio = matches[2][1]
        else:
            sales_to_cap_ratio = sales_to_cap_source

        reinvestment = d.create_array('- Reinvestment', RowIndex.reinvestment)
        reinvestment.append(None)
        elem_end = total_elem + 1
        for i in range(1, elem_end):
            r = "=({next_year}{sales}-{start_year}{sales})/{sales_to_cap_ratio}".format(
                next_year=colnum_string(i+next_year_offset), sales=RowIndex.sales,
                start_year=colnum_string(i+start_year_offset), sales_to_cap_ratio=sales_to_cap_ratio)
            reinvestment.append(r)

        # Terminal growth rate / End of ROIC * End of NOPAT
        term_growth_rate = "{}".format(d.get('Revenue growth rate').last())
        # TODO Cost of capital at year 10 or enter manually
        roic = .15
        terminal_col = total_main_col+1
        nopat_cell = '{term_col}{nopat}'.format(term_col=colnum_string(terminal_col), nopat=RowIndex.nopat)
        reinvestment.append('={term_grate}/{roic} * {nopat}'.format(
            term_grate=term_growth_rate, roic=roic, nopat=nopat_cell))

        # Alternatively Sales to IC ratio = FCFF (computed by TIKR) + NOPAT

    def compute_fcff(self, d):
        fcff = d.create_array('FCFF', RowIndex.fcff)
        fcff.append(None)
        for i in range(2, total_main_col+1):
            fcff_cell = "={prev_year}{nopat}-{prev_year}{reinvest}".format(
                nopat=RowIndex.nopat, reinvest=RowIndex.reinvestment,
                prev_year=colnum_string(i+prev_year_offset))
            fcff.append(fcff_cell)

    def compute_cost_of_capital(self, d):
        # Cost of debt
        interest_expense = 0
        if self.ie[-1] is not None:
            interest_expense = self.ie[-1]

        debt = 0
        if self.debt[-1] is not None:
            debt = self.debt[-1]
        # I have excluded tax rate leading to lower debt number.
        pretax_cost_of_debt = 0
        if debt > 0:
            pretax_cost_of_debt = abs(interest_expense / debt)
        if self.forward_etr == 0:
            cost_of_debt_formula = pretax_cost_of_debt
        else:
            cost_of_debt_formula = "{pretax_cost_of_debt}*(1-{avg_etr}/100)".format(
                pretax_cost_of_debt=pretax_cost_of_debt,
                avg_etr=average(self.forward_etr))

        # Cost of equity
        yf_ticker = self.get_ticker()
        if 'beta' in yf_ticker.info:
            beta = yf_ticker.info['beta']
            print("Obtain beta:", beta)
        else:
            colour_print("Invalid beta: defaulting beta to 1.0", bcolors.WARNING)
            beta = 1.0
        mrp = self.dataset.get_equity_risk_premium()
        cost_of_equity_formula = "({riskfree_rate}+{beta}*{mrp})".format(
            riskfree_rate=self.riskfree_rate, beta=beta, mrp=mrp)

        market_cap = yf_ticker.info['marketCap'] / 1e6
        total_cap_formula = "{market_cap}+{debt}".format(market_cap=market_cap, debt=debt)
        initial_coc_formula = ("{market_cap}/({total_cap})*{cost_of_equity}"
                               "+ {debt}/({total_cap})*{cost_of_debt}").format(
            market_cap=market_cap, debt=debt, total_cap=total_cap_formula,
            cost_of_equity=cost_of_equity_formula, cost_of_debt=cost_of_debt_formula)

        # Mature market ERP set to 4.5% and 0% based on U.S. CRP (Country risk premium)
        # https://www.youtube.com/watch?v=kyKfJ_7-mdg
        # https://pages.stern.nyu.edu/~adamodar/New_Home_Page/datafile/ctryprem.html
        crp = 0
        mature_market_erp = .045
        coc = d.create_array('Cost of capital', RowIndex.cost_of_capital, style="Percent")
        coc.append(None)
        coc.append("={}".format(initial_coc_formula))
        for j in range(1, total_half_elem):
            coc.append("={start_year}{cost_of_capital}".format(start_year=colnum_string(j + start_year_offset),
                                                               cost_of_capital=RowIndex.cost_of_capital))

        prev_coc = "${half_base_offset}${cost_of_capital}".format(
            half_base_offset=colnum_string(half_base_offset), cost_of_capital=RowIndex.cost_of_capital)
        total_erp = crp + mature_market_erp
        for i in range(1, total_half_col):
            # Prev coc - (fixed prev coc - riskfree rate + mature market risk + country risk premium)/5
            current_coc = "={prev_coc} - ({init_coc}-({riskfree_rate}+{erp}))/{half_elem}".format(
                prev_coc=prev_coc, init_coc=initial_coc_formula,
                riskfree_rate=self.riskfree_rate, erp=total_erp, half_elem=total_half_elem)
            coc.append(current_coc)
        coc.append("={riskfree_rate}+{mature_market_erp}".format(riskfree_rate=self.riskfree_rate,
                                                                 mature_market_erp=mature_market_erp))

    def compute_cumulative_df(self, d):
        cumulated_df = d.create_array('Cumulated discount factor', RowIndex.cumulated_df)
        cumulated_df.append(1.0)
        pv = d.create_array('PV (FCFF)', RowIndex.pv)
        pv.append(None)
        for i in range(total_main_col - 2):
            current_df = "={start_year}{cumulated_df}*1/(1+{next_year}{cost_of_capital})".format(
                start_year=colnum_string(i+start_year_offset),
                cumulated_df=RowIndex.cumulated_df,
                next_year=colnum_string(i+next_year_offset),
                cost_of_capital=RowIndex.cost_of_capital)
            cumulated_df.append(current_df)

            current_fcff = "={next_year}{fcff}*{next_year}{cumulated_df}".format(
                next_year=colnum_string(i+next_year_offset), fcff=RowIndex.fcff,
                cumulated_df=RowIndex.cumulated_df)
            pv.append(current_fcff)

    def compute_terminals(self, d):
        roll_number = RowIndex.end_of_roll_number

        def add_rollng_number():
            nonlocal roll_number
            roll_number += 1
            return roll_number

        d.set('Terminal cash flow', d.get('FCFF').last(), add_rollng_number())
        d.set('Terminal cost of capital', d.get('Cost of capital').last(), add_rollng_number(), style='Percent')
        d.set('Terminal value',
              "{term_cash_flow}/({term_cost_of_capital}-{last_sales_grate})".format(
                  term_cash_flow=d.get('Terminal cash flow').value(),
                  term_cost_of_capital=d.get('Terminal cost of capital').value(),
                  last_sales_grate=d.get('Revenue growth rate').last()),
              add_rollng_number())
        d.set('PV (Terminal value)',
              "{term_value}*{cumulated_df}".format(
                  term_value=d.get('Terminal value').value(),
                  cumulated_df=d.get('Cumulated discount factor').last2()),
              add_rollng_number())

        mark = 'PV (FCFF)'
        d.set('PV (Cash flow over next 10 years)', "SUM({start}{row}:{end}{row})".format(
            start=d.get(mark).start(), end=d.get(mark).end(), row=d.get(mark).j), add_rollng_number())

        d.set('Sum of PV', "{pv_term_value}+{pv_next_10_years}".format(
            pv_term_value=d.get('PV (Terminal value)').value(),
            pv_next_10_years=d.get('PV (Cash flow over next 10 years)').value()), add_rollng_number())

        d.set('Value of operating assets', "{}".format(d.get('Sum of PV').value()), add_rollng_number())

        debt = 0
        if self.debt[-1] is not None:
            debt = self.debt[-1]
        d.set('- Debt', "{}".format(debt), add_rollng_number())

        # TODO
        d.set('- Minority interest', 0, add_rollng_number())
        d.set('+ Cash', "{}".format(self.cash[-1]), add_rollng_number())

        non_op = 0
        if type(self.investments) is list:
            # type: List[float]
            if self.investments[-1] is not None:
                non_op = self.investments[-1]
            elif self.investments[-2] is not None:
                colour_print("Latest investment was not defined. Fallback to previous year", bcolors.WARNING)
                non_op = self.investments[-2]
        d.set('+ Non-operating assets', "{}".format(non_op), add_rollng_number())

        d.set('Value of equity',
              "{value_of_equity} - {debt} - {minority_interest} + {cash} + {non_op_asset}".format(
                  value_of_equity=d.get('Value of operating assets').value(),
                  debt=d.get('- Debt').value(),
                  minority_interest=d.get('- Minority interest').value(),
                  cash=d.get('+ Cash').value(),
                  non_op_asset=d.get('+ Non-operating assets').value()),
              add_rollng_number())

        d.set('Number of shares', self.shares[-1], add_rollng_number())
        d.set('Estimated value / share', "{value_of_equity}/{num_shares}".format(
            value_of_equity=d.get('Value of equity').value(),
            num_shares=d.get('Number of shares').value()), add_rollng_number())

        ticker = self.get_ticker()
        avg_price = (ticker.info['regularMarketDayLow'] + ticker.info['regularMarketDayHigh']) / 2.
        d.set('Price', avg_price, add_rollng_number())
        d.set('Price as % of value', "{price}/{value_per_share}".format(
            price=d.get('Price').value(),
            value_per_share=d.get('Estimated value / share').value()),
              add_rollng_number(), style='Percent')

    def get_ticker(self):
        if self.cached_ticker is None:
            tick_name = self.tick
            if re.match(r'\d{4}$', self.tick):
                # Regex to match currency denomination express by 4 digits such as 9618.HK
                suffix = self.dataset.get_currency_suffix(self.sticky_price)
                tick_name = tick_name + '.{}'.format(suffix)

            # Test the ticker by obtaining beta
            ticker = yf.Ticker(tick_name)
            if 'beta' not in ticker.info:
                assert len(self.head) > 0
                initial_query = ' '.join(self.head.split()[:-1])
                print("Waiting to query Yahoo Finance server with '{}'".format(initial_query))
                ticker = yf.Ticker(get_symbol(initial_query))
            self.cached_ticker = ticker
        else:
            ticker = self.cached_ticker
        return ticker

    def compute_return_invested_capital(self, d):
        ## Invested capital
        # TODO =IF(operating_lease="Yes",
        #     IF(rnd_expense_cap="Yes",
        #        book_value_equity + book_value_debt - cash + adjust_debt_outstanding + rnd_converter,
        #        book_value_equity + book_value_debt - cash + adjust_debt_outstanding),
        #     IF(rnd_expense_cap="Yes",
        #        book_value_equity + book_value_debt - cash + rnd_converter,
        #        book_value_equity + book_value_debt - cash))
        d.add_label('Return', RowIndex.returns)
        invested_capital = d.create_array('Invested Capital', RowIndex.invested_capital)
        book_value_debt = 0
        if self.book_value_debt[-1] is not None:
            book_value_debt = self.book_value_debt[-1]
        current_ic = self.book_value_equity[-1] + book_value_debt - self.cash[-1]
        prev_ic = []
        reinvestment = d.get('- Reinvestment')
        prev_ic.append(reinvestment.value())
        invested_capital.append(current_ic)
        for i in range(total_main_col-1):
            current_ic = '={start_year}{invested_cap}+{next_year}{reinvestment}'.format(
                start_year=colnum_string(i+start_year_offset),
                invested_cap=RowIndex.invested_capital,
                next_year=colnum_string(i+next_year_offset),
                reinvestment=RowIndex.reinvestment)
            invested_capital.append(current_ic)

        invested_return = d.create_array('ROIC', RowIndex.roic, style='Percent')
        invested_return.append(None)
        for i in range(1, total_main_col):
            if (i+prev_year_offset) > 1:
                invested_return.append("={start_year}{nopat} / {prev_year}{invested_cap}".format(
                    start_year=colnum_string(i+start_year_offset),
                    nopat=RowIndex.nopat,
                    prev_year=colnum_string(i+prev_year_offset),
                    invested_cap=RowIndex.invested_capital))

        # TODO end of cost of capital
        # invested_return.append(d['Cost of capital'][-1])

    def compute_trade(self, d):
        trade_year = d.create_array('={}1'.format(colnum_string(1)), RowIndex.trade_year, style='')
        sales_growth_rate = d.create_array('Revenue growth rate',
                                           RowIndex.trade_sales_growth_rate, style='Percent')
        sales = d.create_array('Revenue', RowIndex.trade_sales)
        ebit = d.create_array('EBIT', RowIndex.trade_ebit)
        interest_exp = d.create_array('Interest expense', RowIndex.trade_interest_expense)
        eps_proj = d.create_array('EPS projection', RowIndex.trade_eps_proj)
        d.set('ADR ratio', 1.0, RowIndex.trade_adr_ratio)
        adr_conv = d.create_array('EPS proj. after conversion to ADR', RowIndex.trade_adr_convert)
        pe_ratio = d.create_array('Forward P/E ratio', RowIndex.trade_pe_ratio)
        price_target = d.create_array('Price target', RowIndex.trade_price_target)
        for i in range(total_main_col):
            # trade_year.append("={year}".format(year=str(i+datetime.now().year-1)))
            trade_year.append("={}{}".format(
                colnum_string(i+start_year_offset), RowIndex.year))
            sales_growth_rate.append("={year}{sales_growth_row}".format(
                year=colnum_string(i+start_year_offset),
                sales_growth_row=RowIndex.sales_growth_rate))
            sales.append("={year}{sales}".format(
                year=colnum_string(i+start_year_offset),
                sales=RowIndex.sales))
            ebit.append("={year}{ebit}".format(
                year=colnum_string(i+start_year_offset),
                ebit=RowIndex.ebit))
            if 0 <= i < len(self.forward_interest):
                interest_exp.append("={}".format(self.forward_interest[i]))
            # TODO Fix hardcoded "$B$27"
            eps_proj.append(
                "=({year}{ebit}+{year}{interest_exp})*(1-{year}{tax_rate})/$B$27".format(
                    year=colnum_string(i+start_year_offset),
                    ebit=RowIndex.ebit, tax_rate=RowIndex.tax_rate,
                    interest_exp=RowIndex.trade_interest_expense))
            adr_conv.append(
                "={year}{eps_proj}/B{adr_ratio}".format(
                    year=colnum_string(i+start_year_offset),
                    eps_proj=RowIndex.trade_eps_proj,
                    adr_ratio=RowIndex.trade_adr_ratio, ))
            # TODO average P/E ratio?
            pe_ratio.append("={}".format(15.0))
            price_target.append("={start_year}{adr_conv}*{start_year}{pe_ratio}".format(
                adr_conv=RowIndex.trade_adr_convert, pe_ratio=RowIndex.trade_pe_ratio,
                start_year=colnum_string(i+start_year_offset)))

# Damodaran main data page
# https://pages.stern.nyu.edu/~adamodar/New_Home_Page/datacurrent.html
